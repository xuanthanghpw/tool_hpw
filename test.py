import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import requests
import json
import re
import pandas as pd
import threading
import time
from openpyxl.styles import PatternFill, Font

# Import thư viện Gemini của Google (SDK mới: google-genai)
# LƯU Ý: Thư viện cũ "google-generativeai" đã bị Google khai tử (legacy/deprecated),
# các model gemini-1.5-*, gemini-pro, gemini-2.0-* đã bị gỡ khỏi server (luôn trả 404).
# Cài SDK mới bằng lệnh: pip install --upgrade google-genai
try:
    from google import genai
    from google.genai import types as genai_types
    HAS_GEMINI = True
except ImportError:
    HAS_GEMINI = False

# Cấu hình API Smarty
API_URL = "https://us-street.api.smarty.com/street-address"
HEADERS = {
    "Referer": "https://www.smarty.com/",
    "Origin": "https://www.smarty.com",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}

# Model dự phòng cho từng provider (dùng khi không dò được / không cấu hình được)
GROQ_MODELS = ["llama-3.3-70b-versatile", "llama-3.1-8b-instant"]
OPENROUTER_MODELS = [
    "meta-llama/llama-3.1-8b-instruct:free",
    "google/gemini-2.0-flash-exp:free",
    "mistralai/mistral-7b-instruct:free",
]
GEMINI_FALLBACK_MODELS = ['gemini-flash-latest', 'gemini-2.5-flash', 'gemini-2.5-flash-lite', 'gemini-2.0-flash']


class SmartyApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Smarty API & Google Gemini Auto Check (Miễn phí)")
        self.root.geometry("820x780")

        self.excel_path = ""
        self.output_dir = ""
        self.stop_requested = False

        self.session = requests.Session()
        self.session.headers.update(HEADERS)
        self.gemini_client = None

        self.setup_gui()

    def setup_gui(self):
        # 1. Khu vực chọn file Excel
        frame_top = tk.Frame(self.root)
        frame_top.pack(pady=(15, 5), padx=10, fill="x")

        self.btn_select_excel = tk.Button(frame_top, text="1. Chọn file Excel gốc", command=self.select_excel, width=20, bg="#28a745", fg="white", font=("Arial", 10, "bold"))
        self.btn_select_excel.pack(side="left", padx=(0, 10))

        self.lbl_excel_path = tk.Label(frame_top, text="Chưa chọn file...", fg="gray")
        self.lbl_excel_path.pack(side="left")

        # 2. Khu vực cấu hình Smarty Delay
        frame_delay = tk.Frame(self.root)
        frame_delay.pack(pady=5, padx=10, fill="x")

        lbl_delay = tk.Label(frame_delay, text="Độ trễ API Smarty (giây):", font=("Arial", 10, "bold"))
        lbl_delay.pack(side="left")

        self.delay_var = tk.StringVar(value="2.0")
        self.entry_delay = tk.Entry(frame_delay, textvariable=self.delay_var, width=8, font=("Consolas", 11), justify="center")
        self.entry_delay.pack(side="left", padx=10)

        # 3. Khu vực cấu hình AI (kiểm tra tự động kết quả đáng ngờ)
        frame_ai = tk.LabelFrame(self.root, text=" Kiểm tra kết quả đáng ngờ bằng AI (Miễn phí) ", font=("Arial", 10, "bold"), fg="#1da462")
        frame_ai.pack(pady=10, padx=10, fill="x")

        self.use_ai_var = tk.BooleanVar(value=False)
        self.chk_use_ai = tk.Checkbutton(frame_ai, text="Bật tự động kiểm tra trạng thái bằng AI", variable=self.use_ai_var, font=("Arial", 10), command=self.toggle_ai_input)
        self.chk_use_ai.pack(anchor="w", padx=10, pady=5)

        # 3a. Chọn chế độ: Tự động / Thủ công
        frame_mode = tk.Frame(frame_ai)
        frame_mode.pack(anchor="w", padx=20, pady=(0, 5), fill="x")

        self.ai_mode_var = tk.StringVar(value="auto")
        self.rb_mode_auto = tk.Radiobutton(
            frame_mode, variable=self.ai_mode_var, value="auto",
            text="Tự động: hệ thống tự dò & gọi các AI miễn phí (Gemini / Groq / OpenRouter)",
            font=("Arial", 9), justify="left", command=self.toggle_ai_input
        )
        self.rb_mode_auto.pack(anchor="w")

        self.rb_mode_manual = tk.Radiobutton(
            frame_mode, variable=self.ai_mode_var, value="manual",
            text="Thủ công: Tool tạo sẵn Prompt để bạn dán vào AI ngoài (ChatGPT/Gemini web...), rồi dán JSON kết quả về",
            font=("Arial", 9), justify="left", command=self.toggle_ai_input
        )
        self.rb_mode_manual.pack(anchor="w")

        # 3b. Các API Key (đều tùy chọn - chỉ cần điền provider nào bạn có key)
        frame_ai_keys = tk.Frame(frame_ai)
        frame_ai_keys.pack(fill="x", padx=20, pady=5)

        lbl_gemini = tk.Label(frame_ai_keys, text="Gemini API Key:", width=18, anchor="w")
        lbl_gemini.grid(row=0, column=0, sticky="w", pady=2)
        self.entry_ai_key = tk.Entry(frame_ai_keys, width=45, font=("Consolas", 10), show="*")
        self.entry_ai_key.grid(row=0, column=1, padx=5, pady=2)

        lbl_groq = tk.Label(frame_ai_keys, text="Groq API Key (tuỳ chọn):", width=18, anchor="w")
        lbl_groq.grid(row=1, column=0, sticky="w", pady=2)
        self.entry_groq_key = tk.Entry(frame_ai_keys, width=45, font=("Consolas", 10), show="*")
        self.entry_groq_key.grid(row=1, column=1, padx=5, pady=2)

        lbl_openrouter = tk.Label(frame_ai_keys, text="OpenRouter API Key (tuỳ chọn):", width=18, anchor="w")
        lbl_openrouter.grid(row=2, column=0, sticky="w", pady=2)
        self.entry_openrouter_key = tk.Entry(frame_ai_keys, width=45, font=("Consolas", 10), show="*")
        self.entry_openrouter_key.grid(row=2, column=1, padx=5, pady=2)

        self.lbl_ai_note = tk.Label(
            frame_ai,
            text="* Chế độ Thủ công không cần API Key. Chế độ Tự động cần ít nhất 1 trong 3 API Key trên.",
            fg="gray", font=("Arial", 8)
        )
        self.lbl_ai_note.pack(anchor="w", padx=20, pady=(0, 5))

        # Vô hiệu hoá toàn bộ input AI cho tới khi bật checkbox
        self._set_ai_inputs_state(tk.DISABLED)

        if not HAS_GEMINI:
            lbl_gemini.config(text="Gemini (chưa cài 'google-genai'):", fg="red")

        # 4. Nút Hành động
        frame_btns = tk.Frame(self.root)
        frame_btns.pack(pady=10)

        self.btn_start = tk.Button(frame_btns, text="2. Bắt đầu xử lý & Xuất file (TXT + Excel)", command=self.start_processing_thread, width=35, bg="#0078D7", fg="white", font=("Arial", 11, "bold"), state=tk.DISABLED)
        self.btn_start.pack(side="left", padx=5)

        self.btn_stop = tk.Button(frame_btns, text="Dừng lại", command=self.stop_processing, width=10, bg="#dc3545", fg="white", font=("Arial", 11, "bold"), state=tk.DISABLED)
        self.btn_stop.pack(side="left", padx=5)

        # 5. Log hiển thị
        lbl_log = tk.Label(self.root, text="Tiến trình xử lý:", font=("Arial", 10, "bold"))
        lbl_log.pack(anchor="w", padx=10)

        self.log_text = scrolledtext.ScrolledText(self.root, font=("Consolas", 10), width=95, height=18, bg="#1E1E1E", fg="#D4D4D4")
        self.log_text.pack(pady=5, padx=10)

    def _set_ai_inputs_state(self, state):
        self.entry_ai_key.config(state=state)
        self.entry_groq_key.config(state=state)
        self.entry_openrouter_key.config(state=state)
        self.rb_mode_auto.config(state=state)
        self.rb_mode_manual.config(state=state)

    def toggle_ai_input(self):
        if self.use_ai_var.get():
            self._set_ai_inputs_state(tk.NORMAL)
        else:
            self._set_ai_inputs_state(tk.DISABLED)

    def log(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def select_excel(self):
        file_path = filedialog.askopenfilename(title="Chọn file Excel", filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_path:
            self.excel_path = file_path
            self.lbl_excel_path.config(text=self.excel_path, fg="black")
            self.btn_start.config(state=tk.NORMAL)
            self.log(f"Đã chọn file: {self.excel_path}")

    def stop_processing(self):
        self.stop_requested = True
        self.log("\n[HỆ THỐNG] Đang yêu cầu dừng tiến trình...")
        self.btn_stop.config(state=tk.DISABLED)

    def start_processing_thread(self):
        # Validate Form
        try:
            delay_val = float(self.delay_var.get())
            if delay_val < 0: raise ValueError
        except ValueError:
            messagebox.showwarning("Cảnh báo", "Độ trễ không hợp lệ!")
            return

        if self.use_ai_var.get() and self.ai_mode_var.get() == "auto":
            keys = [self.entry_ai_key.get().strip(), self.entry_groq_key.get().strip(), self.entry_openrouter_key.get().strip()]
            if not any(keys):
                messagebox.showwarning("Cảnh báo", "Chế độ Tự động cần ít nhất 1 API Key (Gemini/Groq/OpenRouter)!\nHoặc chuyển sang chế độ Thủ công (không cần Key).")
                return

        save_path = filedialog.asksaveasfilename(
            title="Lưu kết quả (Sẽ xuất cả TXT và Excel)",
            defaultextension="",
            initialfile="KetQua_Smarty",
        )
        if not save_path:
            return

        base_path = save_path.replace('.txt', '').replace('.xlsx', '')
        self.txt_output_path = base_path + ".txt"
        self.excel_output_path = base_path + ".xlsx"

        self.stop_requested = False
        self.btn_start.config(state=tk.DISABLED, text="Đang xử lý...")
        self.btn_select_excel.config(state=tk.DISABLED)
        self.btn_stop.config(state=tk.NORMAL)
        self.entry_delay.config(state=tk.DISABLED)
        self.chk_use_ai.config(state=tk.DISABLED)
        self._set_ai_inputs_state(tk.DISABLED)

        threading.Thread(target=self.process_data, daemon=True).start()

    def process_data(self):
        collected_data = []

        try:
            delay_seconds = float(self.delay_var.get())
            self.log("Đang đọc dữ liệu từ file Excel...")
            df = pd.read_excel(self.excel_path)

            address_cols = ['Shipping Address1', 'Shipping Address2', 'Shipping City', 'Shipping State', 'Shipping PostalCode', 'Shipping Country']
            ref_col = 'Shipment Reference Id'
            required_cols = address_cols + [ref_col]

            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                messagebox.showerror("Lỗi dữ liệu", f"File thiếu các cột:\n{', '.join(missing_cols)}")
                self.reset_ui()
                return

            total_rows = len(df)
            self.log(f"Bắt đầu giai đoạn 1: Gọi API Smarty ({total_rows} dòng)...")

            # GIAI ĐOẠN 1: GỌI SMARTY VÀ GHI TXT (giữ nguyên hoàn toàn logic cũ)
            with open(self.txt_output_path, 'w', encoding='utf-8') as f_out:
                for index, row in df.iterrows():
                    if self.stop_requested:
                        self.log("\n[HỆ THỐNG] Tiến trình dừng bởi người dùng.")
                        break

                    ref_id = str(row[ref_col]).strip() if pd.notna(row[ref_col]) else "NO_REF"
                    parts = [str(row[col]).strip() for col in address_cols if pd.notna(row[col]) and str(row[col]).strip() != ""]
                    input_string = " ".join(parts)

                    if not input_string:
                        continue

                    result_string = self.call_smarty_api_with_retry(input_string, 10, index+1)

                    f_out.write(f"({input_string} : {result_string} [{ref_id}])\n")
                    f_out.flush()

                    collected_data.append({
                        "Reference Id": ref_id,
                        "Chuỗi đầu vào": input_string,
                        "Chuỗi đầu ra": result_string,
                        "Trạng thái xử lý": ""
                    })

                    self.log(f"Smarty API - Dòng {index + 1}/{total_rows} -> OK")
                    time.sleep(delay_seconds)

                # LUÔN LUÔN CHÈN PROMPT BẢN MỚI VÀO CUỐI FILE TXT
                if not self.stop_requested:
                    gpt_prompt = """

=======================================================================================================================
[SYSTEM PROMPT - DÀNH CHO AI/CHATGPT/GEMINI]
Nhiệm vụ của bạn là đóng vai trò chuyên gia kiểm duyệt địa chỉ quốc tế khắt khe. Dựa vào danh sách đối chiếu bên trên (định dạng: (chuỗi gốc : chuỗi chuẩn hóa từ API [Mã đơn hàng])), hãy thực hiện:

1. QUÉT TOÀN BỘ danh sách và CHỈ LỌC RA những đơn hàng CÓ DẤU HIỆU ĐÁNG NGỜ hoặc LỖI.
2. Tiêu chí bắt lỗi (Rất khắt khe):
   - Trả về "Không tìm thấy kết quả từ Server", "Lỗi HTTP...", "Lỗi API...".
   - Trả về "Không tìm thấy delivery_line_1 hoặc last_line".
   - Chuỗi trả về bị thiếu quá nhiều thành phần trọng yếu so với chuỗi gốc (mất hẳn số nhà, tên đường khác hoàn toàn, sai khác mã ZIP nghiêm trọng).
3. ĐẦU RA BẮT BUỘC TRÌNH BÀY DƯỚI DẠNG BẢNG (Markdown Table) với các cột sau:
   | Mã đơn hàng (ID) | Chuỗi đầu vào (Gốc) | Chuỗi đầu ra (API trả về) | Lý do chi tiết không hợp lệ |

TUYỆT ĐỐI KHÔNG đưa các đơn hàng hợp lệ vào bảng này. Nếu tất cả đều hợp lệ, hãy trả lời ngắn gọn: "Tất cả địa chỉ đều hợp lệ."
=======================================================================================================================
"""
                    f_out.write(gpt_prompt)
                    f_out.flush()

            # GIAI ĐOẠN 2: KIỂM TRA BẰNG AI (Tự động nhiều nhà cung cấp, hoặc Thủ công)
            if self.use_ai_var.get() and not self.stop_requested and collected_data:
                self.log("\n===========================================")
                self.log("Bắt đầu giai đoạn 2: Kiểm tra kết quả đáng ngờ bằng AI...")

                mode = self.ai_mode_var.get()
                if mode == "manual":
                    self.log("Chế độ: THỦ CÔNG (bạn tự dán Prompt vào AI ngoài).")
                    self.run_manual_flow(collected_data)
                else:
                    self.log("Chế độ: TỰ ĐỘNG (hệ thống tự dò & gọi AI miễn phí).")
                    self.run_auto_flow(collected_data)

            elif not self.use_ai_var.get():
                self.log("\n[INFO] AI tự động đang TẮT. File TXT đã có Prompt để bạn tự check thủ công.")

            # GIAI ĐOẠN 3: LƯU FILE EXCEL (nhóm "false" lên đầu + tô màu, "true" xếp dưới)
            if collected_data:
                self.log("\nBắt đầu tạo file Excel...")
                self.export_excel(collected_data, self.excel_output_path)
                self.log(f"-> Đã lưu Excel: {self.excel_output_path}")

            if not self.stop_requested:
                self.log(f"\n[THÀNH CÔNG] Toàn bộ tiến trình hoàn tất!")
                messagebox.showinfo("Hoàn tất", f"Đã xuất thành công:\n1. {self.txt_output_path}\n2. {self.excel_output_path}")

        except Exception as e:
            self.log(f"\n[LỖI] Đã xảy ra sự cố: {e}")
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra:\n{e}")
        finally:
            self.reset_ui()

    # =====================================================================================
    # GIAI ĐOẠN 2 - CHẾ ĐỘ TỰ ĐỘNG: thử nhiều nhà cung cấp AI miễn phí, thất bại thì hỏi
    # người dùng có muốn chuyển sang chế độ thủ công không (không tự động chờ lâu).
    # =====================================================================================
    def run_auto_flow(self, collected_data):
        BATCH_SIZE = 20  # batch nhỏ hơn để giảm khả năng vượt giới hạn token/phút của các provider free
        n = len(collected_data)
        self.log(f"Tổng {n} dòng -> chia thành các batch {BATCH_SIZE} dòng/lượt gọi AI.")

        idx = 0
        while idx < n:
            if self.stop_requested:
                return
            batch = collected_data[idx: idx + BATCH_SIZE]
            batch_no = idx // BATCH_SIZE + 1
            total_batches = (n + BATCH_SIZE - 1) // BATCH_SIZE
            self.log(f"-> Đang thử gọi AI tự động cho batch {batch_no}/{total_batches} ({len(batch)} dòng)...")

            results_map = self.check_with_ai_batch_auto(batch)

            if results_map is None:
                # Không có nhà cung cấp AI nào phản hồi được cho batch này -> hỏi người dùng
                self.log("    [!] Không có nhà cung cấp AI miễn phí nào phản hồi thành công cho batch này.")
                remaining_items = collected_data[idx:]
                want_manual = self.ask_switch_to_manual(
                    f"Hệ thống không gọi được AI tự động (đã thử Gemini/Groq/OpenRouter tuỳ theo Key đã nhập).\n"
                    f"Còn {len(remaining_items)} dòng chưa được kiểm tra."
                )
                if want_manual:
                    self.log("-> Người dùng chọn chuyển sang chế độ THỦ CÔNG cho các dòng còn lại.")
                    self.run_manual_flow(remaining_items)
                else:
                    self.log("-> Người dùng chọn bỏ qua kiểm tra AI cho các dòng còn lại.")
                    for item in remaining_items:
                        item["Trạng thái xử lý"] = "unknown"
                        item["Lý do đáng ngờ"] = "Bỏ qua kiểm tra AI (không gọi được AI tự động)"
                return  # đã xử lý xong phần còn lại (thủ công hoặc bỏ qua), kết thúc vòng lặp tự động

            for item in batch:
                rid = item["Reference Id"]
                res = results_map.get(rid)
                if res is None:
                    item["Trạng thái xử lý"] = "unknown"
                    item["Lý do đáng ngờ"] = "Không nhận được kết quả từ AI cho dòng này"
                else:
                    item["Trạng thái xử lý"] = res.get("status", "unknown")
                    item["Lý do đáng ngờ"] = res.get("reason", "")

            self.log(f"-> Batch {batch_no}/{total_batches} hoàn tất.")
            idx += BATCH_SIZE
            if idx < n:
                time.sleep(1.0)  # nghỉ nhẹ giữa các batch

    def check_with_ai_batch_auto(self, items):
        """
        Cố gắng nhanh chóng gọi lần lượt các nhà cung cấp AI miễn phí đã cấu hình
        (Gemini -> Groq -> OpenRouter). Mỗi provider chỉ thử rất ngắn (không chờ
        kiểu exponential backoff dài như trước) để không bắt người dùng chờ lâu.
        Trả về dict {id: {"status", "reason"}} nếu có provider nào thành công,
        hoặc None nếu TẤT CẢ provider đã cấu hình đều thất bại / không có provider nào được cấu hình.
        """
        prompt = self._build_ai_prompt(items)
        raw = self.call_any_available_ai(prompt)
        if raw is None:
            return None
        return self._parse_ai_json_array(raw)

    def call_any_available_ai(self, prompt):
        """Thử lần lượt từng provider đã có API Key. Trả về text thô đầu tiên gọi thành công, hoặc None nếu hết cách."""
        gemini_key = self.entry_ai_key.get().strip()
        groq_key = self.entry_groq_key.get().strip()
        openrouter_key = self.entry_openrouter_key.get().strip()

        providers = []
        if gemini_key and HAS_GEMINI:
            providers.append(("Gemini", lambda: self._call_gemini_fast(prompt, gemini_key)))
        elif gemini_key and not HAS_GEMINI:
            self.log("    [!] Đã nhập Gemini Key nhưng thiếu thư viện 'google-genai' -> bỏ qua Gemini.")
        if groq_key:
            providers.append(("Groq", lambda: self._call_groq(prompt, groq_key)))
        if openrouter_key:
            providers.append(("OpenRouter", lambda: self._call_openrouter(prompt, openrouter_key)))

        if not providers:
            self.log("    [!] Chưa nhập bất kỳ API Key nào cho chế độ Tự động.")
            return None

        for name, fn in providers:
            if self.stop_requested:
                return None
            self.log(f"    -> Đang thử {name}...")
            try:
                raw = fn()
                if raw and raw.strip():
                    self.log(f"    [OK] {name} phản hồi thành công.")
                    return raw
                self.log(f"    [!] {name} trả về rỗng, thử provider kế tiếp...")
            except Exception as e:
                self.log(f"    [!] {name} lỗi: {str(e)[:120]}")
                continue

        return None

    def _call_gemini_fast(self, prompt, api_key):
        """
        Gọi Gemini nhanh: chỉ thử tối đa 2 model, mỗi model tối đa 1 lần retry (chờ 3s)
        nếu gặp 429/503, không backoff dài như bản cũ -> tránh bắt người dùng chờ lâu.
        """
        client = genai.Client(api_key=api_key)
        models = self._quick_discover_gemini_models(client) or GEMINI_FALLBACK_MODELS
        last_err = None

        for model_name in models[:2]:
            for attempt in range(2):
                if self.stop_requested:
                    return None
                try:
                    response = client.models.generate_content(
                        model=model_name,
                        contents=prompt,
                        config=genai_types.GenerateContentConfig(
                            temperature=0.0,
                            response_mime_type="application/json"
                        )
                    )
                    return (response.text or "").strip()
                except Exception as e:
                    last_err = e
                    es = str(e).lower()
                    if "404" in es or "not found" in es:
                        break  # model không tồn tại -> qua model khác ngay
                    if "429" in es or "503" in es or "overloaded" in es or "unavailable" in es or "quota" in es:
                        if attempt == 0:
                            time.sleep(3)
                            continue
                        break  # thử 1 lần rồi thôi, qua model khác
                    break  # lỗi khác (401/403/...) -> không retry, qua model khác

        if last_err:
            raise last_err
        raise RuntimeError("Gemini không khả dụng")

    def _quick_discover_gemini_models(self, client):
        """Dò nhanh model khả dụng, thất bại thì bỏ qua (không được làm chậm luồng chính)."""
        try:
            discovered = []
            for m in client.models.list():
                name = getattr(m, "name", "") or ""
                short = name.split("/")[-1] if "/" in name else name
                if not short:
                    continue
                actions = getattr(m, "supported_actions", None) or []
                if actions and "generateContent" not in actions:
                    continue
                low = short.lower()
                if any(x in low for x in ["embedding", "tts", "image", "video", "vision", "aqa", "gemma"]):
                    continue
                if "flash" not in low and "pro" not in low:
                    continue
                discovered.append(short)

            def priority(name):
                n = name.lower()
                if "flash-lite" in n:
                    return 1
                if n.endswith("-latest") and "flash" in n:
                    return 0
                if "flash" in n:
                    return 2
                if "pro" in n:
                    return 3
                return 4

            discovered = sorted(set(discovered), key=priority)
            return discovered
        except Exception:
            return []

    def _call_groq(self, prompt, api_key):
        """Groq có API tương thích OpenAI, có gói miễn phí. Timeout ngắn để không bắt người dùng chờ lâu."""
        url = "https://api.groq.com/openai/v1/chat/completions"
        headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
        last_err = None
        for model_name in GROQ_MODELS[:2]:
            body = {
                "model": model_name,
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0
            }
            try:
                r = self.session.post(url, headers=headers, json=body, timeout=20)
                if r.status_code == 200:
                    data = r.json()
                    return data["choices"][0]["message"]["content"]
                elif r.status_code == 404:
                    last_err = RuntimeError(f"Model '{model_name}' không tồn tại trên Groq")
                    continue
                elif r.status_code == 429:
                    last_err = RuntimeError("Groq quá tải/hết quota (429)")
                    time.sleep(2)
                    continue
                else:
                    last_err = RuntimeError(f"Groq lỗi HTTP {r.status_code}: {r.text[:150]}")
                    continue
            except requests.exceptions.RequestException as e:
                last_err = e
                continue
        raise last_err or RuntimeError("Groq thất bại")

    def _call_openrouter(self, prompt, api_key):
        """OpenRouter có nhiều model gắn nhãn ':free' dùng được miễn phí."""
        url = "https://openrouter.ai/api/v1/chat/completions"
        headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
        last_err = None
        for model_name in OPENROUTER_MODELS:
            body = {
                "model": model_name,
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0
            }
            try:
                r = self.session.post(url, headers=headers, json=body, timeout=20)
                if r.status_code == 200:
                    data = r.json()
                    return data["choices"][0]["message"]["content"]
                elif r.status_code == 404:
                    last_err = RuntimeError(f"Model '{model_name}' không tồn tại trên OpenRouter")
                    continue
                elif r.status_code == 429:
                    last_err = RuntimeError("OpenRouter quá tải/hết quota (429)")
                    time.sleep(2)
                    continue
                else:
                    last_err = RuntimeError(f"OpenRouter lỗi HTTP {r.status_code}: {r.text[:150]}")
                    continue
            except requests.exceptions.RequestException as e:
                last_err = e
                continue
        raise last_err or RuntimeError("OpenRouter thất bại")

    # =====================================================================================
    # GIAI ĐOẠN 2 - CHẾ ĐỘ THỦ CÔNG: tạo Prompt cho người dùng tự dán vào AI ngoài,
    # rồi dán JSON kết quả trả về vào tool.
    # =====================================================================================
    def run_manual_flow(self, items):
        if not items:
            return
        self.log(f"[THỦ CÔNG] Đang tạo Prompt cho {len(items)} dòng để bạn sao chép...")
        prompt = self._build_ai_prompt(items)

        json_text = self.get_manual_json_from_user(prompt, len(items))

        if not json_text:
            self.log("    [!] Bạn đã bỏ qua bước nhập JSON thủ công -> các dòng này sẽ ở trạng thái 'unknown'.")
            for item in items:
                item["Trạng thái xử lý"] = item.get("Trạng thái xử lý") or "unknown"
                item["Lý do đáng ngờ"] = item.get("Lý do đáng ngờ") or "Bỏ qua kiểm tra AI thủ công"
            return

        results_map = self._parse_ai_json_array(json_text)
        if not results_map:
            self.log("    [!] Không đọc được JSON hợp lệ từ nội dung bạn dán vào -> các dòng này sẽ ở trạng thái 'unknown'.")

        for item in items:
            rid = item["Reference Id"]
            res = results_map.get(rid)
            if res is None:
                item["Trạng thái xử lý"] = "unknown"
                item["Lý do đáng ngờ"] = "Không thấy id này trong JSON đã nhập"
            else:
                item["Trạng thái xử lý"] = res.get("status", "unknown")
                item["Lý do đáng ngờ"] = res.get("reason", "")

        self.log(f"    -> Đã áp dụng kết quả thủ công cho {len(items)} dòng.")

    def ask_switch_to_manual(self, reason_text):
        """Hỏi người dùng (dialog chạy trên main thread) có muốn chuyển sang chế độ thủ công không.
        Được gọi từ luồng nền (background thread) nên phải đồng bộ qua threading.Event."""
        result_holder = {"value": False}
        event = threading.Event()

        def _build():
            try:
                ans = messagebox.askyesno(
                    "Không gọi được AI tự động",
                    f"{reason_text}\n\nBạn có muốn chuyển sang chế độ THỦ CÔNG "
                    f"(tự dán Prompt vào AI ngoài, rồi dán JSON kết quả về) không?"
                )
                result_holder["value"] = bool(ans)
            finally:
                event.set()

        self.root.after(0, _build)
        event.wait()
        return result_holder["value"]

    def get_manual_json_from_user(self, prompt_text, item_count):
        """Mở cửa sổ cho người dùng: sao chép Prompt, dán JSON kết quả. Trả về chuỗi JSON (hoặc None nếu bỏ qua).
        Được gọi từ luồng nền nên phải đồng bộ qua threading.Event."""
        result_holder = {"value": None}
        event = threading.Event()

        def _build():
            dlg = tk.Toplevel(self.root)
            dlg.title("Chế độ thủ công - Dán Prompt vào AI ngoài")
            dlg.geometry("720x650")
            dlg.grab_set()

            tk.Label(
                dlg,
                text=(f"Có {item_count} dòng cần kiểm tra.\n"
                      "1) Bấm 'Sao chép Prompt' rồi dán vào 1 AI bất kỳ (ChatGPT, Gemini web, Claude...).\n"
                      "2) Sao chép TOÀN BỘ phần JSON mà AI trả về, dán vào ô phía dưới.\n"
                      "3) Bấm 'Xác nhận' để tool tự tạo Excel."),
                justify="left", anchor="w"
            ).pack(anchor="w", padx=10, pady=(10, 5))

            tk.Label(dlg, text="Prompt để dán vào AI ngoài:", font=("Arial", 9, "bold")).pack(anchor="w", padx=10)
            txt_prompt = scrolledtext.ScrolledText(dlg, height=14, wrap="word", font=("Consolas", 9))
            txt_prompt.insert("1.0", prompt_text)
            txt_prompt.pack(fill="both", expand=True, padx=10, pady=(0, 5))

            def copy_prompt():
                dlg.clipboard_clear()
                dlg.clipboard_append(prompt_text)
                lbl_copied.config(text="Đã sao chép vào Clipboard!", fg="#28a745")

            frame_copy = tk.Frame(dlg)
            frame_copy.pack(pady=2)
            tk.Button(frame_copy, text="Sao chép Prompt", command=copy_prompt, bg="#0078D7", fg="white", width=20).pack(side="left", padx=5)
            lbl_copied = tk.Label(frame_copy, text="", font=("Arial", 9))
            lbl_copied.pack(side="left", padx=5)

            tk.Label(dlg, text="Dán JSON kết quả từ AI vào đây:", font=("Arial", 9, "bold")).pack(anchor="w", padx=10, pady=(10, 0))
            txt_json = scrolledtext.ScrolledText(dlg, height=10, wrap="word", font=("Consolas", 9))
            txt_json.pack(fill="both", expand=True, padx=10, pady=(0, 5))

            def on_confirm():
                content = txt_json.get("1.0", "end").strip()
                result_holder["value"] = content if content else None
                dlg.grab_release()
                dlg.destroy()
                event.set()

            def on_skip():
                result_holder["value"] = None
                dlg.grab_release()
                dlg.destroy()
                event.set()

            frame_btn = tk.Frame(dlg)
            frame_btn.pack(pady=10)
            tk.Button(frame_btn, text="Xác nhận", command=on_confirm, bg="#28a745", fg="white", width=14, font=("Arial", 10, "bold")).pack(side="left", padx=5)
            tk.Button(frame_btn, text="Bỏ qua kiểm tra AI cho các dòng này", command=on_skip, bg="#dc3545", fg="white", width=32).pack(side="left", padx=5)

            dlg.protocol("WM_DELETE_WINDOW", on_skip)

        self.root.after(0, _build)
        event.wait()
        return result_holder["value"]

    def _build_ai_prompt(self, items):
        """Dựng Prompt dùng chung cho cả 2 chế độ (Tự động gọi API / Thủ công dán vào AI ngoài)."""
        payload = [
            {"id": it["Reference Id"], "in": it["Chuỗi đầu vào"], "out": it["Chuỗi đầu ra"]}
            for it in items
        ]

        prompt = f"""Bạn là chuyên gia kiểm duyệt địa chỉ vận chuyển quốc tế, làm việc rất khắt khe và chính xác.
Dưới đây là một danh sách JSON gồm nhiều đơn hàng, mỗi phần tử có "id" (mã đơn hàng), "in" (chuỗi địa chỉ gốc), "out" (chuỗi địa chỉ đã chuẩn hóa từ API).

Với MỖI phần tử trong danh sách, hãy đánh giá:
- status = "true" NẾU: "out" là địa chỉ chuẩn hóa hợp lệ, giữ đủ thông tin quan trọng (số nhà, tên đường, thành phố, mã ZIP) khớp với "in".
- status = "false" NẾU: "out" chứa các cụm như "Lỗi", "Không tìm thấy kết quả", "Không tìm thấy delivery_line_1", hoặc "out" thiếu/sai lệch nghiêm trọng so với "in" (mất số nhà, tên đường khác hẳn, sai ZIP nghiêm trọng).

YÊU CẦU ĐẦU RA (BẮT BUỘC):
- CHỈ trả về DUY NHẤT một mảng JSON hợp lệ, KHÔNG kèm markdown code fence, KHÔNG giải thích thêm, KHÔNG viết bất kỳ chữ nào trước hoặc sau mảng JSON.
- Số phần tử trong mảng trả về PHẢI khớp chính xác với số phần tử đầu vào (mỗi "id" xuất hiện đúng 1 lần).
- Định dạng mỗi phần tử: {{"id": "<id gốc>", "status": "true" hoặc "false", "reason": "<lý do ngắn gọn, CHỈ điền khi status là false, để trống '' khi true>"}}

Danh sách đầu vào (JSON):
{json.dumps(payload, ensure_ascii=False)}
"""
        return prompt

    def _parse_ai_json_array(self, raw_text):
        """Parse JSON trả về từ AI thành dict {id: {status, reason}}, chịu lỗi tốt (bỏ markdown fence nếu có)."""
        text = raw_text.strip()
        # Phòng trường hợp model vẫn kèm ```json ... ``` dù đã ép response_mime_type
        text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text.strip())
        try:
            data = json.loads(text)
        except json.JSONDecodeError:
            # Cố gắng trích phần mảng JSON [...] nằm trong text nếu model in thêm chữ thừa
            match = re.search(r"\[.*\]", text, re.DOTALL)
            if not match:
                self.log("    [!] Không parse được JSON trả về từ AI.")
                return {}
            try:
                data = json.loads(match.group(0))
            except json.JSONDecodeError:
                self.log("    [!] JSON trả về từ AI bị sai định dạng.")
                return {}

        result_map = {}
        if isinstance(data, list):
            for entry in data:
                if not isinstance(entry, dict) or "id" not in entry:
                    continue
                status = str(entry.get("status", "")).strip().lower()
                status = "true" if status == "true" else ("false" if status == "false" else "unknown")
                result_map[str(entry["id"])] = {
                    "status": status,
                    "reason": entry.get("reason", "") or ""
                }
        return result_map

    def export_excel(self, collected_data, output_path):
        """
        Xuất Excel với TOÀN BỘ các dòng (không chỉ dòng đáng ngờ):
        - Nhóm "false" (đáng ngờ) xếp lên đầu bảng, tô màu đỏ nhạt để dễ nhận diện.
        - Nhóm "unknown" (AI không xác định được / lỗi khi check) xếp giữa, tô màu vàng.
        - Nhóm "true" (hợp lệ) xếp cuối bảng, tô màu xanh nhạt.
        """
        # Đảm bảo mọi dòng đều có đủ cột kể cả khi không bật AI hoặc AI lỗi
        for item in collected_data:
            item.setdefault("Trạng thái xử lý", "")
            item.setdefault("Lý do đáng ngờ", "")

        status_priority = {"false": 0, "unknown": 1, "": 1, "true": 2}
        sorted_data = sorted(
            collected_data,
            key=lambda x: status_priority.get(str(x.get("Trạng thái xử lý", "")).lower(), 1)
        )

        columns = ["Reference Id", "Chuỗi đầu vào", "Chuỗi đầu ra", "Trạng thái xử lý", "Lý do đáng ngờ"]
        out_df = pd.DataFrame(sorted_data, columns=columns)
        out_df.to_excel(output_path, index=False)

        # Tô màu bằng openpyxl (mở lại file vừa lưu để style theo từng dòng)
        import openpyxl
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        fill_false = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # đỏ nhạt
        font_false = Font(color="9C0006", bold=True)
        fill_unknown = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # vàng nhạt
        font_unknown = Font(color="9C6500")
        fill_true = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")     # xanh nhạt
        font_true = Font(color="006100")

        status_col_idx = columns.index("Trạng thái xử lý") + 1  # openpyxl 1-based

        for row_idx in range(2, ws.max_row + 1):  # bỏ qua header ở row 1
            status_val = str(ws.cell(row=row_idx, column=status_col_idx).value or "").strip().lower()
            if status_val == "false":
                fill, font = fill_false, font_false
            elif status_val == "true":
                fill, font = fill_true, font_true
            else:
                fill, font = fill_unknown, font_unknown

            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill
                if col_idx == status_col_idx:
                    cell.font = font

        # Tự động giãn độ rộng cột cho dễ đọc
        for col_idx, col_name in enumerate(columns, start=1):
            max_len = max(
                [len(str(col_name))] + [len(str(row.get(col_name, ""))) for row in sorted_data]
            )
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 60)

        ws.freeze_panes = "A2"
        wb.save(output_path)

    # --- Các hàm Smarty API giữ nguyên ---
    def call_smarty_api_with_retry(self, street_input, max_retries, row_num):
        retries = 0; wait_time = 2
        while retries < max_retries:
            if self.stop_requested: return "Bị dừng"
            result = self.call_smarty_api(street_input)
            if "Too many requests" in result:
                retries += 1
                self.log(f"  [!] Dòng {row_num}: Smarty Too many requests. Chờ {wait_time}s...")
                time.sleep(wait_time)
                wait_time *= 2
            else:
                return result
        return "Lỗi: Quá nhiều request"

    def call_smarty_api(self, street_input):
        params = {
            "key": "21102174564513388", "agent": "smarty (website:demo)",
            "match": "enhanced", "candidates": "5", "geocode": "true",
            "license": "us-rooftop-geocoding-cloud", "street": street_input
        }
        try:
            res = self.session.get(API_URL, params=params, timeout=15)
            if res.status_code != 200:
                try:
                    return f"Lỗi API: {res.json()['errors'][0].get('message', 'Unknown Error')}"
                except: return f"Lỗi HTTP {res.status_code}"

            data = res.json()
            if isinstance(data, list) and len(data) > 0:
                c = data[0]
                dl = c.get("delivery_line_1", "")
                ll = c.get("last_line", "")
                return f"{dl} {ll}".strip() if dl or ll else "Không tìm thấy delivery_line_1 hoặc last_line"
            return "Không tìm thấy kết quả từ Server"
        except requests.exceptions.RequestException: return "Lỗi mạng / Timeout"
        except json.JSONDecodeError: return "Lỗi phân tích JSON"

    def reset_ui(self):
        self.btn_start.config(state=tk.NORMAL, text="2. Bắt đầu xử lý & Xuất file (TXT + Excel)")
        self.btn_select_excel.config(state=tk.NORMAL)
        self.btn_stop.config(state=tk.DISABLED)
        self.entry_delay.config(state=tk.NORMAL)
        self.chk_use_ai.config(state=tk.NORMAL)
        if self.use_ai_var.get():
            self._set_ai_inputs_state(tk.NORMAL)


if __name__ == "__main__":
    root = tk.Tk()
    app = SmartyApp(root)
    root.mainloop()