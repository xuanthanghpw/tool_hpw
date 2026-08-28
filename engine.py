# -*- coding: utf-8 -*-
"""
engine.py
---------
Bien file debit.xlsx (goc) + ct.xlsx (file cong thuc/dinh dang) thanh file
debit_editted.xlsx hoan chinh.

Thiet ke:
- Doc du lieu Summary/All theo TEN HEADER (khong gia dinh cung vi tri dong/cot),
  cho phep dong trong xen giua, nhan dien dong Total an toan.
- Dinh dang (font/fill/border/alignment/number_format) duoc COPY THAT tu cac o
  mau trong sheet 'DEBIT' cua ct.xlsx (khong dung 1 Border cung cho tat ca).
  Chi khi mot o thuc su khong co border nao (rong hoan toan) trong ca ct.xlsx
  lan du lieu thuc te thi moi fallback ve border thin, de tranh mat border so
  voi file goc.

Cach dung (dong lenh):
    python engine.py "duong_dan\\debit.xlsx" "duong_dan\\ct.xlsx" "duong_dan\\output.xlsx" "CI0792"

Cach dung (tu code / GUI) - GIU NGUYEN signature:
    from engine import build_debit_editted
    build_debit_editted(debit_path, ct_path, output_path, ci_code="CI0792", progress=None)
"""
import re
import copy
import datetime as dt
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

THIN = Side(style="thin")
BORDER_THIN_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_NONE = Border()

MONTH_ABBR = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG",
              "SEP", "OCT", "NOV", "DEC"]

ALL_HEADERS = [
    "NO.", "REFERENCE NUMBER", "CLIENT ID", "CARTON NO", "TRACKING NUMBER",
    "GROSS WEIGHT - ACTUAL (KG)", "GROSS WEIGHT - BOOK (KG)",
    "DIM - ACTUAL (CM)", None, None, "DIM - BOOK (CM)", None, None,
    "VOLUME WEIGHT (KG)", "CHARGE WEIGHT (KG)", "DEBIT", "OVERSIZE FREE",
    "TAXES US", "TOTAL (USD)", "ESTIMATE\nLM COST", "ESTIMATE\nIB COST",
    "ESTIMATE\nGORI COST", "ESTIMATE\nCLUTCH COST", "TOTAL COST", "ZIPCODE",
    "WAYBILL NUMBER", "ESTIMATE COST (USD)", "VALUE", "Estimated Zone",
]
N_ALL_COLS = len(ALL_HEADERS)  # 29 -> A..AC

SUMMARY_HEADERS = ["NO", "CLIENT ID", "PAX", "TOTAL TRACK", "TOTAL (USD)",
                    "TOTAL ESTIMATE LM COST", "TOTAL ESTIMATE IB COST",
                    "TOTAL ESTIMATE GORI COST", "TOTAL ESTIMATE CLUTCH COST",
                    "NOTE"]

CURRENCY_NUMBER_FORMAT = "#,##0.00"

# columns (1-based) in the rebuilt 'All' sheet that hold monetary (USD) values
CURRENCY_COLS_ALL = {16, 17, 18, 19, 20, 21, 22, 23, 24, 27, 28}
# P=DEBIT, Q=OVERSIZE FREE, R=TAXES US, S=TOTAL (USD), T..W=ESTIMATE ... COST,
# X=TOTAL COST, AA=ESTIMATE COST (USD), AB=VALUE

# columns (1-based) in the rebuilt 'Summary' sheet that hold monetary values
CURRENCY_COLS_SUMMARY = {5, 6, 7, 8, 9}
# E=TOTAL (USD), F/G/H/I = TOTAL ESTIMATE LM/IB/GORI/CLUTCH COST

# In sheet 'All', these columns already carry a per-client SUBTOTAL at the
# bottom of every client block (see write_group_header/subtotal_row below).
# The grand-total row must NOT re-sum the raw data range for these columns
# (that would double count, since the range already contains one aggregate
# number per client); instead it references ONLY those already-computed
# per-client subtotal cells.
GRAND_TOTAL_NESTED_COLS = [1, 6, 7, 16, 17, 18, 19, 20, 21, 22, 23]
# A=NO. (per-client COUNTA), F/G=GROSS WEIGHT, P..W = DEBIT..ESTIMATE CLUTCH COST
#
# By explicit request, the grand-total ("TONG TAT CA") row must total EXACTLY
# these 11 columns and no others:
#   NO., GROSS WEIGHT ACTUAL, GROSS WEIGHT BOOK, DEBIT, OVERSIZE FREE,
#   TAXES US, TOTAL (USD), ESTIMATE LM/IB/GORI/CLUTCH COST
# All other columns on the grand-total row (DIM actual/book, volume/charge
# weight, TOTAL COST, ESTIMATE COST (USD), VALUE, ...) are intentionally
# left blank -- do NOT add totals for them here.

# columns that are text/identifier fields; no total makes sense for them.
# (2 REFERENCE NUMBER, 3 CLIENT ID, 4 CARTON NO, 5 TRACKING NUMBER,
#  25 ZIPCODE, 26 WAYBILL NUMBER, 29 Estimated Zone)

CURRENCY_HEADER_KEYWORDS = [
    "usd", "cost", "fee", "surcharge", "tax", "insurance", "ioss",
    "extra", "debit", "value", "unit price", "vat", "oversize",
]

# ===========================================================================
# 0. small helpers
# ===========================================================================


def normalize_header(value):
    """Normalize a header cell value for robust, tolerant comparison:
    lowercase, collapse whitespace/newlines, strip."""
    if value is None:
        return ""
    s = str(value).replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def _fmt_excel_text(fmt_code, value):
    """Very small TEXT() emulator for the '00' pattern used in ct.xlsx."""
    if fmt_code.strip() == "00":
        try:
            return f"{int(value):02d}"
        except (TypeError, ValueError):
            return str(value)
    return str(value)


def get_cell_value_with_fallback(ws_data_only, ws_formula, row, col, warnings=None):
    """Return the value of a cell, preferring the cached (data_only) value.
    If the cache is missing (common when a file was saved by a tool that did
    not recalc formulas), fall back to the literal value from the formula
    workbook -- but only if that cell is NOT itself a live formula (since we
    have no formula engine to evaluate it here). If it IS an unresolved
    formula, we return None and record a warning instead of silently
    treating the row as 'no data'.
    """
    v = ws_data_only.cell(row, col).value
    if v is not None:
        return v
    fcell = ws_formula.cell(row, col)
    fv = fcell.value
    if fv is None:
        return None
    if isinstance(fv, str) and fv.startswith("="):
        if warnings is not None:
            warnings.append(
                f"{ws_formula.title}!{get_column_letter(col)}{row}: cong thuc "
                f"chua co gia tri tinh san (cached value), khong doc duoc."
            )
        return None
    return fv


def row_is_blank(ws_data_only, ws_formula, row, cols, warnings=None):
    for c in cols:
        if get_cell_value_with_fallback(ws_data_only, ws_formula, row, c, warnings) not in (None, ""):
            return False
    return True


def find_header_row(ws, alias_map, required_keys, search_rows=40, min_matches=None):
    """Scan the first `search_rows` rows of `ws` looking for the row that best
    matches the expected headers described in `alias_map` (field -> list of
    normalized alias strings). Returns (header_row_index, {field: col_index}).
    Raises ValueError with a descriptive message if nothing good enough is found.
    """
    if min_matches is None:
        min_matches = max(3, len(required_keys) // 2)

    best_row, best_map, best_score = None, {}, -1
    max_col = min(ws.max_column, 80)
    max_row = min(ws.max_row, search_rows)

    for r in range(1, max_row + 1):
        row_values = {}
        for c in range(1, max_col + 1):
            nv = normalize_header(ws.cell(r, c).value)
            if nv:
                row_values.setdefault(nv, c)
        field_map = {}
        for field, aliases in alias_map.items():
            for alias in aliases:
                if alias in row_values:
                    field_map[field] = row_values[alias]
                    break
        score = len(field_map)
        if score > best_score:
            best_score, best_row, best_map = score, r, field_map

    missing_required = [k for k in required_keys if k not in best_map]
    if best_row is None or missing_required:
        preview_row = best_row or 1
        preview = [ws.cell(preview_row, c).value for c in range(1, min(ws.max_column, 20) + 1)]
        raise ValueError(
            f"Khong tim thay header phu hop trong sheet '{ws.title}'.\n"
            f"Cac truong bat buoc con thieu: {missing_required}.\n"
            f"Dong co diem khop cao nhat la dong {preview_row}, noi dung: {preview}"
        )
    return best_row, best_map


def copy_cell_style(source_cell, target_cell, border_fallback=True,
                     skip_cols=None, this_col=None):
    """Copy font/fill/border/alignment/number_format/protection from a real
    template cell onto the target cell using openpyxl's copy(). If the
    template cell has NO border at all on any side, and this column is not
    in `skip_cols` (columns that are intentionally borderless by design),
    fall back to a thin border on all sides so we never lose the boxed-table
    look of the original file.
    """
    target_cell.font = copy.copy(source_cell.font)
    target_cell.fill = copy.copy(source_cell.fill)
    target_cell.alignment = copy.copy(source_cell.alignment)
    target_cell.protection = copy.copy(source_cell.protection)
    target_cell.number_format = source_cell.number_format

    border = source_cell.border
    has_any_border = any([
        border.left and border.left.style,
        border.right and border.right.style,
        border.top and border.top.style,
        border.bottom and border.bottom.style,
    ])
    if has_any_border:
        target_cell.border = copy.copy(border)
    elif border_fallback and (skip_cols is None or this_col not in skip_cols):
        target_cell.border = copy.copy(BORDER_THIN_ALL)
    else:
        target_cell.border = copy.copy(border)


def get_style_row(ws, row, ncols):
    """Return a list of the ncols cell objects (style source) for `row`."""
    return [ws.cell(row, c) for c in range(1, ncols + 1)]


def _set_font_size_bold(cell, size=None, bold=None):
    """Override only font size / bold on a cell, keeping every other font
    attribute (name, color, italic, underline...) exactly as it already is
    (i.e. as copied from the ct.xlsx template by copy_cell_style). Used to
    enforce Summary sheet's explicit sizing rules without touching any of
    the workbook's other styling/logic."""
    f = cell.font
    cell.font = Font(
        name=f.name, size=(size if size is not None else f.size),
        bold=(bold if bold is not None else f.bold),
        italic=f.italic, vertAlign=f.vertAlign, underline=f.underline,
        strike=f.strike, color=f.color,
    )


def is_currency_header(value):
    """True if a header cell's text looks like a monetary/USD column, based
    on a keyword match (handles both the 'All'/'Summary' style headers we
    write ourselves, e.g. 'TOTAL (USD)', and the raw per-client sheet style
    headers, e.g. 'Oversize Fee', 'Import Taxes US', 'Unit price')."""
    nv = normalize_header(value)
    if not nv:
        return False
    return any(kw in nv for kw in CURRENCY_HEADER_KEYWORDS)


def find_any_header_row(ws, search_rows=10, max_col_scan=100):
    """Lightweight, tolerant header-row detector for arbitrary pass-through
    sheets (e.g. the per-client raw sheets '143_002', 'FNKH', 'IN', ...)
    whose exact column layout we do not control/rebuild. A row qualifies as
    a header row if it contains BOTH a 'no.'-like cell and a 'client id'-like
    cell anywhere in the first `search_rows` rows. Returns the row index, or
    None if nothing matches (e.g. an empty/unrelated sheet)."""
    max_r = min(ws.max_row, search_rows)
    max_c = min(ws.max_column, max_col_scan)
    for r in range(1, max_r + 1):
        found_no = False
        found_cid = False
        for c in range(1, max_c + 1):
            nv = normalize_header(ws.cell(r, c).value)
            if nv in ("no.", "no", "stt"):
                found_no = True
            elif nv == "client id":
                found_cid = True
            if found_no and found_cid:
                return r
    return None


def apply_currency_format_to_sheet(ws, log=None, max_col_scan=100):
    """Generic pass: locate the header row of `ws` (if any), find every
    column whose header looks like a monetary/USD field, and force
    '#,##0.00' number formatting on every cell of that column below the
    header (this safely covers repeated title/header/subtotal blocks too,
    since formatting a text or blank cell has no visible effect). Used for
    sheets that are copied through unchanged (per-client detail sheets,
    'FNKH', 'IN', ...) where we don't control the exact layout. Returns the
    number of currency columns formatted."""
    header_row = find_any_header_row(ws)
    if header_row is None:
        return 0
    max_c = min(ws.max_column, max_col_scan)
    currency_cols = [c for c in range(1, max_c + 1)
                      if is_currency_header(ws.cell(header_row, c).value)]
    if not currency_cols:
        return 0
    for r in range(header_row + 1, ws.max_row + 1):
        for c in currency_cols:
            ws.cell(r, c).number_format = CURRENCY_NUMBER_FORMAT
    if log:
        log(f"  - Sheet '{ws.title}': da format {len(currency_cols)} cot tien te "
            f"(cot {[get_column_letter(c) for c in currency_cols]}).")
    return len(currency_cols)


# ===========================================================================
# 1. header alias tables
# ===========================================================================

SUMMARY_ALIASES = {
    "client_id": ["client id", "ma khach", "client code"],
    "service": ["service"],
    "date": ["date", "ngay"],
    "manifest": ["manifest/mawb", "manifest / mawb", "manifest", "mawb"],
    "total_track": ["total track"],
    "total_usd": ["total (usd)", "total usd", "total(usd)"],
}
SUMMARY_REQUIRED = ["client_id", "total_track", "total_usd"]

ALL_ALIASES = {
    "no": ["no.", "no", "stt"],
    "reference_number": ["reference number"],
    "client_id": ["client id"],
    "overpack_id": ["overpack id", "carton no"],
    "tracking_number": ["tracking number"],
    "waybill_number": ["waybill number"],
    "country_code": ["country code"],
    "gross_actual": ["gross weight - actual (kg)", "gross weight actual (kg)", "gross weight - actual"],
    "gross_book": ["gross weight - book (kg)", "gross weight book (kg)", "gross weight - book"],
    "dim_actual": ["dim - actual (cm)", "dim actual (cm)"],
    "dim_book": ["dim - book (cm)", "dim book (cm)"],
    "vol_weight": ["volume weight (kg)", "volume weight"],
    "charge_weight": ["charge weight (kg)", "charge weight"],
    "amount": ["amount (usd)", "amount", "debit"],
    "oversize": ["oversize"],
    "oversize_fee": ["oversize free", "oversize fee"],
    "insurance": ["insurance"],
    "ioss": ["ioss"],
    "import_taxes": ["import taxes"],
    "peak_season": ["peak season surcharges", "peak season surcharge"],
    "vat_eu": ["vat eu 037", "vat eu037"],
    "import_taxes_us": ["taxes us", "import taxes us"],
    "eu_extra": ["eu extra"],
    "total_usd": ["total (usd)", "total usd"],
    "cost": ["cost"],
    "zone_ib": ["zone ib"],
    "zip_code": ["zip code", "zipcode"],
    "est_weight": ["estimated weight (kg)", "estimated weight"],
    "est_zone": ["estimated zone"],
    "est_cost": ["estimated cost (usd)", "estimate cost (usd)"],
    "decl_qty": ["declared quantity"],
    "unit_price": ["unit price"],
    "hpw_id": ["hpw id"],
}
ALL_REQUIRED = ["no", "client_id", "tracking_number", "waybill_number",
                "gross_actual", "gross_book", "dim_actual", "dim_book",
                "vol_weight", "charge_weight", "amount", "total_usd",
                "zip_code", "est_zone", "est_cost"]

ZERO_CHECK_FIELDS = ["insurance", "ioss", "import_taxes", "peak_season",
                      "vat_eu", "eu_extra"]


# ===========================================================================
# 2. robust readers
# ===========================================================================

def read_summary_rows(wb_vals, wb_formula, warnings, log=None):
    ws_v = wb_vals["Summary"]
    ws_f = wb_formula["Summary"]
    header_row, hmap = find_header_row(ws_v, SUMMARY_ALIASES, SUMMARY_REQUIRED)
    if log:
        cols_repr = {k: get_column_letter(v) for k, v in hmap.items()}
        log(f"Sheet Summary: tim thay header o dong {header_row}, cot: {cols_repr}")

    all_cols = list(hmap.values())
    records = []
    consecutive_blank = 0
    r = header_row + 1
    max_r = ws_v.max_row + 5
    while r <= max_r:
        first_val = get_cell_value_with_fallback(ws_v, ws_f, r, 1, warnings)
        client_id = get_cell_value_with_fallback(ws_v, ws_f, r, hmap["client_id"], warnings)

        if isinstance(first_val, str) and first_val.strip().lower() == "total":
            break

        if client_id is None:
            if row_is_blank(ws_v, ws_f, r, all_cols, warnings):
                consecutive_blank += 1
                if consecutive_blank >= 5:
                    break
                r += 1
                continue
            else:
                r += 1
                continue

        consecutive_blank = 0
        rec = dict(
            client_id=client_id,
            service=get_cell_value_with_fallback(ws_v, ws_f, r, hmap.get("service", 0), warnings) if "service" in hmap else None,
            manifest=get_cell_value_with_fallback(ws_v, ws_f, r, hmap.get("manifest", 0), warnings) if "manifest" in hmap else None,
            date=get_cell_value_with_fallback(ws_v, ws_f, r, hmap.get("date", 0), warnings) if "date" in hmap else None,
            total_track=get_cell_value_with_fallback(ws_v, ws_f, r, hmap["total_track"], warnings),
            total_usd=get_cell_value_with_fallback(ws_v, ws_f, r, hmap["total_usd"], warnings),
        )
        records.append(rec)
        r += 1

    if not records:
        raise ValueError(
            "Khong doc duoc dong du lieu nao tu sheet Summary cua debit.xlsx "
            f"(header tim thay o dong {header_row}, cot Client ID = "
            f"{get_column_letter(hmap['client_id'])})."
        )
    return records


def _is_repeated_header_row(ws_v, ws_f, row, hmap, warnings):
    """Sheet 'All' repeats a full Title/blank/Header block for every client
    (Title -> blank -> Header -> data... -> subtotal -> blank -> blank ->
    Title -> ... again). The very first Header row is located once by
    find_header_row(), but every later occurrence of that same Header row
    must be detected here and skipped -- otherwise its text cells (e.g. the
    literal string "CLIENT ID" sitting in the client-id column) get read as
    if they were a real data row, creating a bogus 'client' that swallows
    unrelated rows and corrupts column values throughout the sheet.

    A row is considered a repeated header row when BOTH the 'no' column and
    the 'client_id' column contain text that normalizes to one of their
    known header aliases (e.g. "NO." and "CLIENT ID").
    """
    no_col = hmap.get("no")
    cid_col = hmap.get("client_id")
    if no_col is None or cid_col is None:
        return False
    no_val = normalize_header(get_cell_value_with_fallback(ws_v, ws_f, row, no_col, warnings))
    cid_val = normalize_header(get_cell_value_with_fallback(ws_v, ws_f, row, cid_col, warnings))
    return no_val in ALL_ALIASES["no"] and cid_val in ALL_ALIASES["client_id"]


def read_all_rows(wb_vals, wb_formula, warnings, log=None):
    ws_v = wb_vals["All"]
    ws_f = wb_formula["All"]
    header_row, hmap = find_header_row(ws_v, ALL_ALIASES, ALL_REQUIRED)
    if log:
        log(f"Sheet All: tim thay header o dong {header_row}, {len(hmap)} cot nhan dien duoc.")

    def dim_cols(base_col):
        return [base_col, base_col + 1, base_col + 2]

    watch_cols = [v for k, v in hmap.items() if k not in ("dim_actual", "dim_book")]
    for k in ("dim_actual", "dim_book"):
        if k in hmap:
            watch_cols.extend(dim_cols(hmap[k]))

    records = []
    consecutive_blank = 0
    skipped_header_repeats = 0
    r = header_row + 1
    max_r = ws_v.max_row + 5
    while r <= max_r:
        first_val = get_cell_value_with_fallback(ws_v, ws_f, r, 1, warnings)
        client_id = get_cell_value_with_fallback(ws_v, ws_f, r, hmap["client_id"], warnings)

        if isinstance(first_val, str) and first_val.strip().lower() == "total":
            break

        # 'All' repeats Title/Header blocks per client (see docstring of
        # _is_repeated_header_row). Detect and skip these, otherwise the
        # literal header text (e.g. "CLIENT ID") gets ingested as if it
        # were a real data row.
        if _is_repeated_header_row(ws_v, ws_f, r, hmap, warnings):
            skipped_header_repeats += 1
            consecutive_blank = 0
            r += 1
            continue

        if client_id is None:
            if row_is_blank(ws_v, ws_f, r, watch_cols, warnings):
                consecutive_blank += 1
                if consecutive_blank >= 5:
                    break
                r += 1
                continue
            else:
                r += 1
                continue

        consecutive_blank = 0

        def gv(field, default=None):
            if field not in hmap:
                return default
            return get_cell_value_with_fallback(ws_v, ws_f, r, hmap[field], warnings)

        dim_a = dim_cols(hmap["dim_actual"])
        dim_b = dim_cols(hmap["dim_book"])
        rec = dict(
            reference_number=gv("reference_number"),
            client_id=client_id,
            overpack_id=gv("overpack_id"),
            tracking_number=gv("tracking_number"),
            waybill_number=gv("waybill_number"),
            gross_actual=gv("gross_actual"),
            gross_book=gv("gross_book"),
            dim_actual=[get_cell_value_with_fallback(ws_v, ws_f, r, c, warnings) for c in dim_a],
            dim_book=[get_cell_value_with_fallback(ws_v, ws_f, r, c, warnings) for c in dim_b],
            vol_weight=gv("vol_weight"),
            charge_weight=gv("charge_weight"),
            amount=gv("amount"),
            oversize_fee=gv("oversize_fee"),
            insurance=gv("insurance", 0),
            ioss=gv("ioss", 0),
            import_taxes=gv("import_taxes", 0),
            peak_season=gv("peak_season", 0),
            vat_eu=gv("vat_eu", 0),
            import_taxes_us=gv("import_taxes_us"),
            eu_extra=gv("eu_extra", 0),
            total_usd=gv("total_usd"),
            zip_code=gv("zip_code"),
            est_zone=gv("est_zone"),
            est_cost=gv("est_cost"),
            unit_price=gv("unit_price"),
        )
        records.append(rec)
        r += 1

    if log and skipped_header_repeats:
        log(f"Sheet All: da bo qua {skipped_header_repeats} dong header lap lai "
            f"(moi khach hang co 1 khoi title+header rieng).")

    if not records:
        raise ValueError(
            "Khong doc duoc dong du lieu nao tu sheet All cua debit.xlsx "
            f"(header tim thay o dong {header_row}, cot Client ID = "
            f"{get_column_letter(hmap['client_id'])})."
        )
    return records


# ===========================================================================
# 3. client name / code lookup (giu nguyen)
# ===========================================================================

def load_client_names_and_codes(ct_path):
    """Read the 'CODE KHACH + QB' sheet of ct.xlsx and evaluate, in Python,
    the 'Code debit US' formula (=TEXT(E2,"00")&RIGHT(A,3)&TEXT(F2,"00")&" "&G1)
    for every client, without needing a formula engine.
    Returns: dict code -> name, dict code -> code_debit_us, month(int), period(str)
    """
    wb = load_workbook(ct_path, data_only=True)
    ws = wb["CODE KHÁCH + QB"]
    month = ws["E2"].value
    period = ws["F2"].value
    suffix = ws["G1"].value or "-"
    names = {}
    codes_us = {}
    r = 2
    while True:
        code = ws.cell(r, 1).value
        if code is None:
            if ws.cell(r + 1, 1).value is None and ws.cell(r + 2, 1).value is None:
                break
            r += 1
            continue
        name = ws.cell(r, 2).value
        names[str(code).strip()] = name
        last3 = str(code).strip()[-3:]
        code_us = f"{_fmt_excel_text('00', month)}{last3}{_fmt_excel_text('00', period)} {suffix}"
        codes_us[str(code).strip()] = code_us
        r += 1
    return names, codes_us, month, period


def compute_estimate_costs(waybill, est_cost):
    """Reproduce the 4 classification formulas found in ct.xlsx (DEBIT sheet)
    exactly:
      LM     = LEFT(waybill,2)="LM"
      IB     = contains "-" AND not LM
      GORI   = LEN(waybill)=9
      CLUTCH = numeric, 15<=len<=19, starts with "17"
    """
    w = "" if waybill is None else str(waybill)
    est = est_cost or 0
    lm = est if w[:2] == "LM" else 0
    ib = est if ("-" in w and w[:2] != "LM") else 0
    gori = est if len(w) == 9 else 0
    clutch = 0
    if w[:2] == "17" and 15 <= len(w) <= 19:
        try:
            float(w)
            clutch = est
        except ValueError:
            clutch = 0
    return lm, ib, gori, clutch


def _set(ws, row, col, value, font=None, align=None, border=None, numfmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:
        c.font = font
    if align:
        c.alignment = align
    if border:
        c.border = border
    if numfmt:
        c.number_format = numfmt
    return c


# ===========================================================================
# 4. main entry point
# ===========================================================================

FLIGHT_TITLE_RE = re.compile(
    r"^\s*\S.*?\s*/\s*\d{1,2}\s+[A-Za-z]{3}\s+\S+\s*/\s*\S+\s*$"
)


def build_debit_editted(debit_path, ct_path, output_path, ci_code="",
                         progress=None, flight_title=None, ec_code=None):
    """Main entry point. progress: optional callable(str) for status updates.

    flight_title: optional. If given, this EXACT string is used as the
    flight/manifest title printed at the top of every block in 'All' and
    'Summary' (e.g. "297 3998 4210 / 25 AUG JFK / CI0794"), instead of the
    title being auto-built from manifest/date/service columns + ci_code.
    Expected format: "<MAWB> / <DD MON> <DEST> / <CI code>".
    When flight_title is not given, behavior is unchanged (auto-build from
    data + optional ci_code suffix), so existing callers keep working.

    ec_code: optional. If given, this EXACT string (e.g. "EC260802-") is
    written into the Summary/All header cell instead of the value being
    auto-computed from month/period. When not given, behavior is unchanged
    (auto-computed as before).
    """

    def log(msg):
        if progress:
            progress(msg)

    warnings = []

    if flight_title is not None and flight_title.strip():
        flight_title = flight_title.strip()
        if not FLIGHT_TITLE_RE.match(flight_title):
            raise ValueError(
                "flight_title khong dung dinh dang. Vui long nhap theo mau: "
                '"297 3998 4210 / 25 AUG JFK / CI0794" '
                "(MAWB / NGAY THANG SAN BAY DEN / MA CI)."
            )
    else:
        flight_title = None

    log("Doc du lieu ct.xlsx (bang khach hang)...")
    client_names, client_codes_us, month, period = load_client_names_and_codes(ct_path)

    log("Mo file mau dinh dang ct.xlsx (sheet DEBIT)...")
    ct_wb = load_workbook(ct_path, data_only=False)
    ct_debit = ct_wb["DEBIT"]
    # style templates lifted from the real ct.xlsx DEBIT sheet
    ALL_TITLE1_STYLE = get_style_row(ct_debit, 12, N_ALL_COLS)
    ALL_TITLE2_STYLE = get_style_row(ct_debit, 13, N_ALL_COLS)
    ALL_HEADER_STYLE = get_style_row(ct_debit, 14, N_ALL_COLS)
    ALL_DATA_STYLE = get_style_row(ct_debit, 15, N_ALL_COLS)
    ALL_SUBTOTAL_STYLE = get_style_row(ct_debit, 19, N_ALL_COLS)

    # --- fix: normalize font across the SUBTOTAL row's currency columns ---
    # copy_cell_style() copies font/size verbatim, per column, from the
    # template row (ct.xlsx, sheet DEBIT, row 19). If ONE of the currency
    # columns (P..W = DEBIT..ESTIMATE CLUTCH COST) was saved in ct.xlsx with
    # a different font/size than its neighbours (this has happened for
    # "ESTIMATE GORI COST"), that mismatch gets reproduced identically on
    # every client's subtotal row (and on the grand-total row, which reuses
    # this same style array) -- the number then renders small/unbold in the
    # bottom-right corner of the cell instead of matching the other totals.
    # Fix: pick the font shared by the MAJORITY of the currency columns as
    # the canonical subtotal font, and force any outlier column onto it.
    _subtotal_currency_idx = sorted(c for c in CURRENCY_COLS_ALL if 16 <= c <= 23)

    def _font_key(f):
        return (f.name, f.size, f.bold, f.italic,
                f.color.rgb if f.color and f.color.rgb else None)

    _font_counts = {}
    for _c in _subtotal_currency_idx:
        _font_counts[_font_key(ALL_SUBTOTAL_STYLE[_c - 1].font)] = \
            _font_counts.get(_font_key(ALL_SUBTOTAL_STYLE[_c - 1].font), 0) + 1
    if _font_counts:
        _canonical_key = max(_font_counts, key=_font_counts.get)
        _canonical_font = next(
            copy.copy(ALL_SUBTOTAL_STYLE[_c - 1].font)
            for _c in _subtotal_currency_idx
            if _font_key(ALL_SUBTOTAL_STYLE[_c - 1].font) == _canonical_key
        )
        for _c in _subtotal_currency_idx:
            if _font_key(ALL_SUBTOTAL_STYLE[_c - 1].font) != _canonical_key:
                log(f"  - Canh bao: font o cot subtotal "
                    f"{get_column_letter(_c)} (dong 19 sheet DEBIT cua ct.xlsx) "
                    f"khac voi cac cot tien te con lai -> da tu dong sua ve "
                    f"font/size chung ({_canonical_font.name}/{_canonical_font.size}).")
                ALL_SUBTOTAL_STYLE[_c - 1].font = copy.copy(_canonical_font)

    N_SUM_COLS = len(SUMMARY_HEADERS)
    SUM_TITLE1_STYLE = get_style_row(ct_debit, 85, N_SUM_COLS)
    SUM_TITLE2_STYLE = get_style_row(ct_debit, 86, N_SUM_COLS)
    SUM_HEADER_STYLE = get_style_row(ct_debit, 87, N_SUM_COLS)
    SUM_DATA_STYLE = get_style_row(ct_debit, 88, N_SUM_COLS)

    log("Mo debit.xlsx goc...")
    wb = load_workbook(debit_path, data_only=False)
    wb_vals = load_workbook(debit_path, data_only=True)

    log("Format cac cot tien te (2 so le thap phan, dau phay hang nghin) "
        "tren tat ca cac sheet truoc khi tao du lieu...")
    for sheet_name in wb.sheetnames:
        if sheet_name in ("Summary", "All"):
            # rebuilt from scratch further below; formatted there instead
            continue
        apply_currency_format_to_sheet(wb[sheet_name], log)

    log("Doc sheet Summary (theo header, khong gia dinh vi tri co dinh)...")
    summary_rows = read_summary_rows(wb_vals, wb, warnings, log)

    log("Doc sheet All (theo header, khong gia dinh vi tri co dinh)...")
    all_rows = read_all_rows(wb_vals, wb, warnings, log)

    if warnings:
        log(f"Canh bao: {len(warnings)} o co cong thuc chua duoc tinh san "
            f"(cached value trong), tool se coi la trong. Vd: {warnings[:5]}")

    # ---- global title strings ----
    first = summary_rows[0]
    date_val = first.get("date")
    if isinstance(date_val, str):
        try:
            date_val = dt.datetime.fromisoformat(date_val.split(" ")[0])
        except ValueError:
            date_val = None
    if isinstance(date_val, dt.datetime):
        day_mon = f"{date_val.day:02d} {MONTH_ABBR[date_val.month - 1]}"
        year_2 = f"{date_val.year % 100:02d}"
    else:
        day_mon = ""
        year_2 = ""

    if flight_title is None:
        # legacy behavior: auto-build from data + optional ci_code suffix
        dest = str(first.get("service") or "").split("-")[-1]
        manifest_str = first.get("manifest") or ""
        flight_title = f"{manifest_str} / {day_mon} {dest}".strip()
        if ci_code:
            flight_title += f" / {ci_code}"
        log(f"Tieu de tu dong tao: '{flight_title}'")
    else:
        log(f"Su dung tieu de nguoi dung nhap: '{flight_title}'")

    if ec_code is not None and ec_code.strip():
        ec_code = ec_code.strip()
        log(f"Su dung ma EC nguoi dung nhap: '{ec_code}'")
    else:
        ec_code = f"EC{year_2}{_fmt_excel_text('00', month)}{_fmt_excel_text('00', period)}-"
        log(f"Ma EC tu dong tao: '{ec_code}'")

    # ---- zero-total columns (dropped from All sheet, like the original process) ----
    drop_zero_fields = set()
    for field in ZERO_CHECK_FIELDS:
        s = 0.0
        for row in all_rows:
            v = row.get(field)
            if isinstance(v, (int, float)):
                s += v
        if abs(s) < 1e-9:
            drop_zero_fields.add(field)
    log(f"Truong co tong = 0 (da bi bo qua o output): {sorted(drop_zero_fields)}")

    # ================= build 'CODE KHACH + QB' helper sheet =================
    log("Nhung bang khach hang vao workbook (de dung VLOOKUP noi bo)...")
    if "CODE KHÁCH + QB" in wb.sheetnames:
        del wb["CODE KHÁCH + QB"]
    lookup_ws = wb.create_sheet("CODE KHÁCH + QB")
    lookup_ws.sheet_state = "hidden"
    lookup_ws["A1"] = "CODE"
    lookup_ws["B1"] = "Khách hàng"
    for i, (code, name) in enumerate(client_names.items(), start=2):
        lookup_ws.cell(i, 1, code)
        lookup_ws.cell(i, 2, name)

    # =========================== rebuild SUMMARY =============================
    log("Xay dung lai sheet Summary...")
    ws = wb["Summary"]
    ws.delete_rows(1, ws.max_row)
    ws.merged_cells.ranges = type(ws.merged_cells.ranges)()
    for dim in list(ws.column_dimensions.keys()):
        del ws.column_dimensions[dim]

    ws.merge_cells("A1:I1")
    for ci in range(1, N_SUM_COLS + 1):
        copy_cell_style(SUM_TITLE1_STYLE[ci - 1], ws.cell(1, ci), border_fallback=False)
    _set(ws, 1, 1, flight_title)
    _set_font_size_bold(ws.cell(1, 1), size=20, bold=True)

    ws.merge_cells("G2:H2")
    for ci in range(1, N_SUM_COLS + 1):
        copy_cell_style(SUM_TITLE2_STYLE[ci - 1], ws.cell(2, ci), border_fallback=False)
    _set(ws, 2, 7, ec_code)

    for ci, h in enumerate(SUMMARY_HEADERS, start=1):
        copy_cell_style(SUM_HEADER_STYLE[ci - 1], ws.cell(3, ci))
        ws.cell(3, ci).value = h
        _set_font_size_bold(ws.cell(3, ci), size=15, bold=True)

    start_row = 4
    for i, rec in enumerate(summary_rows):
        rr = start_row + i
        client_id = rec["client_id"]
        vals = [
            i + 1,
            client_id,
            f"=VLOOKUP(B{rr},'CODE KHÁCH + QB'!$A:$B,2,0)",
            rec["total_track"],
            rec["total_usd"],
            f"=SUMIF(All!C:C,Summary!B{rr},All!T:T)",
            f"=SUMIF(All!C:C,Summary!B{rr},All!U:U)",
            f"=SUMIF(All!C:C,Summary!B{rr},All!V:V)",
            f"=SUMIF(All!C:C,Summary!B{rr},All!W:W)",
            None,
        ]
        for ci, val in enumerate(vals, start=1):
            copy_cell_style(SUM_DATA_STYLE[ci - 1], ws.cell(rr, ci), border_fallback=True)
            ws.cell(rr, ci).value = val
            if ci in CURRENCY_COLS_SUMMARY:
                ws.cell(rr, ci).number_format = CURRENCY_NUMBER_FORMAT
            _set_font_size_bold(ws.cell(rr, ci), size=15, bold=False)

    total_row = start_row + len(summary_rows)
    for ci in range(1, N_SUM_COLS + 1):
        copy_cell_style(SUM_HEADER_STYLE[ci - 1], ws.cell(total_row, ci), border_fallback=True)
        _set_font_size_bold(ws.cell(total_row, ci), size=15, bold=True)
    _set(ws, total_row, 1, "Total")
    for col, colletter in [(4, "D"), (5, "E"), (6, "F"), (7, "G"), (8, "H"), (9, "I")]:
        ws.cell(total_row, col).value = f"=SUM({colletter}{start_row}:{colletter}{total_row - 1})"
        if col in CURRENCY_COLS_SUMMARY:
            ws.cell(total_row, col).number_format = CURRENCY_NUMBER_FORMAT

    widths = {"A": 8, "B": 11, "C": 23, "D": 10, "E": 15, "F": 15, "G": 15, "H": 15, "I": 15, "J": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 65
    ws.row_dimensions[2].height = 65
    ws.row_dimensions[3].height = 78
    for rr in range(start_row, total_row + 1):
        ws.row_dimensions[rr].height = 65

    # =========================== rebuild ALL ==================================
    log("Xay dung lai sheet All (nhom theo khach hang, them cong thuc)...")
    ws2 = wb["All"]
    ws2.delete_rows(1, ws2.max_row)
    for dim in list(ws2.column_dimensions.keys()):
        del ws2.column_dimensions[dim]
    ws2.merged_cells.ranges = type(ws2.merged_cells.ranges)()

    groups = []
    seen = {}
    for row in all_rows:
        cid = row["client_id"]
        if cid not in seen:
            seen[cid] = len(groups)
            groups.append([])
        groups[seen[cid]].append(row)

    # column A ("NO.") is intentionally borderless in data rows, matching both
    # ct.xlsx's own template and the reference completed file.
    NO_BORDER_DATA_COLS = {1}

    def write_group_header(row_idx, client_id):
        name = client_names.get(str(client_id).strip(), "")
        code_us = client_codes_us.get(str(client_id).strip(), "")

        for ci in range(1, N_ALL_COLS + 1):
            copy_cell_style(ALL_TITLE1_STYLE[ci - 1], ws2.cell(row_idx, ci), border_fallback=False)
        ws2.cell(row_idx, 1).value = flight_title
        ws2.merge_cells(start_row=row_idx, start_column=14, end_row=row_idx, end_column=18)
        ws2.cell(row_idx, 14).value = name
        ws2.merge_cells(start_row=row_idx, start_column=19, end_row=row_idx, end_column=21)
        ws2.cell(row_idx, 19).value = code_us

        blank_row = row_idx + 1
        for ci in range(1, N_ALL_COLS + 1):
            copy_cell_style(ALL_TITLE2_STYLE[ci - 1], ws2.cell(blank_row, ci), border_fallback=False)

        hr = row_idx + 2
        for ci, h in enumerate(ALL_HEADERS, start=1):
            copy_cell_style(ALL_HEADER_STYLE[ci - 1], ws2.cell(hr, ci))
            ws2.cell(hr, ci).value = h
        ws2.merge_cells(start_row=hr, start_column=8, end_row=hr, end_column=10)
        ws2.merge_cells(start_row=hr, start_column=11, end_row=hr, end_column=13)
        return hr

    cur_row = 3
    table_first_data_row = None
    table_last_data_row = None
    subtotal_rows = []
    for gi, grp in enumerate(groups):
        header_row = write_group_header(cur_row, grp[0]["client_id"])
        data_start = header_row + 1
        if table_first_data_row is None:
            table_first_data_row = data_start
        for j, row in enumerate(grp):
            rr = data_start + j
            waybill = row["waybill_number"]
            est_cost = row["est_cost"]
            lm, ib, gori, clutch = compute_estimate_costs(waybill, est_cost)
            dim_a = row["dim_actual"] + [None, None, None]
            dim_b = row["dim_book"] + [None, None, None]
            values = [
                j + 1,
                row["reference_number"],
                row["client_id"],
                row["overpack_id"],
                row["tracking_number"],
                row["gross_actual"],
                row["gross_book"],
                dim_a[0], dim_a[1], dim_a[2],
                dim_b[0], dim_b[1], dim_b[2],
                row["vol_weight"],
                row["charge_weight"],
                row["amount"],
                row["oversize_fee"],
                row["import_taxes_us"],
                row["total_usd"],
                None, None, None, None,   # T,U,V,W -> formulas below
                None,                       # X TOTAL COST -> formula
                row["zip_code"],
                waybill,
                est_cost,
                row["unit_price"],
                row["est_zone"],
            ]
            for ci, val in enumerate(values, start=1):
                copy_cell_style(ALL_DATA_STYLE[ci - 1], ws2.cell(rr, ci),
                                 border_fallback=True, skip_cols=NO_BORDER_DATA_COLS,
                                 this_col=ci)
                ws2.cell(rr, ci).value = val
                if ci in CURRENCY_COLS_ALL:
                    ws2.cell(rr, ci).number_format = CURRENCY_NUMBER_FORMAT
            ws2.cell(rr, 20).value = f"=IF(LEFT(Z{rr},2)=\"LM\",AA{rr},0)"
            ws2.cell(rr, 21).value = f"=IF(AND(ISNUMBER(FIND(\"-\",Z{rr})), LEFT(Z{rr},2)<>\"LM\"), AA{rr}, 0)"
            ws2.cell(rr, 22).value = f"=IF(LEN(Z{rr})=9, AA{rr}, 0)"
            ws2.cell(rr, 23).value = f"=IF(AND(ISNUMBER(--Z{rr}), LEN(Z{rr})>=15, LEN(Z{rr})<=19, LEFT(Z{rr},2)=\"17\"), AA{rr}, 0)"
            ws2.cell(rr, 24).value = f"=SUM(T{rr}:W{rr})"
            for ci in (20, 21, 22, 23, 24):
                ws2.cell(rr, ci).number_format = CURRENCY_NUMBER_FORMAT

        data_end = data_start + len(grp) - 1
        table_last_data_row = data_end
        subtotal_row = data_end + 1
        subtotal_rows.append(subtotal_row)
        for ci in range(1, N_ALL_COLS + 1):
            copy_cell_style(ALL_SUBTOTAL_STYLE[ci - 1], ws2.cell(subtotal_row, ci), border_fallback=True)
        ws2.cell(subtotal_row, 1).value = f"=COUNTIFS($C:$C,$C{data_end})"
        # IMPORTANT: these formulas now mirror the LIVE template formulas
        # found in ct.xlsx, sheet DEBIT, row 9 (confirmed by inspecting the
        # actual file) instead of a hand-written SUM(range):
        #   A9 = COUNTIFS($C:$C,$C8)
        #   F9/G9 = SUMIF($C:$C,$C8,F:F) / SUMIF($C:$C,$C8,G:G)   (no ROUND)
        #   P9..W9 = ROUND(SUMIF($C:$C,$C8,P:P),2) ... (every currency
        #            column is wrapped in ROUND(...,2))
        # $C8 in the template is simply "the CLIENT ID cell one row above
        # the subtotal row" -- i.e. the last data row of that client's
        # block -- so here that becomes $C{data_end}. SUMIF matches by
        # CLIENT ID across the whole column rather than a hardcoded row
        # range, exactly like the template, and -- like SUM -- it is NOT
        # affected by hidden/filtered rows, so it still works correctly if
        # this sheet is later filtered down to just the subtotal rows.
        # Rounding every client's currency subtotal to the cent (as the
        # template does) is also what keeps the final grand total's
        # floating-point summation from drifting off by a cent.
        for col_letter, col_idx in [("F", 6), ("G", 7)]:
            ws2.cell(subtotal_row, col_idx).value = (
                f"=SUMIF($C:$C,$C{data_end},{col_letter}:{col_letter})"
            )
        for col_letter, col_idx in [("P", 16), ("Q", 17), ("R", 18), ("S", 19),
                                     ("T", 20), ("U", 21), ("V", 22), ("W", 23)]:
            ws2.cell(subtotal_row, col_idx).value = (
                f"=ROUND(SUMIF($C:$C,$C{data_end},{col_letter}:{col_letter}),2)"
            )
            ws2.cell(subtotal_row, col_idx).number_format = CURRENCY_NUMBER_FORMAT
        cur_row = subtotal_row + 3  # 2 blank rows then next title

    # ---- grand total row: 1 blank row below the very last client subtotal ----
    # Because the per-client subtotal cells above are plain SUM(...) (see
    # note above -- required so they survive being filtered/hidden), Excel's
    # "ignore nested SUBTOTAL results" trick no longer applies here: a flat
    # SUBTOTAL/SUM over the whole data range would double count (raw rows +
    # their SUM already counted once). So the grand total goes back to
    # referencing ONLY the already-computed per-client subtotal cells
    # explicitly: =SUBTOTAL(9, <cell1>, <cell2>, ...). This is safe against
    # double counting, and (being SUBTOTAL) it also correctly excludes any
    # of those client rows that get hidden by a filter.
    if subtotal_rows:
        grand_row = subtotal_rows[-1] + 2  # 1 blank row gap, as requested
        log(f"Them dong TONG TAT CA (grand total) o dong {grand_row}, "
            f"cach dong subtotal cuoi cung 1 dong trong...")

        for ci in range(1, N_ALL_COLS + 1):
            copy_cell_style(ALL_SUBTOTAL_STYLE[ci - 1], ws2.cell(grand_row, ci), border_fallback=True)
        # NOTE: per request, this label cell is intentionally left BLANK
        # (no "TONG TAT CA / GRAND TOTAL" text) -- matching the reference
        # file, where the grand-total row has no text in column B either.

        for col_idx in GRAND_TOTAL_NESTED_COLS:
            letter = get_column_letter(col_idx)
            refs = ",".join(f"{letter}{sr}" for sr in subtotal_rows)
            ws2.cell(grand_row, col_idx).value = f"=SUBTOTAL(9,{refs})"
            if col_idx in CURRENCY_COLS_ALL:
                ws2.cell(grand_row, col_idx).number_format = CURRENCY_NUMBER_FORMAT

        # NOTE: by explicit request, the grand-total row totals ONLY the 11
        # columns in GRAND_TOTAL_NESTED_COLS above. All other columns (DIM,
        # volume/charge weight, TOTAL COST, ESTIMATE COST (USD), VALUE, ...)
        # are left blank on this row -- no SUBTOTAL/SUM formulas for them.

    # ---- center-align every used cell in 'All' (horizontal + vertical) ----
    # Requested: all data in sheet 'All' should be centered both
    # horizontally and vertically, regardless of what the ct.xlsx template
    # cell's original alignment was. wrap_text is preserved so multi-line
    # headers (e.g. "ESTIMATE\nLM COST") keep wrapping correctly.
    last_used_row = grand_row if subtotal_rows else cur_row
    for rr in range(1, last_used_row + 1):
        for ci in range(1, N_ALL_COLS + 1):
            c = ws2.cell(rr, ci)
            c.alignment = Alignment(
                horizontal="center", vertical="center",
                wrap_text=c.alignment.wrap_text,
            )

    all_widths = {"A": 8, "B": 20, "C": 14, "D": 15, "E": 24, "F": 12, "G": 12,
                  "H": 7, "I": 7, "J": 7, "K": 7, "L": 7, "M": 7, "N": 12, "O": 12,
                  "P": 12, "Q": 12, "R": 12, "S": 12, "T": 11, "U": 11, "V": 11,
                  "W": 11, "X": 11, "Y": 13, "Z": 26, "AA": 13, "AB": 9, "AC": 12}
    for col, w in all_widths.items():
        ws2.column_dimensions[col].width = w

    log("Luu file...")
    wb.save(output_path)
    if warnings:
        log(f"Hoan tat voi {len(warnings)} canh bao (xem chi tiet ben tren). "
            f"Neu debit.xlsx co cong thuc chua tinh, hay mo va luu lai (Ctrl+S) "
            f"trong Excel truoc khi chay tool de dam bao du lieu day du.")
    else:
        log("Hoan tat, khong co canh bao.")
    return output_path


if __name__ == "__main__":
    import sys
    # Cach dung:
    #   python engine.py debit.xlsx ct.xlsx output.xlsx "CI0794"
    #   python engine.py debit.xlsx ct.xlsx output.xlsx --title "297 3998 4210 / 25 AUG JFK / CI0794"
    debit_path, ct_path, output_path = sys.argv[1], sys.argv[2], sys.argv[3]
    rest = sys.argv[4:]
    ci = ""
    title = None
    if rest:
        if rest[0] == "--title" and len(rest) > 1:
            title = rest[1]
        else:
            ci = rest[0]
    build_debit_editted(debit_path, ct_path, output_path, ci_code=ci,
                         flight_title=title, progress=print)