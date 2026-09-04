import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import requests
import json
import re
import pandas as pd
import threading
import time
import os
import random
import tempfile
import concurrent.futures
import webbrowser
import sys
import socket
from urllib.parse import urlparse
from openpyxl.styles import PatternFill, Font

try:
    from google import genai
    from google.genai import types as genai_types
    HAS_GEMINI = True
except ImportError:
    HAS_GEMINI = False

API_URL = "https://us-street.api.smarty.com/street-address"
DEFAULT_SMARTY_KEY = "21102174564513388"
MIN_SMARTY_REQUEST_INTERVAL = 10.0
STABLE_USER_AGENT = "smarty-address-tool/1.0"

APP_DIR = os.path.dirname(os.path.abspath(sys.executable if getattr(sys, "frozen", False) else __file__))
DATA_DIR = os.path.join(APP_DIR, "data")
try:
    os.makedirs(DATA_DIR, exist_ok=True)
except OSError:
    DATA_DIR = os.path.join(os.getenv("LOCALAPPDATA", APP_DIR), "SmartyAddressTool", "data")
    os.makedirs(DATA_DIR, exist_ok=True)

# Pool User-Agent để xoay vòng theo từng request/attempt, tránh bị nhận diện
# là 1 "client" cố định gửi liên tục -> giảm khả năng dính "Too many requests".
USER_AGENT_POOL = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Edg/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
]

ACCEPT_LANGUAGE_POOL = [
    "en-US,en;q=0.9",
    "en-US,en;q=0.8,vi;q=0.5",
    "en-GB,en;q=0.9",
    "en-US,en;q=0.9,fr;q=0.6",
    "en-CA,en;q=0.9",
]

GROQ_MODELS = ["llama-3.3-70b-versatile", "llama-3.1-8b-instant"]
OPENROUTER_MODELS = [
    "meta-llama/llama-3.1-8b-instruct:free",
    "google/gemini-2.0-flash-exp:free",
    "mistralai/mistral-7b-instruct:free",
]
GEMINI_FALLBACK_MODELS = ['gemini-flash-latest', 'gemini-2.5-flash', 'gemini-2.5-flash-lite', 'gemini-2.0-flash']

# File lưu lại PHIÊN LÀM VIỆC GẦN NHẤT (toàn bộ collected_data - tức là các dòng đã gọi
# xong API Smarty, KÈM trạng thái/lý do AI đã chấm gần nhất nếu có). Nhờ vậy, dù đã đóng
# hẳn cửa sổ 1 lần chạy xong, người dùng vẫn có thể mở lại tool và bấm "Kiểm tra lại" để
# hỏi AI lại (vd hỏi lại GPT lần nữa) MÀ KHÔNG cần gọi lại Smarty API từ đầu (tốn thời
# gian + có thể dính rate-limit lại). File này được ghi đè mỗi khi có 1 phiên xử lý/kiểm
# tra lại mới hoàn tất, để luôn phản ánh phiên GẦN NHẤT.
SESSION_STATE_FILE = "smarty_last_session.json"
ADDRESS_CACHE_FILE = "smarty_address_cache.json"

class SmartyApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Smarty API & Google Gemini Auto Check (Xoay vòng Proxy)")
        self.root.geometry("820x780")

        self.excel_path = ""
        self.output_dir = ""
        self.stop_requested = False

        self.session_state_dir = DATA_DIR
        self._migrate_legacy_data_files()

        self.session = requests.Session()
        self._last_smarty_request_at = 0.0
        self.gemini_client = None
        self.proxy_list = self._load_proxies()
        self.api_key_list = self._load_api_keys()
        self._identity_counter = 0

        # ----- Trạng thái xoay vòng KEY/PROXY LINH ĐỘNG -----
        # Thay vì xoay vòng "mù" theo thứ tự cố định (round-robin cứng nhắc như trước), tool
        # theo dõi từng KEY và từng PROXY xem có đang bị 429/lỗi mạng gần đây không, để lần
        # chọn tiếp theo TỰ ĐỘNG ưu tiên né các key/proxy vừa bị chặn, chỉ quay lại dùng khi
        # đã hết thời gian "khóa tạm" (cooldown). Mỗi mục có dạng:
        #   {"blocked_until": <timestamp thời điểm hết bị khóa>, "fail_streak": <số lần bị liên tiếp>}
        self.key_status = {}
        self.proxy_status = {}
        self._key_rr_counter = 0
        self._proxy_rr_counter = 0
        # Chỉ cảnh báo 1 lần/phiên chạy (không spam log) khi phát hiện dấu hiệu API Key duy
        # nhất có thể đã hết quota (bị 429 liên tục nhiều lần dù đã đổi proxy/User-Agent).
        self._warned_single_key_session = False

        self.setup_gui()
        # Ngay khi mở tool, dò xem có phiên cũ (đã lưu từ lần chạy trước) hay không, để
        # hiện nút "Kiểm tra lại" nếu có -> người dùng có thể kiểm tra lại BẤT CỨ LÚC NÀO,
        # kể cả sau khi đã tắt hẳn tool rồi mở lại.
        self._refresh_recheck_button()

    def _load_proxies(self):
        path = os.path.join(self.session_state_dir, "proxies.txt")
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                return [line.strip() for line in f if line.strip()]
        return []

    def _load_api_keys(self):
        """Tuỳ chọn: nếu có file smarty_keys.txt (mỗi dòng 1 key Smarty),
        tool sẽ XOAY VÒNG nhiều key cùng với Proxy/User-Agent để né rate-limit
        triệt để hơn. Nếu không có file, dùng key mặc định."""
        path = os.path.join(self.session_state_dir, "smarty_keys.txt")
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                keys = [line.strip() for line in f if line.strip()]
            if keys:
                return keys
        return [DEFAULT_SMARTY_KEY]

    def _migrate_legacy_data_files(self):
        for filename in (
            "proxies.txt",
            "smarty_keys.txt",
            SESSION_STATE_FILE,
            ADDRESS_CACHE_FILE,
        ):
            target = os.path.join(self.session_state_dir, filename)
            if os.path.exists(target):
                continue
            for legacy_dir in (os.getcwd(), APP_DIR):
                source = os.path.join(legacy_dir, filename)
                if source == target or not os.path.exists(source):
                    continue
                try:
                    os.replace(source, target)
                except OSError:
                    pass
                break

    def _pick_available(self, values, status_dict, counter_attr):
        """Chọn 1 phần tử từ 'values' (danh sách key hoặc proxy), ƯU TIÊN các phần tử KHÔNG
        đang bị khóa tạm (cooldown). Nếu còn ít nhất 1 phần tử rảnh, xoay vòng round-robin
        CHỈ trong nhóm đó (bỏ qua hẳn phần tử đang bị khóa). Nếu TẤT CẢ đều đang bị khóa (vd
        chỉ có 1 key duy nhất và nó vừa bị 429), đành phải dùng lại phần tử nào sắp hết khóa
        SỚM NHẤT (đỡ phải chờ lâu nhất có thể)."""
        if not values:
            return None
        now = time.time()
        available = [v for v in values if status_dict.get(v, {}).get("blocked_until", 0) <= now]
        counter = getattr(self, counter_attr)
        if available:
            chosen = available[counter % len(available)]
            setattr(self, counter_attr, counter + 1)
            return chosen
        # Không còn phần tử nào rảnh -> chọn phần tử hết cooldown sớm nhất, thay vì random.
        return min(values, key=lambda v: status_dict.get(v, {}).get("blocked_until", 0))

    def _mark_key_rate_limited(self, key):
        """Đánh dấu 1 API Key vừa bị Smarty trả về 429 (rate limit). Khóa tạm key này lại
        trong 1 khoảng thời gian tăng dần theo cấp số nhân (backoff: 5s, 10s, 20s, 40s...,
        tối đa 5 phút) mỗi lần liên tiếp bị 429, để những lần chọn identity SAU sẽ tự động
        né key này ra và ưu tiên key khác (nếu có) - đây chính là phần 'tự đổi key linh động'."""
        st = self.key_status.setdefault(key, {"blocked_until": 0.0, "fail_streak": 0})
        st["fail_streak"] += 1
        cooldown = min(5 * (2 ** (st["fail_streak"] - 1)), 300)
        st["blocked_until"] = time.time() + cooldown
        return cooldown, st["fail_streak"]

    def _mark_key_success(self, key):
        """Gọi thành công (hoặc lỗi khác không phải rate-limit) bằng key này -> coi như key
        đang khỏe mạnh trở lại, xóa hết cờ 'nghi ngờ' để lần sau được ưu tiên chọn lại."""
        st = self.key_status.setdefault(key, {"blocked_until": 0.0, "fail_streak": 0})
        st["fail_streak"] = 0
        st["blocked_until"] = 0.0

    def _mark_proxy_issue(self, proxy_url):
        """Tương tự _mark_key_rate_limited nhưng cho PROXY khi gặp lỗi mạng/timeout (thường là
        proxy chập chờn/chết tạm thời) - khóa tạm ngắn hơn key vì lỗi mạng thường qua nhanh."""
        if not proxy_url:
            return 0
        st = self.proxy_status.setdefault(proxy_url, {"blocked_until": 0.0, "fail_streak": 0})
        st["fail_streak"] += 1
        cooldown = min(3 * (2 ** (st["fail_streak"] - 1)), 120)
        st["blocked_until"] = time.time() + cooldown
        return cooldown

    def _mark_proxy_success(self, proxy_url):
        if not proxy_url:
            return
        st = self.proxy_status.setdefault(proxy_url, {"blocked_until": 0.0, "fail_streak": 0})
        st["fail_streak"] = 0
        st["blocked_until"] = 0.0

    def _mask_key(self, key):
        """Che bớt API Key khi in ra log (chỉ hiện 4 ký tự cuối), tránh lộ trọn key ra màn
        hình/log file trong khi vẫn đủ để người dùng phân biệt được đang là key nào."""
        key = str(key or "")
        return f"...{key[-4:]}" if len(key) > 4 else "*" * len(key)

    def _proxy_label(self, proxy_url):
        if not proxy_url:
            return "không dùng proxy"
        try:
            return f"proxy #{self.proxy_list.index(proxy_url)}"
        except ValueError:
            return "proxy (?)"

    def _request_route_label(self, proxy_url):
        if not proxy_url:
            return "IP thật (không qua proxy)"
        try:
            parsed = urlparse(proxy_url if "://" in proxy_url else f"http://{proxy_url}")
            host = parsed.hostname or "?"
            try:
                resolved_ip = socket.gethostbyname(host)
                host_text = f"{host} ({resolved_ip})" if resolved_ip != host else host
            except socket.gaierror:
                host_text = host
            port_text = f":{parsed.port}" if parsed.port else ""
            return f"{self._proxy_label(proxy_url)} - {host_text}{port_text}"
        except ValueError:
            return f"{self._proxy_label(proxy_url)} - {proxy_url}"

    def _next_identity(self):
        """Trả về 1 'danh tính' request mới: Proxy + Smarty Key + User-Agent + Accept-Language.
        KHÁC với round-robin cứng nhắc trước đây, giờ Proxy VÀ Key được chọn LINH ĐỘNG qua
        _pick_available(): tự động BỎ QUA những key/proxy vừa bị 429/lỗi mạng gần đây (đang
        trong thời gian cooldown) và ưu tiên dùng key/proxy còn 'khỏe mạnh', thay vì cứ lặp
        lại đúng thứ tự cố định kể cả khi biết chắc phần tử đó vừa mới bị chặn."""
        i = self._identity_counter
        self._identity_counter += 1

        proxy_url = self._pick_available(self.proxy_list, self.proxy_status, "_proxy_rr_counter")
        api_key = self._pick_available(self.api_key_list, self.key_status, "_key_rr_counter")
        user_agent = STABLE_USER_AGENT
        accept_language = "en-US,en;q=0.9"

        headers = {
            "Referer": "https://www.smarty.com/",
            "Origin": "https://www.smarty.com",
            "User-Agent": user_agent,
            "Accept-Language": accept_language,
        }
        return {
            "proxy_url": proxy_url,
            "api_key": api_key,
            "headers": headers,
            "index": i,
        }

    def _warmup_proxies(self):
        """'Đánh thức' từng proxy TRƯỚC khi bắt đầu xử lý dòng thật, thay vì để dòng
        đầu tiên lãnh đủ hàng loạt lỗi timeout. Nhiều proxy (đặc biệt proxy free/dùng
        chung) có kết nối 'nguội': lần chạm đầu tiên tới 1 proxy mới phải bắt tay
        TCP/TLS lại từ đầu (và nếu đó là proxy xoay IP/dạng gateway thì backend có
        thể cần thời gian để cấp phiên mới) nên hay timeout, nhưng các lần sau đó
        thường ổn định. Việc này chạy song song, timeout ngắn, và CHỈ in 1 dòng
        tổng kết cuối cùng - không in từng lỗi riêng lẻ để tránh gây cảm giác
        "app bị lỗi liên tục" ngay khi vừa bắt đầu."""
        if not self.proxy_list:
            return

        self.log(f"Đang khởi động {len(self.proxy_list)} proxy trước khi xử lý (có thể mất vài giây)...")

        def _check(proxy_url):
            try:
                requests.get(
                    "https://www.smarty.com/",
                    proxies={"http": proxy_url, "https": proxy_url},
                    timeout=8,
                )
                return True
            except requests.exceptions.RequestException:
                return False

        results = []
        with concurrent.futures.ThreadPoolExecutor(max_workers=min(10, len(self.proxy_list))) as executor:
            futures = [executor.submit(_check, p) for p in self.proxy_list]
            for fut in concurrent.futures.as_completed(futures, timeout=20):
                try:
                    results.append(fut.result())
                except Exception:
                    results.append(False)

        ok_count = sum(1 for r in results if r)
        self.log(
            f"Đã khởi động xong proxy: {ok_count}/{len(self.proxy_list)} phản hồi tốt "
            f"(số còn lại vẫn sẽ được thử lại tự động trong lúc chạy nếu cần)."
        )

    def setup_gui(self):
        frame_top = tk.Frame(self.root)
        frame_top.pack(pady=(15, 5), padx=10, fill="x")

        self.btn_select_excel = tk.Button(frame_top, text="1. Chọn file Excel gốc", command=self.select_excel, width=20, bg="#28a745", fg="white", font=("Arial", 10, "bold"))
        self.btn_select_excel.pack(side="left", padx=(0, 10))

        self.lbl_excel_path = tk.Label(frame_top, text="Chưa chọn file...", fg="gray")
        self.lbl_excel_path.pack(side="left")

        frame_delay = tk.Frame(self.root)
        frame_delay.pack(pady=5, padx=10, fill="x")

        lbl_delay = tk.Label(frame_delay, text="Độ trễ API Smarty (giây):", font=("Arial", 10, "bold"))
        lbl_delay.pack(side="left")

        self.delay_var = tk.StringVar(value=str(MIN_SMARTY_REQUEST_INTERVAL))
        self.entry_delay = tk.Entry(frame_delay, textvariable=self.delay_var, width=8, font=("Consolas", 11), justify="center")
        self.entry_delay.pack(side="left", padx=10)

        frame_match = tk.LabelFrame(self.root, text=" Chế độ khớp địa chỉ Smarty (match) ", font=("Arial", 10, "bold"), fg="#0078D7")
        frame_match.pack(pady=10, padx=10, fill="x")

        # Mặc định "strict" - đúng như mặc định của Smarty khi nhập tay trên web, chỉ trả
        # kết quả dựa hoàn toàn trên dữ liệu gốc, không tự suy đoán/vá thêm bớt.
        self.match_mode_var = tk.StringVar(value="strict")

        self.rb_match_strict = tk.Radiobutton(
            frame_match, variable=self.match_mode_var, value="strict",
            text="Strict (Khuyến khích - mặc định): chỉ khớp dựa hoàn toàn trên dữ liệu gốc, giống hệt khi nhập tay vào Smarty, không tự vá/suy đoán",
            font=("Arial", 9, "bold"), justify="left", fg="#1da462"
        )
        self.rb_match_strict.pack(anchor="w", padx=10, pady=(5, 0))

        self.rb_match_enhanced = tk.Radiobutton(
            frame_match, variable=self.match_mode_var, value="enhanced",
            text="Enhanced: Smarty tự khớp mạnh tay hơn, có thể tự suy đoán/sửa dữ liệu mơ hồ để cố trả ra kết quả",
            font=("Arial", 9), justify="left"
        )
        self.rb_match_enhanced.pack(anchor="w", padx=10, pady=(0, 2))

        self.lbl_match_note = tk.Label(
            frame_match,
            text="* Khuyến khích dùng Strict để kết quả trung thực với dữ liệu gốc trong Excel. Chỉ chọn Enhanced nếu bạn "
                 "hiểu rõ và chấp nhận việc Smarty có thể tự suy đoán/sửa đổi địa chỉ để cố khớp.",
            fg="gray", font=("Arial", 8), justify="left", wraplength=760
        )
        self.lbl_match_note.pack(anchor="w", padx=10, pady=(0, 5))

        frame_ai = tk.LabelFrame(self.root, text=" Kiểm tra kết quả đáng ngờ bằng AI (Miễn phí) ", font=("Arial", 10, "bold"), fg="#1da462")
        frame_ai.pack(pady=10, padx=10, fill="x")

        self.use_ai_var = tk.BooleanVar(value=True)
        self.chk_use_ai = tk.Checkbutton(frame_ai, text="Bật tự động kiểm tra trạng thái bằng AI", variable=self.use_ai_var, font=("Arial", 10), command=self.toggle_ai_input)
        self.chk_use_ai.pack(anchor="w", padx=10, pady=5)

        frame_mode = tk.Frame(frame_ai)
        frame_mode.pack(anchor="w", padx=20, pady=(0, 5), fill="x")

        self.ai_mode_var = tk.StringVar(value="manual")
        self.rb_mode_auto = tk.Radiobutton(
            frame_mode, variable=self.ai_mode_var, value="auto",
            text="Tự động: hệ thống tự dò & gọi các AI miễn phí (Gemini / Groq / OpenRouter)",
            font=("Arial", 9), justify="left", command=self.toggle_ai_input
        )
        self.rb_mode_auto.pack(anchor="w")

        self.rb_mode_manual = tk.Radiobutton(
            frame_mode, variable=self.ai_mode_var, value="manual",
            text="Thủ công: Tool tạo sẵn Prompt để bạn dán vào AI ngoài (ChatGPT/Gemini web...), rồi dán JSON kết quả về",
            font=("Arial", 9), justify="left", command=self.toggle_ai_input
        )
        self.rb_mode_manual.pack(anchor="w")

        frame_ai_keys = tk.Frame(frame_ai)
        frame_ai_keys.pack(fill="x", padx=20, pady=5)

        lbl_gemini = tk.Label(frame_ai_keys, text="Gemini API Key:", width=18, anchor="w")
        lbl_gemini.grid(row=0, column=0, sticky="w", pady=2)
        self.entry_ai_key = tk.Entry(frame_ai_keys, width=45, font=("Consolas", 10), show="*")
        self.entry_ai_key.grid(row=0, column=1, padx=5, pady=2)

        lbl_groq = tk.Label(frame_ai_keys, text="Groq API Key (tuỳ chọn):", width=18, anchor="w")
        lbl_groq.grid(row=1, column=0, sticky="w", pady=2)
        self.entry_groq_key = tk.Entry(frame_ai_keys, width=45, font=("Consolas", 10), show="*")
        self.entry_groq_key.grid(row=1, column=1, padx=5, pady=2)

        lbl_openrouter = tk.Label(frame_ai_keys, text="OpenRouter API Key (tuỳ chọn):", width=18, anchor="w")
        lbl_openrouter.grid(row=2, column=0, sticky="w", pady=2)
        self.entry_openrouter_key = tk.Entry(frame_ai_keys, width=45, font=("Consolas", 10), show="*")
        self.entry_openrouter_key.grid(row=2, column=1, padx=5, pady=2)

        self.lbl_ai_note = tk.Label(
            frame_ai,
            text="* Chế độ Thủ công không cần API Key. Chế độ Tự động cần ít nhất 1 trong 3 API Key trên.",
            fg="gray", font=("Arial", 8)
        )
        self.lbl_ai_note.pack(anchor="w", padx=20, pady=(0, 5))

        self.lbl_analysis_note = tk.Label(
            frame_ai,
            text=("* Dù BẬT hay TẮT AI, tool luôn tự kiểm tra thêm phần \"analysis\" (DPV) của Smarty "
                  "(vacant / CMRA / no_stat / inactive / dpv_match_code). Kết quả cuối cùng có 3 mức: "
                  "TRUE (chắc chắn đúng) / SUSPECT (nghi ngờ - cần xem lại) / FALSE (lỗi). "
                  "AI không thể tự hạ một dòng đã bị đánh dấu nghi ngờ xuống thành TRUE."),
            fg="#1da462", font=("Arial", 8), justify="left", wraplength=760
        )
        self.lbl_analysis_note.pack(anchor="w", padx=20, pady=(0, 5))

        # Khởi tạo trạng thái enable/disable theo đúng giá trị mặc định hiện tại của use_ai_var
        # (giờ mặc định BẬT), thay vì luôn khóa cứng như trước.
        self.toggle_ai_input()

        if not HAS_GEMINI:
            lbl_gemini.config(text="Gemini (chưa cài 'google-genai'):", fg="red")

        frame_btns = tk.Frame(self.root)
        frame_btns.pack(pady=10)

        self.btn_start = tk.Button(frame_btns, text="2. Bắt đầu xử lý & Xuất file (TXT + Excel)", command=self.start_processing_thread, width=35, bg="#0078D7", fg="white", font=("Arial", 11, "bold"), state=tk.DISABLED)
        self.btn_start.pack(side="left", padx=5)

        self.btn_stop = tk.Button(frame_btns, text="Dừng lại", command=self.stop_processing, width=10, bg="#dc3545", fg="white", font=("Arial", 11, "bold"), state=tk.DISABLED)
        self.btn_stop.pack(side="left", padx=5)

        self.btn_clear_cache = tk.Button(
            frame_btns, text="Xóa cache địa chỉ", command=self.clear_address_cache,
            width=17, bg="#6c757d", fg="white", font=("Arial", 10, "bold")
        )
        self.btn_clear_cache.pack(side="left", padx=5)

        self.lbl_cache_info = tk.Label(self.root, text="", fg="gray", font=("Arial", 8))
        self.lbl_cache_info.pack(anchor="w", padx=10, pady=(0, 3))
        self._refresh_cache_info()

        # ----- Khu vực "Kiểm tra lại phiên cũ" -----
        # Khung này LUÔN tồn tại nhưng RỖNG (không có gì bên trong) nếu chưa từng có phiên
        # nào được lưu. Nút "Kiểm tra lại (<n> dòng)" chỉ được pack() vào khung này khi
        # _refresh_recheck_button() phát hiện có file phiên cũ hợp lệ trên đĩa - tức là
        # đã có ít nhất 1 lần tool gọi xong Smarty (dù có thể chưa hoàn tất bước AI).
        self.frame_recheck = tk.Frame(self.root)
        self.frame_recheck.pack(pady=(0, 5), padx=10, fill="x")

        self.btn_recheck = tk.Button(
            self.frame_recheck, text="Kiểm tra lại (0 dòng)",
            command=self.start_recheck_session, width=28,
            bg="#6f42c1", fg="white", font=("Arial", 10, "bold")
        )
        # Chưa pack() vội - _refresh_recheck_button() sẽ tự pack/ẩn tuỳ tình trạng.

        self.lbl_recheck_info = tk.Label(self.frame_recheck, text="", fg="gray", font=("Arial", 8), justify="left")
        # Cũng chưa pack() - hiện cùng lúc với nút.

        lbl_log = tk.Label(self.root, text="Tiến trình xử lý:", font=("Arial", 10, "bold"))
        lbl_log.pack(anchor="w", padx=10)

        self.log_text = scrolledtext.ScrolledText(self.root, font=("Consolas", 10), width=95, height=18, bg="#1E1E1E", fg="#D4D4D4")
        self.log_text.pack(pady=5, padx=10)

    def _set_ai_inputs_state(self, state):
        self.entry_ai_key.config(state=state)
        self.entry_groq_key.config(state=state)
        self.entry_openrouter_key.config(state=state)
        self.rb_mode_auto.config(state=state)
        self.rb_mode_manual.config(state=state)

    def toggle_ai_input(self):
        if self.use_ai_var.get():
            self._set_ai_inputs_state(tk.NORMAL)
        else:
            self._set_ai_inputs_state(tk.DISABLED)

    def log(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def _refresh_cache_info(self):
        path = os.path.join(self.session_state_dir, ADDRESS_CACHE_FILE)
        if not os.path.exists(path):
            self.lbl_cache_info.config(text="Cache địa chỉ: chưa có dữ liệu")
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            count = len(data) if isinstance(data, dict) else 0
            updated_at = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(os.path.getmtime(path)))
            self.lbl_cache_info.config(text=f"Cache địa chỉ: {count} bản ghi | Cập nhật: {updated_at}")
        except (OSError, json.JSONDecodeError):
            self.lbl_cache_info.config(text="Cache địa chỉ: không đọc được dữ liệu")

    def clear_address_cache(self):
        path = os.path.join(self.session_state_dir, ADDRESS_CACHE_FILE)
        if not os.path.exists(path):
            self._refresh_cache_info()
            return
        if not messagebox.askyesno("Xóa cache", "Xóa toàn bộ cache địa chỉ Smarty đã lưu?\n\nLần chạy sau sẽ gọi lại Smarty cho các địa chỉ này."):
            return
        try:
            os.remove(path)
            self._refresh_cache_info()
            self.log("[CACHE] Đã xóa toàn bộ cache địa chỉ Smarty.")
        except OSError as e:
            messagebox.showerror("Lỗi xóa cache", f"Không thể xóa cache:\n{e}")

    def select_excel(self):
        file_path = filedialog.askopenfilename(title="Chọn file Excel", filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_path:
            self.excel_path = file_path
            self.lbl_excel_path.config(text=self.excel_path, fg="black")
            self.btn_start.config(state=tk.NORMAL)
            self.log(f"Đã chọn file: {self.excel_path}")
            if self.proxy_list:
                self.log(f"Kết nối: dùng {len(self.proxy_list)} proxy từ data/proxies.txt; User-Agent cố định.")
            else:
                self.log("Kết nối: không có proxy, dùng IP thật; User-Agent cố định.")
            if len(self.api_key_list) > 1:
                self.log(f"Đã nạp {len(self.api_key_list)} Smarty Key từ file smarty_keys.txt (xoay vòng).")
            else:
                self.log(
                    "Chỉ đang dùng 1 API Key Smarty duy nhất (mặc định, không có smarty_keys.txt). "
                    "Nếu Smarty trả về 429, tool sẽ dừng an toàn để không phát sinh thêm request bị chặn. "
                    "Hãy kiểm tra quota hoặc cấu hình API key hợp lệ trong thư mục data."
                )
            self.log(
                f"Khoảng cách tối thiểu giữa các request Smarty: {MIN_SMARTY_REQUEST_INTERVAL:.0f} giây; "
                "khi 429 sẽ tự chờ rồi thử lại."
            )

    # ================================================================================
    # LƯU / ĐỌC PHIÊN LÀM VIỆC CŨ (để có thể "Kiểm tra lại" bất cứ lúc nào, không cần
    # gọi lại Smarty API - chỉ cần hỏi lại AI/GPT và dán JSON kết quả mới vào).
    # ================================================================================

    def _session_file_path(self):
        return os.path.join(self.session_state_dir, SESSION_STATE_FILE)

    def _address_cache_key(self, street_input, match_mode):
        normalized = " ".join(str(street_input).split()).casefold()
        return f"{match_mode}|{normalized}"

    def _load_address_cache(self, match_mode):
        path = os.path.join(self.session_state_dir, ADDRESS_CACHE_FILE)
        if not os.path.exists(path):
            return {}
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, dict):
                return {}
            prefix = f"{match_mode}|"
            return {key: value for key, value in data.items() if key.startswith(prefix) and isinstance(value, dict)}
        except (OSError, json.JSONDecodeError):
            return {}

    def _save_address_cache(self, cache):
        path = os.path.join(self.session_state_dir, ADDRESS_CACHE_FILE)
        try:
            all_cache = {}
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    stored = json.load(f)
                if isinstance(stored, dict):
                    all_cache.update(stored)
            all_cache.update(cache)
            with open(path, "w", encoding="utf-8") as f:
                json.dump(all_cache, f, ensure_ascii=False)
            self.root.after(0, self._refresh_cache_info)
        except (OSError, json.JSONDecodeError) as e:
            self.log(f"    [!] Không lưu được cache Smarty: {e}")

    def _save_session_to_disk(self, collected_data, source_excel_path="", output_excel_path=""):
        """Ghi đè lại file phiên gần nhất trên đĩa. Lưu TOÀN BỘ collected_data (bao gồm cả
        các dòng đã 'true' lẫn 'suspect'/'false'), kèm trạng thái/lý do AI mới nhất, để lần
        'Kiểm tra lại' sau vẫn thấy đúng dữ liệu Smarty gốc + trạng thái mới nhất đã có."""
        try:
            payload = {
                "saved_at": time.strftime("%Y-%m-%d %H:%M:%S"),
                "row_count": len(collected_data),
                "source_excel_path": source_excel_path,
                "output_excel_path": output_excel_path,
                "collected_data": collected_data,
            }
            with open(self._session_file_path(), "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False)
        except Exception as e:
            # Lỗi lưu phiên KHÔNG được phép làm hỏng luồng chính (người dùng vẫn đã có Excel
            # kết quả rồi) - chỉ log lại để họ biết tính năng "Kiểm tra lại" có thể không dùng
            # được ở lần mở tool tiếp theo.
            self.log(f"    [!] Không lưu được phiên làm việc để kiểm tra lại sau này: {e}")

    def _load_session_from_disk(self):
        path = self._session_file_path()
        if not os.path.exists(path):
            return None
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, dict) or not isinstance(data.get("collected_data"), list):
                return None
            if not data["collected_data"]:
                return None
            return data
        except Exception:
            return None

    def _refresh_recheck_button(self):
        """Dò lại file phiên trên đĩa và cập nhật hiển thị nút 'Kiểm tra lại'. An toàn để gọi
        từ luồng nền (thread xử lý) thông qua root.after, hoặc trực tiếp từ luồng GUI chính."""
        session = self._load_session_from_disk()
        if session:
            n = session.get("row_count", len(session.get("collected_data", [])))
            self.btn_recheck.config(text=f"Kiểm tra lại ({n} dòng)", state=tk.NORMAL)
            if not self.btn_recheck.winfo_ismapped():
                self.btn_recheck.pack(side="left", padx=(0, 10))
            info_bits = [f"Phiên gần nhất lưu lúc {session.get('saved_at', '?')}"]
            src = session.get("source_excel_path", "")
            if src:
                info_bits.append(f"nguồn: {os.path.basename(src)}")
            self.lbl_recheck_info.config(text=" - ".join(info_bits))
            if not self.lbl_recheck_info.winfo_ismapped():
                self.lbl_recheck_info.pack(side="left")
        else:
            self.btn_recheck.pack_forget()
            self.lbl_recheck_info.pack_forget()

    def start_recheck_session(self):
        """Xử lý khi bấm nút 'Kiểm tra lại (<n> dòng)'. Mở lại đúng cửa sổ dán Prompt cho AI
        (giống hệt luồng thủ công lúc chạy chính), cho phép người dùng hỏi AI/GPT lại từ đầu
        (vd bù các dòng bị GPT bỏ sót/cắt bớt ở lần trước) rồi dán JSON mới để xuất lại Excel,
        MÀ KHÔNG cần gọi lại Smarty API cho các dòng đó."""
        session = self._load_session_from_disk()
        if not session:
            messagebox.showinfo("Không có phiên cũ", "Không tìm thấy phiên làm việc cũ nào để kiểm tra lại.")
            self._refresh_recheck_button()
            return

        collected_data = session["collected_data"]
        n = len(collected_data)

        proceed = messagebox.askyesno(
            "Kiểm tra lại phiên cũ",
            f"Phiên cũ có {n} dòng (đã gọi Smarty API từ trước - sẽ KHÔNG gọi lại Smarty).\n\n"
            f"Tool sẽ mở lại cửa sổ để bạn dán Prompt qua AI ngoài (ChatGPT/Gemini...) và dán "
            f"JSON kết quả mới vào, sau đó xuất lại 1 file Excel mới.\n\nBạn có muốn tiếp tục không?"
        )
        if not proceed:
            return

        save_path = filedialog.asksaveasfilename(
            title="Lưu kết quả kiểm tra lại",
            defaultextension=".xlsx",
            initialfile="KetQua_Smarty_KiemTraLai",
        )
        if not save_path:
            return
        if not save_path.lower().endswith(".xlsx"):
            save_path += ".xlsx"

        self.btn_recheck.config(state=tk.DISABLED)
        self.btn_start.config(state=tk.DISABLED)
        self.btn_select_excel.config(state=tk.DISABLED)

        threading.Thread(
            target=self._recheck_session_worker,
            args=(collected_data, save_path, session.get("source_excel_path", "")),
            daemon=True,
        ).start()

    def _recheck_session_worker(self, collected_data, save_path, source_excel_path):
        try:
            self.log("\n===========================================")
            self.log(f"[KIỂM TRA LẠI] Đang mở lại {len(collected_data)} dòng đã gọi Smarty trước đó "
                      f"để hỏi AI lại (không gọi lại Smarty API)...")

            self.run_manual_flow(collected_data)

            self.log("\nĐang tạo lại file Excel...")
            final_path = self.export_excel(collected_data, save_path)
            self.log(f"-> Đã lưu Excel: {final_path}")

            # Ghi đè lại phiên trên đĩa với trạng thái/lý do AI MỚI NHẤT, để lần "Kiểm tra lại"
            # tiếp theo (nếu có) tiếp tục dựa trên kết quả vừa cập nhật này.
            self._save_session_to_disk(collected_data, source_excel_path=source_excel_path, output_excel_path=final_path)

            self.log("\n[THÀNH CÔNG] Đã kiểm tra lại và xuất file Excel mới!")
            messagebox.showinfo("Hoàn tất", f"Đã xuất lại thành công:\n{final_path}")
        except Exception as e:
            self.log(f"\n[LỖI] Kiểm tra lại thất bại: {e}")
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi kiểm tra lại:\n{e}")
        finally:
            self.root.after(0, self._reset_recheck_ui)

    def _reset_recheck_ui(self):
        self.btn_start.config(state=tk.NORMAL if self.excel_path else tk.DISABLED)
        self.btn_select_excel.config(state=tk.NORMAL)
        self._refresh_recheck_button()

    def stop_processing(self):
        self.stop_requested = True
        self.log("\n[HỆ THỐNG] Đang yêu cầu dừng tiến trình...")
        self.btn_stop.config(state=tk.DISABLED)

    def start_processing_thread(self):
        try:
            delay_val = float(self.delay_var.get())
            if delay_val < 0: raise ValueError
        except ValueError:
            messagebox.showwarning("Cảnh báo", "Độ trễ không hợp lệ!")
            return

        if self.use_ai_var.get() and self.ai_mode_var.get() == "auto":
            keys = [self.entry_ai_key.get().strip(), self.entry_groq_key.get().strip(), self.entry_openrouter_key.get().strip()]
            if not any(keys):
                messagebox.showwarning("Cảnh báo", "Chế độ Tự động cần ít nhất 1 API Key!\nHoặc chuyển sang chế độ Thủ công.")
                return

        save_path = filedialog.asksaveasfilename(
            title="Lưu kết quả",
            defaultextension="",
            initialfile="KetQua_Smarty",
        )
        if not save_path:
            return

        base_path = save_path.replace('.txt', '').replace('.xlsx', '')
        self.txt_output_path = base_path + ".txt"
        self.excel_output_path = base_path + ".xlsx"

        self.stop_requested = False
        self.btn_start.config(state=tk.DISABLED, text="Đang xử lý...")
        self.btn_select_excel.config(state=tk.DISABLED)
        self.btn_stop.config(state=tk.NORMAL)
        self.btn_clear_cache.config(state=tk.DISABLED)
        self.entry_delay.config(state=tk.DISABLED)
        self.rb_match_strict.config(state=tk.DISABLED)
        self.rb_match_enhanced.config(state=tk.DISABLED)
        self.chk_use_ai.config(state=tk.DISABLED)
        self._set_ai_inputs_state(tk.DISABLED)

        threading.Thread(target=self.process_data, daemon=True).start()

    def process_data(self):
        collected_data = []
        unresolved_error = False

        # Bắt đầu 1 phiên gọi Smarty mới -> reset lại trạng thái xoay vòng key/proxy (cooldown)
        # về sạch, vì người dùng có thể đã bổ sung thêm key mới/đợi qua giới hạn từ lần trước.
        self.key_status = {}
        self.proxy_status = {}
        self._key_rr_counter = 0
        self._proxy_rr_counter = 0
        self._warned_single_key_session = False

        try:
            delay_seconds = float(self.delay_var.get())
            self.log("Đang đọc dữ liệu từ file Excel...")
            df = pd.read_excel(self.excel_path)

            address_cols = ['Shipping Address1', 'Shipping Address2', 'Shipping City', 'Shipping State', 'Shipping PostalCode', 'Shipping Country']
            ref_col = 'Shipment Reference Id'
            required_cols = address_cols + [ref_col]

            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                messagebox.showerror("Lỗi dữ liệu", f"File thiếu các cột:\n{', '.join(missing_cols)}")
                self.reset_ui()
                return

            total_rows = len(df)
            match_mode = self.match_mode_var.get()
            self.log(f"Chế độ khớp Smarty: {match_mode.upper()}" + (" (khuyến khích)" if match_mode == "strict" else ""))

            self.log(f"Bắt đầu giai đoạn 1: Gọi API Smarty ({total_rows} dòng)...")

            # Một file thường có nhiều dòng dùng chung một địa chỉ. Cache trong phiên giúp
            # mỗi chuỗi địa chỉ chỉ gọi Smarty tối đa một lần, kể cả sau khi mở lại tool.
            smarty_cache = self._load_address_cache(match_mode)
            cache_dirty = False
            new_cache_count = 0

            f_handle = self._open_output_txt_with_retry(self.txt_output_path)
            self.txt_output_path = f_handle.name
            with f_handle as f_out:
                for index, row in df.iterrows():
                    if self.stop_requested:
                        self.log("\n[HỆ THỐNG] Tiến trình dừng bởi người dùng.")
                        break

                    ref_id = str(row[ref_col]).strip() if pd.notna(row[ref_col]) else "NO_REF"
                    # QUAN TRỌNG: KHÔNG được biến đổi/giải mã dữ liệu đầu vào từ Excel dưới bất kỳ
                    # hình thức nào (kể cả HTML entity như "&#39;", "&amp;"...). Dữ liệu gốc phải
                    # được giữ nguyên 100% và gửi y hệt như vậy cho API, để đảm bảo tính trung thực
                    # và nhất quán với đúng những gì có trong file Excel gốc của người dùng.
                    parts = [str(row[col]).strip() for col in address_cols if pd.notna(row[col]) and str(row[col]).strip() != ""]
                    input_string = " ".join(parts)

                    # Giữ riêng giá trị GỐC của 5 cột (không qua Smarty, không qua AI) để chạy lớp
                    # lọc định dạng cuối cùng bên dưới.
                    addr1_raw = str(row['Shipping Address1']).strip() if pd.notna(row['Shipping Address1']) else ""
                    addr2_raw = str(row['Shipping Address2']).strip() if pd.notna(row['Shipping Address2']) else ""
                    city_raw = str(row['Shipping City']).strip() if pd.notna(row['Shipping City']) else ""
                    state_raw = str(row['Shipping State']).strip() if pd.notna(row['Shipping State']) else ""
                    postal_raw = str(row['Shipping PostalCode']).strip() if pd.notna(row['Shipping PostalCode']) else ""

                    if not input_string:
                        continue

                    cache_key = self._address_cache_key(input_string, match_mode)
                    used_cache = False
                    if cache_key in smarty_cache:
                        outcome = smarty_cache[cache_key].copy()
                        used_cache = True
                    else:
                        outcome = self.call_smarty_api_with_retry(input_string, 3, index+1)
                        request_proxy = outcome.pop("_proxy_url", None)
                        if outcome.get("stopped"):
                            self.log("[DỪNG] Người dùng đã dừng trước khi tất cả địa chỉ có kết quả.")
                            break
                        if outcome.get("status_code") != 200:
                            unresolved_error = True
                            self.log(
                                f"[KHÔNG XUẤT EXCEL] Dòng {index + 1} chưa có phản hồi HTTP 200 "
                                f"(HTTP {outcome.get('status_code')})."
                            )
                            break
                        if outcome.get("status_code") == 200:
                            smarty_cache[cache_key] = outcome.copy()
                            cache_dirty = True
                            new_cache_count += 1
                            if new_cache_count % 50 == 0:
                                self._save_address_cache(smarty_cache)
                    result_string = outcome.get("result", "")
                    analysis = outcome.get("analysis", {}) or {}

                    dpv_match_code = analysis.get("dpv_match_code", "")
                    dpv_vacant = analysis.get("dpv_vacant", "")
                    dpv_cmra = analysis.get("dpv_cmra", "")
                    dpv_no_stat = analysis.get("dpv_no_stat", "")
                    active = analysis.get("active", "")
                    # Trước đây 2 trường này bị GỘP LẪN vào 1 biến "footnotes" duy nhất (và không
                    # hề được xuất ra Excel) -> không ai kiểm chứng được vì sao 1 dòng bị/không bị
                    # gắn cờ. Giờ tách riêng, xuất cả 2 raw fields ra Excel để minh bạch/debug.
                    dpv_footnotes_raw = analysis.get("dpv_footnotes", "")
                    footnotes_raw = analysis.get("footnotes", "")

                    analysis_reasons = self._compute_analysis_flags(analysis)
                    # Kiểm tra ĐỘC LẬP trên chính chuỗi input gốc (chưa qua Smarty). Cộng dồn với
                    # analysis_reasons: dù Smarty (đặc biệt match=enhanced) có tự "vá" và trả về
                    # dpv_match_code=Y cho input bẩn, dòng này vẫn phải bị đưa vào diện nghi ngờ vì
                    # bản thân dữ liệu nguồn có vấn đề (vd còn dính "&#39;" - mã HTML entity).
                    input_anomaly_reasons = self._detect_input_anomalies(input_string)
                    # Lớp lọc ĐỊNH DẠNG CUỐI CÙNG trên 5 cột gốc - chạy TRƯỚC khi qua AI (thủ công
                    # và tự động), độc lập với cả Smarty lẫn AI.
                    format_reasons = self._detect_column_format_issues(addr1_raw, addr2_raw, city_raw, state_raw, postal_raw)
                    all_reasons = analysis_reasons + input_anomaly_reasons + format_reasons

                    analysis_summary = "; ".join(analysis_reasons) if analysis_reasons else (
                        "Không có cảnh báo" if analysis else "Không có dữ liệu analysis"
                    )
                    if input_anomaly_reasons:
                        analysis_summary += " | " + "; ".join(input_anomaly_reasons)
                    if format_reasons:
                        analysis_summary += " | " + "; ".join(format_reasons)

                    f_out.write(f"({input_string} : {result_string} | Analysis: {analysis_summary} [{ref_id}])\n")
                    f_out.flush()

                    item = {
                        "Reference Id": ref_id,
                        "Chuỗi đầu vào": input_string,
                        "Chuỗi đầu ra": result_string,
                        "DPV Match Code": dpv_match_code,
                        "DPV Vacant": dpv_vacant,
                        "DPV CMRA": dpv_cmra,
                        "DPV No Stat": dpv_no_stat,
                        "Active": active,
                        "DPV Footnotes": dpv_footnotes_raw,
                        "Footnotes": footnotes_raw,
                        # Tự động đánh dấu "suspect" ngay từ bước gọi API (kể cả khi TẮT AI),
                        # dựa trên dữ liệu "analysis" của Smarty VÀ chất lượng của chính input gốc,
                        # không chỉ dựa vào so khớp chuỗi hay việc Smarty có match được hay không.
                        "Trạng thái xử lý": "suspect" if all_reasons else "",
                        "Lý do đáng ngờ": ("Cảnh báo: " + "; ".join(all_reasons)) if all_reasons else "",
                    }
                    item["_analysis_reasons"] = all_reasons
                    collected_data.append(item)

                    if used_cache:
                        self.log(f"Smarty API - Dòng {index + 1}/{total_rows} -> dùng kết quả cache (không tạo request)")
                    elif outcome.get("success"):
                        self.log(
                            f"Smarty API - Dòng {index + 1}/{total_rows} -> OK | "
                            f"{self._request_route_label(request_proxy)}"
                        )
                    else:
                        self.log(
                            f"Smarty API - Dòng {index + 1}/{total_rows} -> LỖI: {result_string} | "
                            f"{self._request_route_label(request_proxy)}"
                        )

                    # Chỉ delay nếu không dùng proxy hoặc đang chạy luồng bình thường
                    if delay_seconds > 0:
                        time.sleep(delay_seconds)

                if cache_dirty:
                    self._save_address_cache(smarty_cache)

                if not self.stop_requested:
                    gpt_prompt = """
=======================================================================================================================
[SYSTEM PROMPT - DÀNH CHO AI/CHATGPT/GEMINI]
Nhiệm vụ của bạn là đóng vai trò chuyên gia kiểm duyệt địa chỉ quốc tế khắt khe. Dựa vào danh sách đối chiếu bên trên (định dạng: (chuỗi gốc : chuỗi chuẩn hóa từ API | Analysis: <cảnh báo DPV nếu có> [Mã đơn hàng])), hãy thực hiện:

1. QUÉT TOÀN BỘ danh sách và LỌC RA những đơn hàng CÓ DẤU HIỆU ĐÁNG NGỜ hoặc LỖI.
2. Tiêu chí bắt LỖI (out không hợp lệ):
   - Trả về "Không tìm thấy kết quả từ Server", "Lỗi HTTP...", "Lỗi API...".
   - Trả về "Không tìm thấy delivery_line_1 hoặc last_line".
   - Chuỗi trả về bị thiếu quá nhiều thành phần trọng yếu so với chuỗi gốc (mất hẳn số nhà, tên đường khác hoàn toàn, sai khác mã ZIP nghiêm trọng).
3. Tiêu chí bắt NGHI NGỜ (out khớp chuỗi nhưng vẫn cần xem lại) — QUAN TRỌNG, đừng chỉ so khớp chuỗi:
   - Phần "Analysis" đi kèm có cảnh báo, ví dụ: địa chỉ đang bị BỎ TRỐNG (vacant), là hộp thư CMRA, không có dữ liệu xác thực (no_stat), địa chỉ KHÔNG hoạt động (active=N), hoặc mã khớp DPV chưa xác nhận đầy đủ.
   - Những trường hợp này KHÔNG được coi là "hợp lệ" chỉ vì chuỗi địa chỉ khớp — phải liệt kê vào diện đáng ngờ.
4. ĐẦU RA BẮT BUỘC TRÌNH BÀY DƯỚI DẠNG BẢNG (Markdown Table) với các cột sau:
   | Mã đơn hàng (ID) | Chuỗi đầu vào (Gốc) | Chuỗi đầu ra (API trả về) | Loại (Lỗi/Nghi ngờ) | Lý do chi tiết |

TUYỆT ĐỐI KHÔNG đưa các đơn hàng hoàn toàn hợp lệ (không lỗi, không cảnh báo Analysis) vào bảng này. Nếu tất cả đều hợp lệ, hãy trả lời ngắn gọn: "Tất cả địa chỉ đều hợp lệ."
=======================================================================================================================
"""
                    f_out.write(gpt_prompt)
                    f_out.flush()

            if unresolved_error or self.stop_requested:
                self.log(
                    "\n[CHƯA HOÀN TẤT] Không xuất Excel vì vẫn còn dòng chưa nhận được kết quả rõ ràng. "
                    "Hãy sửa kết nối/API rồi chạy lại; các dòng lỗi không được đưa vào file kết quả."
                )
                return

            if self.use_ai_var.get() and not self.stop_requested and collected_data:
                self.log("\n===========================================")
                self.log("Bắt đầu giai đoạn 2: Kiểm tra kết quả đáng ngờ bằng AI...")

                mode = self.ai_mode_var.get()
                if mode == "manual":
                    self.log("Chế độ: THỦ CÔNG (bạn tự dán Prompt vào AI ngoài).")
                    self.run_manual_flow(collected_data)
                else:
                    self.log("Chế độ: TỰ ĐỘNG (hệ thống tự dò & gọi AI miễn phí).")
                    self.run_auto_flow(collected_data)

            elif not self.use_ai_var.get():
                self.log("\n[INFO] AI tự động đang TẮT. File TXT đã có Prompt để bạn tự check thủ công.")

            if collected_data:
                self.log("\nBắt đầu tạo file Excel...")
                self.excel_output_path = self.export_excel(collected_data, self.excel_output_path)
                self.log(f"-> Đã lưu Excel: {self.excel_output_path}")

                # Lưu lại TOÀN BỘ phiên này (các dòng đã gọi xong Smarty + trạng thái AI hiện
                # tại, nếu có) xuống đĩa. Nhờ vậy, dù đóng hẳn tool đi, người dùng vẫn có thể mở
                # lại và bấm nút "Kiểm tra lại" để hỏi AI lại/dán JSON mới, KHÔNG cần gọi lại
                # Smarty API cho các dòng đã có kết quả rồi.
                self._save_session_to_disk(
                    collected_data,
                    source_excel_path=self.excel_path,
                    output_excel_path=self.excel_output_path,
                )
                self.root.after(0, self._refresh_recheck_button)

            if not self.stop_requested:
                self.log(f"\n[THÀNH CÔNG] Toàn bộ tiến trình hoàn tất!")
                messagebox.showinfo("Hoàn tất", f"Đã xuất thành công:\n1. {self.txt_output_path}\n2. {self.excel_output_path}")

        except Exception as e:
            self.log(f"\n[LỖI] Đã xảy ra sự cố: {e}")
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra:\n{e}")
        finally:
            self.reset_ui()

    def run_auto_flow(self, collected_data):
        BATCH_SIZE = 20
        n = len(collected_data)
        self.log(f"Tổng {n} dòng -> chia thành các batch {BATCH_SIZE} dòng/lượt gọi AI.")

        idx = 0
        while idx < n:
            if self.stop_requested:
                return
            batch = collected_data[idx: idx + BATCH_SIZE]
            batch_no = idx // BATCH_SIZE + 1
            total_batches = (n + BATCH_SIZE - 1) // BATCH_SIZE
            self.log(f"-> Đang thử gọi AI tự động cho batch {batch_no}/{total_batches} ({len(batch)} dòng)...")

            results_map = self.check_with_ai_batch_auto(batch)

            if results_map is None:
                self.log("    [!] Không có nhà cung cấp AI miễn phí nào phản hồi thành công cho batch này.")
                remaining_items = collected_data[idx:]
                want_manual = self.ask_switch_to_manual(
                    f"Hệ thống không gọi được AI tự động.\n"
                    f"Còn {len(remaining_items)} dòng chưa được kiểm tra."
                )
                if want_manual:
                    self.log("-> Người dùng chọn chuyển sang chế độ THỦ CÔNG cho các dòng còn lại.")
                    self.run_manual_flow(remaining_items)
                else:
                    self.log("-> Người dùng chọn bỏ qua kiểm tra AI cho các dòng còn lại.")
                    for item in remaining_items:
                        self._apply_ai_result_with_analysis(
                            item, None,
                            missing_reason="Bỏ qua kiểm tra AI (không gọi được AI tự động)"
                        )
                return

            for item in batch:
                rid = item["Reference Id"]
                res = results_map.get(rid)
                self._apply_ai_result_with_analysis(item, res)

            self.log(f"-> Batch {batch_no}/{total_batches} hoàn tất.")
            idx += BATCH_SIZE
            if idx < n:
                time.sleep(1.0)

    def check_with_ai_batch_auto(self, items):
        prompt = self._build_ai_prompt(items)
        raw = self.call_any_available_ai(prompt)
        if raw is None:
            return None
        return self._parse_ai_json_array(raw)

    def call_any_available_ai(self, prompt):
        gemini_key = self.entry_ai_key.get().strip()
        groq_key = self.entry_groq_key.get().strip()
        openrouter_key = self.entry_openrouter_key.get().strip()

        providers = []
        if gemini_key and HAS_GEMINI:
            providers.append(("Gemini", lambda: self._call_gemini_fast(prompt, gemini_key)))
        if groq_key:
            providers.append(("Groq", lambda: self._call_groq(prompt, groq_key)))
        if openrouter_key:
            providers.append(("OpenRouter", lambda: self._call_openrouter(prompt, openrouter_key)))

        if not providers:
            self.log("    [!] Chưa nhập bất kỳ API Key nào cho chế độ Tự động.")
            return None

        for name, fn in providers:
            if self.stop_requested:
                return None
            self.log(f"    -> Đang thử {name}...")
            try:
                raw = fn()
                if raw and raw.strip():
                    self.log(f"    [OK] {name} phản hồi thành công.")
                    return raw
                self.log(f"    [!] {name} trả về rỗng, thử provider kế tiếp...")
            except Exception as e:
                self.log(f"    [!] {name} lỗi: {str(e)[:120]}")
                continue

        return None

    def _call_gemini_fast(self, prompt, api_key):
        client = genai.Client(api_key=api_key)
        models = self._quick_discover_gemini_models(client) or GEMINI_FALLBACK_MODELS
        last_err = None

        for model_name in models[:2]:
            for attempt in range(2):
                if self.stop_requested:
                    return None
                try:
                    response = client.models.generate_content(
                        model=model_name,
                        contents=prompt,
                        config=genai_types.GenerateContentConfig(
                            temperature=0.0,
                            response_mime_type="application/json"
                        )
                    )
                    return (response.text or "").strip()
                except Exception as e:
                    last_err = e
                    es = str(e).lower()
                    if "404" in es or "not found" in es:
                        break
                    if "429" in es or "503" in es or "overloaded" in es or "unavailable" in es or "quota" in es:
                        if attempt == 0:
                            time.sleep(3)
                            continue
                        break
                    break

        if last_err:
            raise last_err
        raise RuntimeError("Gemini không khả dụng")

    def _quick_discover_gemini_models(self, client):
        try:
            discovered = []
            for m in client.models.list():
                name = getattr(m, "name", "") or ""
                short = name.split("/")[-1] if "/" in name else name
                if not short:
                    continue
                actions = getattr(m, "supported_actions", None) or []
                if actions and "generateContent" not in actions:
                    continue
                low = short.lower()
                if any(x in low for x in ["embedding", "tts", "image", "video", "vision", "aqa", "gemma"]):
                    continue
                if "flash" not in low and "pro" not in low:
                    continue
                discovered.append(short)

            def priority(name):
                n = name.lower()
                if "flash-lite" in n:
                    return 1
                if n.endswith("-latest") and "flash" in n:
                    return 0
                if "flash" in n:
                    return 2
                if "pro" in n:
                    return 3
                return 4

            discovered = sorted(set(discovered), key=priority)
            return discovered
        except Exception:
            return []

    def _call_groq(self, prompt, api_key):
        url = "https://api.groq.com/openai/v1/chat/completions"
        headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
        last_err = None
        for model_name in GROQ_MODELS[:2]:
            body = {
                "model": model_name,
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0
            }
            try:
                r = self.session.post(url, headers=headers, json=body, timeout=20)
                if r.status_code == 200:
                    data = r.json()
                    return data["choices"][0]["message"]["content"]
                elif r.status_code == 404:
                    last_err = RuntimeError(f"Model '{model_name}' không tồn tại trên Groq")
                    continue
                elif r.status_code == 429:
                    last_err = RuntimeError("Groq quá tải/hết quota (429)")
                    time.sleep(2)
                    continue
                else:
                    last_err = RuntimeError(f"Groq lỗi HTTP {r.status_code}: {r.text[:150]}")
                    continue
            except requests.exceptions.RequestException as e:
                last_err = e
                continue
        raise last_err or RuntimeError("Groq thất bại")

    def _call_openrouter(self, prompt, api_key):
        url = "https://openrouter.ai/api/v1/chat/completions"
        headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
        last_err = None
        for model_name in OPENROUTER_MODELS:
            body = {
                "model": model_name,
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0
            }
            try:
                r = self.session.post(url, headers=headers, json=body, timeout=20)
                if r.status_code == 200:
                    data = r.json()
                    return data["choices"][0]["message"]["content"]
                elif r.status_code == 404:
                    last_err = RuntimeError(f"Model '{model_name}' không tồn tại trên OpenRouter")
                    continue
                elif r.status_code == 429:
                    last_err = RuntimeError("OpenRouter quá tải/hết quota (429)")
                    time.sleep(2)
                    continue
                else:
                    last_err = RuntimeError(f"OpenRouter lỗi HTTP {r.status_code}: {r.text[:150]}")
                    continue
            except requests.exceptions.RequestException as e:
                last_err = e
                continue
        raise last_err or RuntimeError("OpenRouter thất bại")

    def run_manual_flow(self, items):
        if not items:
            return
        self.log(f"[THỦ CÔNG] Đang tạo Prompt cho {len(items)} dòng để bạn sao chép...")
        prompt = self._build_ai_prompt(items)

        json_text = self.get_manual_json_from_user(prompt, len(items))

        if not json_text:
            self.log("    [!] Bạn đã bỏ qua bước nhập JSON thủ công -> giữ nguyên cờ Analysis (nếu có), còn lại 'unknown'.")
            for item in items:
                self._apply_ai_result_with_analysis(
                    item, None,
                    missing_reason="Bỏ qua kiểm tra AI thủ công"
                )
            return

        results_map = self._parse_ai_json_array(json_text)
        if not results_map:
            self.log("    [!] Không đọc được JSON hợp lệ từ nội dung bạn dán vào -> các dòng này sẽ ở trạng thái 'unknown'.")

        for item in items:
            rid = item["Reference Id"]
            res = results_map.get(rid)
            if res is None:
                self._apply_ai_result_with_analysis(item, None, missing_reason="Không thấy id này trong JSON đã nhập")
            else:
                self._apply_ai_result_with_analysis(item, res)

        self.log(f"    -> Đã áp dụng kết quả thủ công cho {len(items)} dòng.")

    def _apply_ai_result_with_analysis(self, item, res, missing_reason="Không nhận được kết quả từ AI cho dòng này"):
        """Gộp kết quả AI (true/suspect/false) với cờ nghi ngờ tính sẵn từ 'analysis'
        của Smarty (dpv_vacant, dpv_cmra, dpv_no_stat, active, dpv_match_code...).
        Nguyên tắc: nếu analysis đã cảnh báo, AI KHÔNG được phép hạ xuống 'true'."""
        analysis_reasons = item.get("_analysis_reasons", [])
        analysis_text = "; ".join(analysis_reasons)

        if res is None:
            if analysis_reasons:
                # Đã có cờ nghi ngờ từ analysis từ trước -> giữ nguyên "suspect", chỉ bổ sung ghi chú
                base_reason = item.get("Lý do đáng ngờ", "") or f"Cảnh báo Analysis: {analysis_text}"
                item["Trạng thái xử lý"] = "suspect"
                item["Lý do đáng ngờ"] = f"{base_reason} | {missing_reason}"
            else:
                item["Trạng thái xử lý"] = "unknown"
                item["Lý do đáng ngờ"] = missing_reason
            return

        ai_status = res.get("status", "unknown")
        ai_reason = (res.get("reason", "") or "").strip()

        if not analysis_reasons:
            item["Trạng thái xử lý"] = ai_status
            item["Lý do đáng ngờ"] = ai_reason
            return

        # Có cảnh báo từ analysis: AI nói "false" -> vẫn giữ "false" (nặng hơn).
        # AI nói "true" hoặc "suspect" -> luôn là "suspect" tối thiểu, không được là "true".
        if ai_status == "false":
            final_status = "false"
        else:
            final_status = "suspect"

        reason_parts = []
        if ai_reason:
            reason_parts.append(f"AI: {ai_reason}")
        reason_parts.append(f"Cảnh báo Analysis: {analysis_text}")
        item["Trạng thái xử lý"] = final_status
        item["Lý do đáng ngờ"] = " | ".join(reason_parts)

    def _ask_retry_cancel(self, title, message):
        """Hỏi người dùng 'Thử lại' / 'Huỷ' theo kiểu thread-safe (an toàn khi gọi từ luồng xử lý
        nền process_data, vì messagebox phải được tạo trên luồng GUI chính)."""
        result_holder = {"value": False}
        event = threading.Event()

        def _build():
            try:
                result_holder["value"] = messagebox.askretrycancel(title, message)
            finally:
                event.set()

        self.root.after(0, _build)
        event.wait()
        return result_holder["value"]

    def _alt_path_with_timestamp(self, path):
        """Sinh ra 1 đường dẫn khác (thêm hậu tố thời gian) để lưu dự phòng khi không thể ghi
        đè vào file gốc, tránh làm mất toàn bộ dữ liệu đã xử lý."""
        base, ext = os.path.splitext(path)
        return f"{base}_{time.strftime('%Y%m%d_%H%M%S')}{ext}"

    def _open_output_txt_with_retry(self, path):
        """Mở file TXT kết quả ở chế độ ghi (tạo mới/ghi đè). Nếu file đang bị một chương trình
        khác (Excel, Notepad, v.v.) mở và khoá nên không ghi được, hỏi người dùng:
        - 'Retry': thử mở lại (sau khi họ đã đóng file đó).
        - 'Cancel': tự động đổi sang tên file khác (kèm timestamp) để không mất dữ liệu."""
        current = path
        while True:
            try:
                return open(current, 'w', encoding='utf-8')
            except PermissionError:
                self.log(f"  [!] Không thể ghi file TXT '{os.path.basename(current)}' - có thể đang mở ở chương trình khác.")
                retry = self._ask_retry_cancel(
                    "Không thể ghi file TXT",
                    f"File:\n{current}\n\ncó thể đang được MỞ trong một chương trình khác (Excel, Notepad, v.v.) "
                    f"nên không thể ghi vào.\n\nBấm 'Retry' SAU KHI bạn đã đóng file đó lại.\n"
                    f"Bấm 'Cancel' để tự động lưu sang một tên file khác (không mất dữ liệu)."
                )
                if retry:
                    continue
                current = self._alt_path_with_timestamp(path)
                self.log(f"  [i] Sẽ lưu file TXT sang tên khác: {current}")

    def _save_workbook_with_retry(self, wb, desired_path):
        """Lưu workbook openpyxl vào desired_path. Đây chính là chỗ hay gặp lỗi 'PermissionError'
        khi file Excel kết quả đang được người dùng MỞ SẴN (trùng tên với file họ định lưu).
        Thay vì crash toàn bộ tiến trình và bắt người dùng tự đóng file rồi chạy lại từ đầu, ở đây
        sẽ hỏi:
        - 'Retry': thử lưu lại (sau khi người dùng đã đóng file Excel đang mở đó).
        - 'Cancel': tự động lưu kết quả sang một tên file khác (kèm timestamp), không mất dữ liệu
          đã xử lý (đặc biệt quan trọng vì bước này diễn ra SAU KHI đã gọi API + AI xong)."""
        path = desired_path
        while True:
            try:
                wb.save(path)
                return path
            except PermissionError:
                self.log(f"  [!] Không thể lưu Excel '{os.path.basename(path)}' - file đang được MỞ ở chương trình khác.")
                retry = self._ask_retry_cancel(
                    "Không thể lưu file Excel",
                    f"File:\n{path}\n\ncó thể đang được MỞ trong Excel (hoặc chương trình khác) nên không thể "
                    f"ghi đè lên được.\n\nHãy ĐÓNG file đó lại, sau đó bấm 'Retry' để lưu lại.\n"
                    f"Bấm 'Cancel' để tự động lưu kết quả sang một tên file khác (không mất dữ liệu đã xử lý)."
                )
                if retry:
                    continue
                path = self._alt_path_with_timestamp(desired_path)
                self.log(f"  [i] Sẽ lưu file Excel sang tên khác: {path}")
            except OSError as e:
                # Các lỗi OS khác (đĩa đầy, đường dẫn không hợp lệ...) cũng không nên làm mất
                # toàn bộ dữ liệu đã xử lý -> vẫn cho người dùng cơ hội thử lại / đổi tên.
                self.log(f"  [!] Lỗi khi lưu file Excel: {e}")
                retry = self._ask_retry_cancel(
                    "Không thể lưu file Excel",
                    f"Đã xảy ra lỗi khi lưu file:\n{path}\n\nLỗi: {e}\n\n"
                    f"Bấm 'Retry' để thử lưu lại.\nBấm 'Cancel' để tự động lưu sang một tên file khác."
                )
                if retry:
                    continue
                path = self._alt_path_with_timestamp(desired_path)
                self.log(f"  [i] Sẽ lưu file Excel sang tên khác: {path}")

    def ask_switch_to_manual(self, reason_text):
        result_holder = {"value": False}
        event = threading.Event()

        def _build():
            try:
                ans = messagebox.askyesno(
                    "Không gọi được AI tự động",
                    f"{reason_text}\n\nBạn có muốn chuyển sang chế độ THỦ CÔNG "
                    f"(tự dán Prompt vào AI ngoài, rồi dán JSON kết quả về) không?"
                )
                result_holder["value"] = bool(ans)
            finally:
                event.set()

        self.root.after(0, _build)
        event.wait()
        return result_holder["value"]

    def get_manual_json_from_user(self, prompt_text, item_count):
        result_holder = {"value": None}
        event = threading.Event()

        def _build():
            dlg = tk.Toplevel(self.root)
            dlg.title("Chế độ thủ công - Dán Prompt vào AI ngoài")
            dlg.geometry("720x650")
            dlg.grab_set()

            # Tự động mở sẵn 1 tab trình duyệt tới ChatGPT ngay khi vào giai đoạn này, để người
            # dùng chỉ cần bấm "Sao chép Prompt" rồi qua tab đó dán vào, không cần tự mở tay.
            # Best-effort: nếu máy không có trình duyệt mặc định hoặc mở thất bại thì bỏ qua,
            # không làm gián đoạn luồng thủ công (người dùng vẫn có thể tự mở trình duyệt).
            try:
                webbrowser.open("https://chatgpt.com/")
            except Exception as e:
                self.log(f"    [!] Không tự mở được trình duyệt ChatGPT: {e}")

            tk.Label(
                dlg,
                text=(f"Có {item_count} dòng cần kiểm tra.\n"
                      "Đã tự mở sẵn 1 tab trình duyệt ChatGPT VÀ tự sao chép Prompt vào Clipboard cho bạn.\n"
                      "1) Qua tab ChatGPT (hoặc AI bất kỳ khác: Gemini web, Claude...), dán (Ctrl+V) là được. "
                      "(Nút 'Sao chép Prompt' bên dưới vẫn dùng được để bấm lại nếu cần.)\n"
                      "2) Sao chép TOÀN BỘ phần JSON mà AI trả về, dán vào ô phía dưới.\n"
                      "3) Bấm 'Xác nhận' để tool tự tạo Excel."),
                justify="left", anchor="w"
            ).pack(anchor="w", padx=10, pady=(10, 5))

            tk.Label(dlg, text="Prompt để dán vào AI ngoài:", font=("Arial", 9, "bold")).pack(anchor="w", padx=10)
            txt_prompt = scrolledtext.ScrolledText(dlg, height=14, wrap="word", font=("Consolas", 9))
            txt_prompt.insert("1.0", prompt_text)
            txt_prompt.pack(fill="both", expand=True, padx=10, pady=(0, 5))

            def copy_prompt():
                dlg.clipboard_clear()
                dlg.clipboard_append(prompt_text)
                lbl_copied.config(text="Đã sao chép vào Clipboard!", fg="#28a745")

            frame_copy = tk.Frame(dlg)
            frame_copy.pack(pady=2)
            tk.Button(frame_copy, text="Sao chép Prompt", command=copy_prompt, bg="#0078D7", fg="white", width=20).pack(side="left", padx=5)
            lbl_copied = tk.Label(frame_copy, text="", font=("Arial", 9))
            lbl_copied.pack(side="left", padx=5)

            # Tự động sao chép Prompt vào Clipboard ngay khi dialog mở ra, để người dùng đỡ phải
            # bấm nút thủ công - qua thẳng tab ChatGPT (đã tự mở ở trên) và dán (Ctrl+V) là được.
            # Nút "Sao chép Prompt" vẫn giữ nguyên để bấm lại bất cứ lúc nào (vd lỡ copy đè thứ
            # khác vào Clipboard).
            copy_prompt()

            tk.Label(dlg, text="Dán JSON kết quả từ AI vào đây:", font=("Arial", 9, "bold")).pack(anchor="w", padx=10, pady=(10, 0))
            txt_json = scrolledtext.ScrolledText(dlg, height=10, wrap="word", font=("Consolas", 9))
            txt_json.pack(fill="both", expand=True, padx=10, pady=(0, 5))

            def on_confirm():
                content = txt_json.get("1.0", "end").strip()
                result_holder["value"] = content if content else None
                dlg.grab_release()
                dlg.destroy()
                event.set()

            def on_skip():
                result_holder["value"] = None
                dlg.grab_release()
                dlg.destroy()
                event.set()

            frame_btn = tk.Frame(dlg)
            frame_btn.pack(pady=10)
            tk.Button(frame_btn, text="Xác nhận", command=on_confirm, bg="#28a745", fg="white", width=14, font=("Arial", 10, "bold")).pack(side="left", padx=5)
            tk.Button(frame_btn, text="Bỏ qua kiểm tra AI cho các dòng này", command=on_skip, bg="#dc3545", fg="white", width=32).pack(side="left", padx=5)

            dlg.protocol("WM_DELETE_WINDOW", on_skip)

        self.root.after(0, _build)
        event.wait()
        return result_holder["value"]

    def _build_ai_prompt(self, items):
        payload = [
            {
                "id": it["Reference Id"],
                "in": it["Chuỗi đầu vào"],
                "out": it["Chuỗi đầu ra"],
                "analysis": {
                    "dpv_match_code": it.get("DPV Match Code", ""),
                    "dpv_vacant": it.get("DPV Vacant", ""),
                    "dpv_cmra": it.get("DPV CMRA", ""),
                    "dpv_no_stat": it.get("DPV No Stat", ""),
                    "active": it.get("Active", ""),
                    "dpv_footnotes": it.get("DPV Footnotes", ""),
                    "footnotes": it.get("Footnotes", ""),
                },
            }
            for it in items
        ]

        prompt = f"""Bạn là chuyên gia kiểm duyệt địa chỉ vận chuyển quốc tế, làm việc rất khắt khe và chính xác.
Dưới đây là một danh sách JSON gồm nhiều đơn hàng. Mỗi phần tử có:
- "id": mã đơn hàng
- "in": chuỗi địa chỉ gốc
- "out": chuỗi địa chỉ đã chuẩn hóa từ API Smarty
- "analysis": dữ liệu xác thực DPV (Delivery Point Validation) của USPS đi kèm, gồm:
   + "dpv_match_code": "Y" = khớp hoàn toàn, đã xác nhận; "D"/"S"/"N" hoặc rỗng = CHƯA xác nhận đầy đủ.
   + "dpv_vacant": "Y" = USPS liệt kê địa chỉ này ĐANG BỊ BỎ TRỐNG (vacant) — rất đáng ngờ dù chuỗi khớp.
   + "dpv_cmra": "Y" = đây là hộp thư nhận hộ (CMRA), không phải địa chỉ nhà/công ty thực tế.
   + "dpv_no_stat": "Y" = USPS không có đủ dữ liệu để xác thực địa chỉ này.
   + "active": "N" = địa chỉ hiện KHÔNG còn hoạt động.

QUAN TRỌNG: Việc "out" khớp với "in" về mặt CHUỖI KÝ TỰ là CHƯA ĐỦ để kết luận địa chỉ chắc chắn đúng. Bạn PHẢI kết hợp cả phần "analysis" để đánh giá — đây chính là mục đích của việc kiểm tra này.

Với MỖI phần tử, hãy gán MỘT trong 3 trạng thái:
- status = "true": CHỈ khi "out" khớp tốt với "in" (đủ số nhà, tên đường, thành phố, mã ZIP) VÀ "analysis" không có cảnh báo nào (dpv_match_code="Y", dpv_vacant khác "Y", dpv_cmra khác "Y", dpv_no_stat khác "Y", active khác "N").
- status = "suspect": khi "out" khớp với "in" về chuỗi NHƯNG "analysis" có ít nhất một cảnh báo (dpv_vacant="Y", dpv_cmra="Y", dpv_no_stat="Y", active="N", hoặc dpv_match_code khác "Y"). Đây là các địa chỉ ĐƯỢC TÌM THẤY nhưng cần con người xem lại trước khi giao hàng — KHÔNG được gán "true" cho các trường hợp này.
- status = "false": khi "out" chứa các cụm lỗi như "Lỗi", "Không tìm thấy kết quả", "Không tìm thấy delivery_line_1", hoặc "out" thiếu/sai lệch nghiêm trọng so với "in" (mất số nhà, tên đường khác hẳn, sai ZIP nghiêm trọng).

YÊU CẦU ĐẦU RA (BẮT BUỘC):
- CHỈ trả về DUY NHẤT một mảng JSON hợp lệ, KHÔNG kèm markdown code fence, KHÔNG giải thích thêm, KHÔNG viết bất kỳ chữ nào trước hoặc sau mảng JSON.
- Số phần tử trong mảng trả về PHẢI khớp chính xác với số phần tử đầu vào (mỗi "id" xuất hiện đúng 1 lần).
- Định dạng mỗi phần tử: {{"id": "<id gốc>", "status": "true" | "suspect" | "false", "reason": "<lý do ngắn gọn; BẮT BUỘC điền khi status là 'suspect' hoặc 'false', để trống '' khi 'true'>"}}

Danh sách đầu vào (JSON):
{json.dumps(payload, ensure_ascii=False)}
"""
        return prompt

    def _parse_ai_json_array(self, raw_text):
        text = raw_text.strip()
        text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text.strip())
        try:
            data = json.loads(text)
        except json.JSONDecodeError:
            match = re.search(r"\[.*\]", text, re.DOTALL)
            if not match:
                self.log("    [!] Không parse được JSON trả về từ AI.")
                return {}
            try:
                data = json.loads(match.group(0))
            except json.JSONDecodeError:
                self.log("    [!] JSON trả về từ AI bị sai định dạng.")
                return {}

        result_map = {}
        if isinstance(data, list):
            for entry in data:
                if not isinstance(entry, dict) or "id" not in entry:
                    continue
                status = str(entry.get("status", "")).strip().lower()
                status = status if status in ("true", "false", "suspect") else "unknown"
                result_map[str(entry["id"])] = {
                    "status": status,
                    "reason": entry.get("reason", "") or ""
                }
        return result_map

    def export_excel(self, collected_data, output_path):
        for item in collected_data:
            item.setdefault("DPV Match Code", "")
            item.setdefault("DPV Vacant", "")
            item.setdefault("DPV CMRA", "")
            item.setdefault("DPV No Stat", "")
            item.setdefault("Active", "")
            item.setdefault("DPV Footnotes", "")
            item.setdefault("Footnotes", "")
            item.setdefault("Trạng thái xử lý", "")
            item.setdefault("Lý do đáng ngờ", "")

        # false (lỗi) đứng trước, kế đến suspect (nghi ngờ - cần xem lại analysis),
        # rồi unknown (chưa kiểm tra được), cuối cùng true (chắc chắn đúng).
        status_priority = {"false": 0, "suspect": 1, "unknown": 2, "": 2, "true": 3}
        sorted_data = sorted(
            collected_data,
            key=lambda x: status_priority.get(str(x.get("Trạng thái xử lý", "")).lower(), 2)
        )

        columns = [
            "Reference Id", "Chuỗi đầu vào", "Chuỗi đầu ra",
            "DPV Match Code", "DPV Vacant", "DPV CMRA", "DPV No Stat", "Active",
            "DPV Footnotes", "Footnotes",
            "Trạng thái xử lý", "Lý do đáng ngờ",
        ]
        out_df = pd.DataFrame(sorted_data, columns=columns)

        # QUAN TRỌNG: xử lý/định dạng dữ liệu ra 1 FILE TẠM trước (không đụng tới output_path lúc
        # này), để việc file đích đang bị khoá (đang mở sẵn ở Excel) không làm hỏng hay chặn mất
        # công đoạn xử lý dữ liệu đã tốn công gọi API/AI. Chỉ bước SAO CHÉP/LƯU vào output_path ở
        # cuối mới cần cơ chế retry khi bị khoá.
        tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(tmp_fd)
        try:
            out_df.to_excel(tmp_path, index=False)

            import openpyxl
            wb = openpyxl.load_workbook(tmp_path)
            ws = wb.active

            fill_false = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            font_false = Font(color="9C0006", bold=True)
            fill_suspect = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            font_suspect = Font(color="C65911", bold=True)
            fill_unknown = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            font_unknown = Font(color="9C6500")
            fill_true = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            font_true = Font(color="006100")

            status_col_idx = columns.index("Trạng thái xử lý") + 1

            for row_idx in range(2, ws.max_row + 1):
                status_val = str(ws.cell(row=row_idx, column=status_col_idx).value or "").strip().lower()
                if status_val == "false":
                    fill, font = fill_false, font_false
                elif status_val == "suspect":
                    fill, font = fill_suspect, font_suspect
                elif status_val == "true":
                    fill, font = fill_true, font_true
                else:
                    fill, font = fill_unknown, font_unknown

                for col_idx in range(1, len(columns) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = fill
                    if col_idx == status_col_idx:
                        cell.font = font

            for col_idx, col_name in enumerate(columns, start=1):
                max_len = max(
                    [len(str(col_name))] + [len(str(row.get(col_name, ""))) for row in sorted_data]
                )
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 60)

            ws.freeze_panes = "A2"

            # Bước lưu thực sự vào đích cuối - có retry nếu file đang bị khoá (đang mở ở chương
            # trình khác). Trả về đường dẫn thực tế đã lưu thành công (có thể khác output_path
            # nếu người dùng chọn đổi tên thay vì đóng file đang mở).
            final_path = self._save_workbook_with_retry(wb, output_path)
            return final_path
        finally:
            try:
                os.remove(tmp_path)
            except OSError:
                pass

    def call_smarty_api(self, street_input, identity):
        # match lấy từ lựa chọn của người dùng trên GUI (mặc định "strict" - giống hệt kết
        # quả khi nhập thủ công vào Smarty, chỉ khớp dựa hoàn toàn trên dữ liệu gốc, không tự
        # "vá"/suy đoán). Người dùng có thể tự đổi sang "enhanced" nếu muốn Smarty khớp mạnh
        # tay hơn, nhưng "strict" là lựa chọn được khuyến khích.
        match_mode = self.match_mode_var.get() if self.match_mode_var.get() in ("strict", "enhanced") else "strict"
        params = {
            "key": identity["api_key"], "agent": "smarty (website:demo)",
            "match": match_mode, "candidates": "5", "geocode": "true",
            "license": "us-rooftop-geocoding-cloud", "street": street_input
        }

        proxy_url = identity.get("proxy_url")
        proxies = {"http": proxy_url, "https": proxy_url} if proxy_url else None

        try:
            elapsed = time.monotonic() - self._last_smarty_request_at
            wait_for = MIN_SMARTY_REQUEST_INTERVAL - elapsed
            if wait_for > 0:
                time.sleep(wait_for)
            self._last_smarty_request_at = time.monotonic()

            # Xoá cookie trước mỗi lần gọi để mỗi request trông như 1 phiên hoàn toàn mới,
            # tránh Smarty gộp nhiều request lại thành 1 "client" duy nhất qua cookie.
            res = self.session.get(
                API_URL, params=params, headers=identity["headers"],
                proxies=proxies, timeout=15
            )

            rate_limited = res.status_code == 429
            if res.status_code != 200:
                try:
                    msg = res.json()['errors'][0].get('message', 'Unknown Error')
                except Exception:
                    msg = f"HTTP {res.status_code}"
                if "too many requests" in msg.lower() or "rate limit" in msg.lower():
                    rate_limited = True
                return {
                    "result": f"Lỗi API: {msg}",
                    "analysis": {},
                    "rate_limited": rate_limited,
                    "status_code": res.status_code,
                    "retry_after": res.headers.get("Retry-After", ""),
                    "request_id": res.headers.get("X-Request-Id", ""),
                    "success": False,
                }

            data = res.json()
            if isinstance(data, list) and len(data) > 0:
                c = data[0]
                dl = c.get("delivery_line_1", "")
                ll = c.get("last_line", "")
                result_string = f"{dl} {ll}".strip() if dl or ll else "Không tìm thấy delivery_line_1 hoặc last_line"
                analysis = c.get("analysis", {}) or {}
                return {"result": result_string, "analysis": analysis, "rate_limited": False, "status_code": 200, "success": True}
            return {"result": "Không tìm thấy kết quả từ Server", "analysis": {}, "rate_limited": False, "status_code": 200, "success": False}
        except requests.exceptions.RequestException as e:
            # Lỗi mạng/timeout/kết nối: đây thường là lỗi TẠM THỜI (proxy chập chờn, mạng lag...),
            # KHÔNG phải lỗi do dữ liệu địa chỉ sai -> phải được retry giống như rate-limit,
            # nếu không sẽ bỏ sót cả những dòng đáng ra hợp lệ chỉ vì 1 lần mạng bị giật.
            return {
                "result": f"Lỗi mạng / Timeout: {str(e)[:120]}",
                "analysis": {}, "rate_limited": False, "network_error": True, "status_code": None, "success": False,
            }
        except json.JSONDecodeError:
            return {
                "result": "Lỗi phân tích JSON",
                "analysis": {}, "rate_limited": False, "network_error": True, "status_code": None, "success": False,
            }

    def call_smarty_api_with_retry(self, street_input, max_retries, row_num):
        # Chỉ trả về khi Smarty phản hồi HTTP 200 hoặc người dùng bấm Dừng. Các lỗi
        # tạm thời không được phép lọt vào collected_data và xuất hiện trong Excel.
        network_wait = 5
        attempt = 0
        retry_identity = None
        while not self.stop_requested:
            attempt += 1
            if self.stop_requested:
                return {"result": "Bị dừng", "analysis": {}, "stopped": True}

            # Mỗi LƯỢT GỌI (kể cả lần đầu, không chỉ khi bị chặn) đều dùng 1 danh tính mới:
            # proxy khác + Smarty key khác (nếu có) + User-Agent/Accept-Language khác. Việc
            # chọn key/proxy giờ LINH ĐỘNG (xem _next_identity/_pick_available): tự động né
            # những key/proxy vừa bị 429/lỗi mạng thay vì cứ lặp lại mù quáng.
            identity = retry_identity or self._next_identity()
            retry_identity = None
            api_key = identity["api_key"]
            proxy_url = identity.get("proxy_url")
            outcome = self.call_smarty_api(street_input, identity)

            is_rate_limited = outcome.get("rate_limited")
            is_network_error = outcome.get("network_error")

            if not is_rate_limited and not is_network_error:
                # Gọi được (dù thành công hay lỗi khác không liên quan rate-limit/mạng) -> key
                # và proxy này vẫn "khỏe", xóa cờ nghi ngờ để lần sau lại được ưu tiên chọn.
                self._mark_key_success(api_key)
                self._mark_proxy_success(proxy_url)
                outcome["_proxy_url"] = proxy_url
                return outcome

            key_label = self._mask_key(api_key)
            proxy_label = self._proxy_label(proxy_url)

            if is_rate_limited:
                retry_after = outcome.get("retry_after", "")
                try:
                    wait_seconds = int(retry_after)
                except (TypeError, ValueError):
                    wait_seconds = 60
                wait_seconds = min(max(wait_seconds, 5), 300)
                self.log(
                    f"  [!] Dòng {row_num}: Smarty trả về 429 với key {key_label}, {proxy_label}. "
                    f"Tạm dừng {wait_seconds}s rồi thử lại cùng proxy (lần {attempt})."
                )
                time.sleep(wait_seconds)
                retry_identity = identity
                continue

            self._mark_proxy_issue(proxy_url)
            self.log(
                f"  [!] Dòng {row_num}: Lỗi mạng/Timeout (lần {attempt}, {proxy_label}). "
                f"Tạm dừng {network_wait}s rồi tự thử lại..."
            )
            time.sleep(network_wait)
            network_wait = min(network_wait * 1.5, 60)

        return {"result": "Bị dừng", "analysis": {}, "stopped": True}

    # ---- Bảng mã tra cứu chính thức của Smarty (US Street API) ----
    DPV_FOOTNOTE_INFO = {
        "AA": ("Street/city/state/ZIP hợp lệ", False),
        "A1": ("Địa chỉ KHÔNG có trong dữ liệu USPS", True),
        "BB": ("Toàn bộ địa chỉ hợp lệ", False),
        "CC": ("Thông tin phụ không được nhận diện, KHÔNG bắt buộc", True),
        "C1": ("Thông tin phụ không được nhận diện, BẮT BUỘC", True),
        "F1": ("Địa chỉ quân sự/ngoại giao", True),
        "G1": ("Địa chỉ General Delivery", True),
        "M1": ("Thiếu số nhà", True),
        "M3": ("Số nhà không hợp lệ", True),
        "N1": ("Thiếu thông tin phụ bắt buộc", True),
        "PB": ("Địa chỉ dạng PO Box kiểu đường phố", True),
        "P1": ("Thiếu số hộp PO/RR/HC", True),
        "P3": ("Số hộp PO/RR/HC không hợp lệ", True),
        "RR": ("Có thông tin hộp thư riêng PMB", True),
        "R1": ("Không có thông tin hộp thư riêng PMB", False),
        "R7": ("Không giao hàng tận nhà", True),
        "TA": ("Khớp sau khi bỏ ký tự chữ cái cuối số nhà", True),
        "U1": ("ZIP Code dạng unique", True),
    }

    FOOTNOTE_INFO_ONLY = {"N", "Q", "Y", "Z", "LL", "LI"}
    FOOTNOTE_INFO = {
        "A": "USPS đã sửa ZIP Code",
        "B": "USPS đã sửa chính tả city/state",
        "C": "Không xác định được ZIP",
        "D": "Địa chỉ không có trong USPS",
        "E": "Nhiều bản ghi cùng ZIP",
        "F": "Không tìm thấy địa chỉ như đã nhập",
        "G": "Đã dùng dữ liệu addressee",
        "H": "Thiếu số phụ Apt/Suite",
        "I": "Dữ liệu không đủ để xác định ZIP+4",
        "J": "Có hai địa chỉ trong input",
        "K": "Đã đổi hướng N/S/E/W để khớp",
        "L": "Đã thêm/sửa/xóa thành phần địa chỉ",
        "M": "Đã sửa chính tả tên đường",
        "N": "Đã chuẩn hóa viết tắt",
        "O": "Có nhiều ZIP+4, lấy mã thấp nhất",
        "P": "Có tên địa chỉ ưu tiên hơn",
        "Q": "ZIP Code dạng unique",
        "R": "EWS báo địa chỉ sắp có dữ liệu",
        "S": "Thông tin phụ không được nhận diện",
        "T": "Magnet street syndrome",
        "U": "Tên thành phố không chính thức",
        "V": "City/state không khớp ZIP",
        "W": "ZIP không giao hàng tận đường",
        "X": "ZIP unique dùng ZIP+4 mặc định",
        "Y": "Địa chỉ quân sự",
        "Z": "Khớp qua ZIPMOVE",
        "LL": "Được gửi sang LACSLink",
        "LI": "Được gửi sang LACSLink",
    }

    def _parse_dpv_footnotes(self, raw):
        """Tách chuỗi dpv_footnotes (ghép liên tục từng cặp 2 ký tự, vd 'AABBCC') thành list mã."""
        s = str(raw or "").strip().upper()
        return [s[i:i + 2] for i in range(0, len(s) - 1, 2)] if len(s) >= 2 else ([s] if s else [])

    def _parse_footnotes(self, raw):
        """Tách chuỗi footnotes (ngăn cách bởi '#', vd 'H#L#') thành list mã chữ cái."""
        s = str(raw or "").strip().upper()
        if not s:
            return []
        return [code for code in s.split("#") if code]

    # Các mẫu regex phát hiện dữ liệu ĐẦU VÀO (từ Excel gốc) bị lỗi encode / còn dính HTML entity.
    # Đây là lớp kiểm tra ĐỘC LẬP với Smarty: dù Smarty (đặc biệt ở chế độ "enhanced" - khớp mạnh
    # tay, có thể tự "vá" qua các cụm rác không đọc hiểu được) có trả về dpv_match_code=Y hay
    # không, một input còn dính rác kiểu này vẫn cho thấy DỮ LIỆU NGUỒN có vấn đề và không nên
    # được tin tưởng tuyệt đối - phải luôn đưa vào diện cần người kiểm tra lại.
    SUSPICIOUS_INPUT_PATTERNS = [
        (re.compile(r"&#\d+;"), "chứa mã HTML numeric entity (vd \"&#39;\")"),
        (re.compile(r"&#x[0-9a-fA-F]+;"), "chứa mã HTML hex entity (vd \"&#x27;\")"),
        (re.compile(r"&(amp|lt|gt|quot|apos|nbsp);", re.IGNORECASE), "chứa mã HTML named entity (vd \"&amp;\", \"&quot;\")"),
        (re.compile(r"[\uFFFD]"), "chứa ký tự lỗi encode (replacement character \\uFFFD)"),
        (re.compile(r"<[^>]+>"), "chứa thẻ HTML còn sót lại (vd \"<br>\")"),
    ]

    # Bảng mã viết tắt 2 ký tự HỢP LỆ theo chuẩn USPS: 50 bang + DC + vùng lãnh thổ + mã quân sự.
    VALID_US_STATE_CODES = {
        "AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA", "HI", "ID", "IL", "IN", "IA",
        "KS", "KY", "LA", "ME", "MD", "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ",
        "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI", "SC", "SD", "TN", "TX", "UT", "VT",
        "VA", "WA", "WV", "WI", "WY",
        "DC",  # Washington D.C.
        "AS", "GU", "MP", "PR", "VI",  # Vùng lãnh thổ
        "AA", "AE", "AP",  # Địa chỉ quân sự (Armed Forces)
    }

    def _detect_column_format_issues(self, addr1, addr2, city, state, postal):
        """LỚP LỌC ĐỊNH DẠNG CUỐI CÙNG - chạy trên DỮ LIỆU GỐC (chưa qua Smarty, chưa qua AI)
        của 5 cột: Shipping Address1, Address2, City, State, PostalCode. Chạy NGAY TRƯỚC khi
        đưa dữ liệu qua bước kiểm tra AI (cả THỦ CÔNG lẫn TỰ ĐỘNG). Mọi lý do trả về ở đây được
        cộng dồn vào "_analysis_reasons" của dòng đó -> tự động đưa vào diện 'suspect' và AI
        KHÔNG được phép hạ xuống 'true', giống hệt cơ chế đã áp dụng cho cảnh báo DPV/input bẩn.

        Quy tắc:
        1) Address1/Address2: Address1 được phép chứa LUÔN cả nội dung kiểu Address2 (số căn
           hộ/tòa nhà) - ĐIỀU NÀY CHẤP NHẬN ĐƯỢC, không bị coi là lỗi. Chỉ bị coi là lỗi khi
           Address1/Address2 bị "dính" thêm dữ liệu vốn thuộc City/State/PostalCode (thường do
           người dùng/nguồn dữ liệu gộp nhầm nhiều cột vào 1 ô). Việc so khớp dùng ranh giới từ
           (word-boundary) và với City chỉ tính là nghi ngờ khi xuất hiện CÙNG LÚC với State
           hoặc ZIP, để tránh nhận nhầm (vd bang viết tắt "IN", "OR", "ME", "HI"... trùng với
           từ tiếng Anh thông thường nếu chỉ so khớp đơn lẻ).
        2) City: không được trống, không chứa chữ số, không chứa dấu phẩy/chấm phẩy, không dài
           bất thường - những dấu hiệu cho thấy ô này bị dính thêm dữ liệu khác.
        3) State: không được trống, phải khớp đúng 1 trong các mã viết tắt 2 ký tự chuẩn USPS.
        4) PostalCode: CHỈ kiểm tra không được trống. KHÔNG so khớp với postal code mà Smarty
           trả về, vì input gốc lệch so với Smarty vẫn có thể ĐÚNG (Smarty tự khớp/chuẩn hoá)."""
        reasons = []

        # --- (2) Shipping City ---
        if not city:
            reasons.append("[Lọc định dạng] Shipping City đang TRỐNG")
        else:
            if re.search(r"\d", city):
                reasons.append(f"[Lọc định dạng] Shipping City chứa chữ số bất thường ('{city}') - nghi dính dữ liệu khác (vd mã ZIP)")
            if re.search(r"[,;]", city):
                reasons.append(f"[Lọc định dạng] Shipping City chứa dấu phẩy/chấm phẩy ('{city}') - nghi bị gộp nhầm City/State/PostalCode vào cùng 1 ô")
            if len(city) > 40:
                reasons.append(f"[Lọc định dạng] Shipping City dài bất thường ({len(city)} ký tự) - nghi bị dính thêm dữ liệu khác")

        # --- (3) Shipping State ---
        if not state:
            reasons.append("[Lọc định dạng] Shipping State đang TRỐNG")
        elif state.upper() not in self.VALID_US_STATE_CODES:
            reasons.append(f"[Lọc định dạng] Shipping State ('{state}') không đúng định dạng viết tắt 2 ký tự chuẩn USPS")

        # --- (4) Shipping PostalCode: CHỈ kiểm tra không trống ---
        if not postal:
            reasons.append("[Lọc định dạng] Shipping PostalCode đang TRỐNG")

        # --- (1) Shipping Address1 / Address2 ---
        if not addr1 and not addr2:
            reasons.append("[Lọc định dạng] Shipping Address1 và Address2 đều đang TRỐNG - không có số nhà/tên đường")
        else:
            addr_combo = f"{addr1} {addr2}"

            zip5_match = re.search(r"\d{5}", postal) if postal else None
            zip5 = zip5_match.group(0) if zip5_match else ""
            has_zip_leak = bool(zip5) and re.search(rf"\b{re.escape(zip5)}\b", addr_combo)

            has_state_leak = bool(state) and re.search(rf"\b{re.escape(state)}\b", addr_combo, re.IGNORECASE)
            has_city_leak = bool(city) and len(city) >= 3 and re.search(rf"\b{re.escape(city)}\b", addr_combo, re.IGNORECASE)

            if has_zip_leak:
                reasons.append(f"[Lọc định dạng] Shipping Address1/Address2 chứa mã ZIP ('{zip5}') trùng Shipping PostalCode - nghi dính nhầm dữ liệu cột khác")
            if has_city_leak and (has_state_leak or has_zip_leak):
                reasons.append(f"[Lọc định dạng] Shipping Address1/Address2 chứa tên thành phố ('{city}') CÙNG với State/ZIP - nghi bị gộp nhầm Shipping City/State/PostalCode vào Address1/Address2")

        return reasons

    def _detect_input_anomalies(self, input_string):
        """Quét CHUỖI ĐẦU VÀO GỐC (chưa qua Smarty) để tìm dấu hiệu lỗi encode/HTML entity còn sót
        lại từ nguồn dữ liệu (Excel/CRM/website export...). Trả về danh sách lý do (rỗng nếu input
        sạch). Đây là kiểm tra độc lập với 'analysis' của Smarty - áp dụng NGAY CẢ KHI Smarty (đặc
        biệt match=enhanced) vẫn khớp được địa chỉ, vì bản thân input bẩn là 1 rủi ro về chất lượng
        dữ liệu, không nên tự động coi là 'true' chỉ vì Smarty đã cố gắng vá được."""
        reasons = []
        if not input_string:
            return reasons
        for pattern, desc in self.SUSPICIOUS_INPUT_PATTERNS:
            if pattern.search(input_string):
                reasons.append(f"Input gốc {desc} - dữ liệu nguồn có thể bị lỗi encode, cần kiểm tra lại thủ công")
        return reasons

    def _compute_analysis_flags(self, analysis):
        """Đối chiếu TOÀN BỘ object 'analysis' Smarty trả về (không chỉ vacant/CMRA/no_stat/
        active/match_code, mà còn cả 'dpv_footnotes' VÀ 'footnotes' - trường trước đây bị bỏ sót).
        Nguyên tắc KHẮT KHE: hễ còn BẤT KỲ dòng cảnh báo nào (kể cả các mã như 'L#' - đổi suffix/
        directional - vốn không ảnh hưởng dpv_match_code) thì vẫn bị đánh dấu 'suspect', không được
        coi là hợp lệ tuyệt đối chỉ vì dpv_match_code=Y. Trả về danh sách lý do (rỗng nếu thực sự
        sạch, không có bất kỳ cảnh báo nào)."""
        reasons = []
        if not analysis:
            return reasons

        vacant = str(analysis.get("dpv_vacant", "")).strip().upper()
        cmra = str(analysis.get("dpv_cmra", "")).strip().upper()
        no_stat = str(analysis.get("dpv_no_stat", "")).strip().upper()
        active = str(analysis.get("active", "")).strip().upper()
        match_code = str(analysis.get("dpv_match_code", "")).strip().upper()

        if vacant == "Y":
            reasons.append("USPS đánh dấu địa chỉ đang BỊ BỎ TRỐNG (dpv_vacant=Y)")
        if cmra == "Y":
            reasons.append("Địa chỉ là hộp thư nhận hộ CMRA, không phải địa chỉ thực (dpv_cmra=Y)")
        if no_stat == "Y":
            reasons.append("USPS không có đủ dữ liệu xác thực cho địa chỉ này (dpv_no_stat=Y)")
        if active == "N":
            reasons.append("Địa chỉ hiện KHÔNG hoạt động (active=N)")
        if match_code and match_code != "Y":
            reasons.append(f"Mã khớp DPV chưa xác nhận đầy đủ (dpv_match_code={match_code})")

        # --- MỚI: kiểm tra dpv_footnotes theo từng mã 2 ký tự ---
        for code in self._parse_dpv_footnotes(analysis.get("dpv_footnotes", "")):
            info = self.DPV_FOOTNOTE_INFO.get(code)
            if info is None:
                # Mã lạ / chưa có trong bảng tra -> vẫn đưa vào diện nghi ngờ để con người xem lại,
                # thay vì mặc định coi là an toàn.
                reasons.append(f"Mã dpv_footnotes lạ, cần xem lại thủ công ({code})")
                continue
            desc, should_block = info
            if should_block:
                reasons.append(f"Cảnh báo dpv_footnotes: {desc} ({code})")

        # --- MỚI (quan trọng nhất, khắc phục đúng lỗi bạn báo): kiểm tra 'footnotes' ---
        # Đây là trường trước đây KHÔNG được kiểm tra, nên các cảnh báo như 'L#'
        # (đổi suffix/directional để khớp) đã lọt qua dù dpv_match_code=Y.
        for code in self._parse_footnotes(analysis.get("footnotes", "")):
            if code in self.FOOTNOTE_INFO_ONLY:
                continue  # Mã thuần thông tin (N#, Q#, Y#, Z#, LL#, LI#) - không chặn
            desc = self.FOOTNOTE_INFO.get(code, "Mã footnotes lạ, cần xem lại thủ công")
            reasons.append(f"Cảnh báo footnotes: {desc} ({code}#)")

        return reasons

    def reset_ui(self):
        self.btn_start.config(state=tk.NORMAL, text="2. Bắt đầu xử lý & Xuất file (TXT + Excel)")
        self.btn_select_excel.config(state=tk.NORMAL)
        self.btn_stop.config(state=tk.DISABLED)
        self.btn_clear_cache.config(state=tk.NORMAL)
        self.entry_delay.config(state=tk.NORMAL)
        self.rb_match_strict.config(state=tk.NORMAL)
        self.rb_match_enhanced.config(state=tk.NORMAL)
        self.chk_use_ai.config(state=tk.NORMAL)
        if self.use_ai_var.get():
            self._set_ai_inputs_state(tk.NORMAL)

if __name__ == "__main__":
    root = tk.Tk()
    app = SmartyApp(root)
    root.mainloop()