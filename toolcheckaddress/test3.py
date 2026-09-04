import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import requests
import json
import re
import pandas as pd
import threading
import time
import os
import random
import tempfile
import concurrent.futures
import webbrowser
from openpyxl.styles import PatternFill, Font

try:
    from google import genai
    from google.genai import types as genai_types
    HAS_GEMINI = True
except ImportError:
    HAS_GEMINI = False

API_URL = "https://us-street.api.smarty.com/street-address"
DEFAULT_SMARTY_KEY = "21102174564513388"

USER_AGENT_POOL = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Edg/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
]

ACCEPT_LANGUAGE_POOL = [
    "en-US,en;q=0.9",
    "en-US,en;q=0.8,vi;q=0.5",
    "en-GB,en;q=0.9",
    "en-US,en;q=0.9,fr;q=0.6",
    "en-CA,en;q=0.9",
]

GROQ_MODELS = ["llama-3.3-70b-versatile", "llama-3.1-8b-instant"]
OPENROUTER_MODELS = [
    "meta-llama/llama-3.1-8b-instruct:free",
    "google/gemini-2.0-flash-exp:free",
    "mistralai/mistral-7b-instruct:free",
]
GEMINI_FALLBACK_MODELS = ['gemini-flash-latest', 'gemini-2.5-flash', 'gemini-2.5-flash-lite', 'gemini-2.0-flash']

SESSION_STATE_FILE = "smarty_last_session.json"

class SmartyApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Smarty API & Google Gemini Auto Check")
        self.root.geometry("820x780")

        self.excel_path = ""
        self.output_dir = ""
        self.stop_requested = False

        self.session = requests.Session()
        self.gemini_client = None
        self.proxy_list = self._load_proxies()
        self.api_key_list = self._load_api_keys()
        self._identity_counter = 0

        self.key_status = {}
        self.proxy_status = {}
        self._key_rr_counter = 0
        self._proxy_rr_counter = 0
        self._warned_single_key_session = False

        self.session_state_dir = os.getcwd()

        self.setup_gui()
        self._refresh_recheck_button()

    def _load_proxies(self):
        if os.path.exists("proxies.txt"):
            try:
                with open("proxies.txt", "r", encoding="utf-8") as f:
                    return [line.strip() for line in f if line.strip()]
            except Exception:
                return []
        return []

    def _load_api_keys(self):
        if os.path.exists("smarty_keys.txt"):
            try:
                with open("smarty_keys.txt", "r", encoding="utf-8") as f:
                    keys = [line.strip() for line in f if line.strip()]
                if keys:
                    return keys
            except Exception:
                pass
        return [DEFAULT_SMARTY_KEY]

    def _pick_available(self, values, status_dict, counter_attr):
        if not values:
            return None
        now = time.time()
        available = [v for v in values if status_dict.get(v, {}).get("blocked_until", 0) <= now]
        counter = getattr(self, counter_attr)
        if available:
            chosen = available[counter % len(available)]
            setattr(self, counter_attr, counter + 1)
            return chosen
        return min(values, key=lambda v: status_dict.get(v, {}).get("blocked_until", 0))

    def _mark_key_rate_limited(self, key):
        st = self.key_status.setdefault(key, {"blocked_until": 0.0, "fail_streak": 0})
        st["fail_streak"] += 1
        cooldown = min(5 * (2 ** (st["fail_streak"] - 1)), 300)
        st["blocked_until"] = time.time() + cooldown
        return cooldown, st["fail_streak"]

    def _mark_key_success(self, key):
        st = self.key_status.setdefault(key, {"blocked_until": 0.0, "fail_streak": 0})
        st["fail_streak"] = 0
        st["blocked_until"] = 0.0

    def _mark_proxy_issue(self, proxy_url):
        if not proxy_url:
            return 0
        st = self.proxy_status.setdefault(proxy_url, {"blocked_until": 0.0, "fail_streak": 0})
        st["fail_streak"] += 1
        cooldown = min(3 * (2 ** (st["fail_streak"] - 1)), 120)
        st["blocked_until"] = time.time() + cooldown
        return cooldown

    def _mark_proxy_success(self, proxy_url):
        if not proxy_url:
            return
        st = self.proxy_status.setdefault(proxy_url, {"blocked_until": 0.0, "fail_streak": 0})
        st["fail_streak"] = 0
        st["blocked_until"] = 0.0

    def _mask_key(self, key):
        key = str(key or "")
        return f"...{key[-4:]}" if len(key) > 4 else "*" * len(key)

    def _proxy_label(self, proxy_url):
        if not proxy_url:
            return "không dùng proxy"
        try:
            return f"proxy #{self.proxy_list.index(proxy_url)}"
        except ValueError:
            return "proxy (?)"

    def _next_identity(self):
        i = self._identity_counter
        self._identity_counter += 1

        proxy_url = self._pick_available(self.proxy_list, self.proxy_status, "_proxy_rr_counter")
        api_key = self._pick_available(self.api_key_list, self.key_status, "_key_rr_counter")
        user_agent = USER_AGENT_POOL[i % len(USER_AGENT_POOL)]
        accept_language = ACCEPT_LANGUAGE_POOL[i % len(ACCEPT_LANGUAGE_POOL)]

        headers = {
            "Referer": "https://www.smarty.com/",
            "Origin": "https://www.smarty.com",
            "User-Agent": user_agent,
            "Accept-Language": accept_language,
            "Accept": "*/*",
            "Accept-Encoding": "gzip, deflate, br, zstd",
            "Sec-Ch-Ua": '"Not=A?Brand";v="99", "Opera";v="135", "Chromium";v="151"',
            "Sec-Ch-Ua-Mobile": "?0",
            "Sec-Ch-Ua-Platform": '"Windows"',
            "Sec-Fetch-Dest": "empty",
            "Sec-Fetch-Mode": "cors",
            "Sec-Fetch-Site": "same-site",
        }
        return {
            "proxy_url": proxy_url,
            "api_key": api_key,
            "headers": headers,
            "index": i,
        }

    def _warmup_proxies(self):
        if not self.proxy_list:
            return

        self.log(f"Đang khởi động {len(self.proxy_list)} proxy trước khi xử lý...")

        def _check(proxy_url):
            try:
                requests.get(
                    "https://www.smarty.com/",
                    proxies={"http": proxy_url, "https": proxy_url},
                    timeout=8,
                )
                return True
            except requests.exceptions.RequestException:
                return False

        results = []
        with concurrent.futures.ThreadPoolExecutor(max_workers=min(10, len(self.proxy_list))) as executor:
            futures = [executor.submit(_check, p) for p in self.proxy_list]
            for fut in concurrent.futures.as_completed(futures, timeout=20):
                try:
                    results.append(fut.result())
                except Exception:
                    results.append(False)

        ok_count = sum(1 for r in results if r)
        self.log(f"Đã khởi động xong proxy: {ok_count}/{len(self.proxy_list)} phản hồi tốt.")

    def setup_gui(self):
        frame_top = tk.Frame(self.root)
        frame_top.pack(pady=(15, 5), padx=10, fill="x")

        self.btn_select_excel = tk.Button(frame_top, text="1. Chọn file Excel gốc", command=self.select_excel, width=20, bg="#28a745", fg="white", font=("Arial", 10, "bold"))
        self.btn_select_excel.pack(side="left", padx=(0, 10))

        self.lbl_excel_path = tk.Label(frame_top, text="Chưa chọn file...", fg="gray")
        self.lbl_excel_path.pack(side="left")

        frame_delay = tk.Frame(self.root)
        frame_delay.pack(pady=5, padx=10, fill="x")

        lbl_delay = tk.Label(frame_delay, text="Độ trễ API Smarty (giây):", font=("Arial", 10, "bold"))
        lbl_delay.pack(side="left")

        self.delay_var = tk.StringVar(value="12.0")
        self.entry_delay = tk.Entry(frame_delay, textvariable=self.delay_var, width=8, font=("Consolas", 11), justify="center")
        self.entry_delay.pack(side="left", padx=10)

        frame_match = tk.LabelFrame(self.root, text=" Chế độ khớp địa chỉ Smarty (match) ", font=("Arial", 10, "bold"), fg="#0078D7")
        frame_match.pack(pady=10, padx=10, fill="x")

        self.match_mode_var = tk.StringVar(value="strict")

        self.rb_match_strict = tk.Radiobutton(
            frame_match, variable=self.match_mode_var, value="strict",
            text="Strict (Khuyến khích - mặc định): chỉ khớp dựa hoàn toàn trên dữ liệu gốc",
            font=("Arial", 9, "bold"), justify="left", fg="#1da462"
        )
        self.rb_match_strict.pack(anchor="w", padx=10, pady=(5, 0))

        self.rb_match_enhanced = tk.Radiobutton(
            frame_match, variable=self.match_mode_var, value="enhanced",
            text="Enhanced: Smarty tự khớp mạnh tay hơn",
            font=("Arial", 9), justify="left"
        )
        self.rb_match_enhanced.pack(anchor="w", padx=10, pady=(0, 2))

        self.lbl_match_note = tk.Label(
            frame_match,
            text="* Khuyến khích dùng Strict để kết quả trung thực với dữ liệu gốc.",
            fg="gray", font=("Arial", 8), justify="left", wraplength=760
        )
        self.lbl_match_note.pack(anchor="w", padx=10, pady=(0, 5))

        frame_ai = tk.LabelFrame(self.root, text=" Kiểm tra kết quả đáng ngờ bằng AI (Miễn phí) ", font=("Arial", 10, "bold"), fg="#1da462")
        frame_ai.pack(pady=10, padx=10, fill="x")

        self.use_ai_var = tk.BooleanVar(value=True)
        self.chk_use_ai = tk.Checkbutton(frame_ai, text="Bật tự động kiểm tra trạng thái bằng AI", variable=self.use_ai_var, font=("Arial", 10), command=self.toggle_ai_input)
        self.chk_use_ai.pack(anchor="w", padx=10, pady=5)

        frame_mode = tk.Frame(frame_ai)
        frame_mode.pack(anchor="w", padx=20, pady=(0, 5), fill="x")

        self.ai_mode_var = tk.StringVar(value="manual")
        self.rb_mode_auto = tk.Radiobutton(
            frame_mode, variable=self.ai_mode_var, value="auto",
            text="Tự động: hệ thống tự dò & gọi AI miễn phí",
            font=("Arial", 9), justify="left", command=self.toggle_ai_input
        )
        self.rb_mode_auto.pack(anchor="w")

        self.rb_mode_manual = tk.Radiobutton(
            frame_mode, variable=self.ai_mode_var, value="manual",
            text="Thủ công: tạo Prompt để dán vào AI ngoài",
            font=("Arial", 9), justify="left", command=self.toggle_ai_input
        )
        self.rb_mode_manual.pack(anchor="w")

        frame_ai_keys = tk.Frame(frame_ai)
        frame_ai_keys.pack(fill="x", padx=20, pady=5)

        lbl_gemini = tk.Label(frame_ai_keys, text="Gemini API Key:", width=18, anchor="w")
        lbl_gemini.grid(row=0, column=0, sticky="w", pady=2)
        self.entry_ai_key = tk.Entry(frame_ai_keys, width=45, font=("Consolas", 10), show="*")
        self.entry_ai_key.grid(row=0, column=1, padx=5, pady=2)

        lbl_groq = tk.Label(frame_ai_keys, text="Groq API Key (tuỳ chọn):", width=18, anchor="w")
        lbl_groq.grid(row=1, column=0, sticky="w", pady=2)
        self.entry_groq_key = tk.Entry(frame_ai_keys, width=45, font=("Consolas", 10), show="*")
        self.entry_groq_key.grid(row=1, column=1, padx=5, pady=2)

        lbl_openrouter = tk.Label(frame_ai_keys, text="OpenRouter API Key (tuỳ chọn):", width=18, anchor="w")
        lbl_openrouter.grid(row=2, column=0, sticky="w", pady=2)
        self.entry_openrouter_key = tk.Entry(frame_ai_keys, width=45, font=("Consolas", 10), show="*")
        self.entry_openrouter_key.grid(row=2, column=1, padx=5, pady=2)

        self.lbl_ai_note = tk.Label(
            frame_ai,
            text="* Chế độ Thủ công không cần API Key. Chế độ Tự động cần ít nhất 1 trong 3 API Key trên.",
            fg="gray", font=("Arial", 8)
        )
        self.lbl_ai_note.pack(anchor="w", padx=20, pady=(0, 5))

        self.lbl_analysis_note = tk.Label(
            frame_ai,
            text=("* Dù BẬT hay TẮT AI, tool luôn tự kiểm tra thêm phần analysis (DPV) của Smarty."),
            fg="#1da462", font=("Arial", 8), justify="left", wraplength=760
        )
        self.lbl_analysis_note.pack(anchor="w", padx=20, pady=(0, 5))

        self.toggle_ai_input()

        if not HAS_GEMINI:
            lbl_gemini.config(text="Gemini (chưa cài 'google-genai'):", fg="red")

        frame_btns = tk.Frame(self.root)
        frame_btns.pack(pady=10)

        self.btn_start = tk.Button(frame_btns, text="2. Bắt đầu xử lý & Xuất file (TXT + Excel)", command=self.start_processing_thread, width=35, bg="#0078D7", fg="white", font=("Arial", 11, "bold"), state=tk.DISABLED)
        self.btn_start.pack(side="left", padx=5)

        self.btn_stop = tk.Button(frame_btns, text="Dừng lại", command=self.stop_processing, width=10, bg="#dc3545", fg="white", font=("Arial", 11, "bold"), state=tk.DISABLED)
        self.btn_stop.pack(side="left", padx=5)

        self.frame_recheck = tk.Frame(self.root)
        self.frame_recheck.pack(pady=(0, 5), padx=10, fill="x")

        self.btn_recheck = tk.Button(
            self.frame_recheck, text="Kiểm tra lại (0 dòng)",
            command=self.start_recheck_session, width=28,
            bg="#6f42c1", fg="white", font=("Arial", 10, "bold")
        )

        self.lbl_recheck_info = tk.Label(self.frame_recheck, text="", fg="gray", font=("Arial", 8), justify="left")

        lbl_log = tk.Label(self.root, text="Tiến trình xử lý:", font=("Arial", 10, "bold"))
        lbl_log.pack(anchor="w", padx=10)

        self.log_text = scrolledtext.ScrolledText(self.root, font=("Consolas", 10), width=95, height=18, bg="#1E1E1E", fg="#D4D4D4")
        self.log_text.pack(pady=5, padx=10)

    def _set_ai_inputs_state(self, state):
        self.entry_ai_key.config(state=state)
        self.entry_groq_key.config(state=state)
        self.entry_openrouter_key.config(state=state)
        self.rb_mode_auto.config(state=state)
        self.rb_mode_manual.config(state=state)

    def toggle_ai_input(self):
        if self.use_ai_var.get():
            self._set_ai_inputs_state(tk.NORMAL)
        else:
            self._set_ai_inputs_state(tk.DISABLED)

    def log(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def select_excel(self):
        file_path = filedialog.askopenfilename(title="Chọn file Excel", filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_path:
            self.excel_path = file_path
            self.lbl_excel_path.config(text=self.excel_path, fg="black")
            self.btn_start.config(state=tk.NORMAL)
            self.log(f"Đã chọn file: {self.excel_path}")
            if self.proxy_list:
                self.log(f"Đã nạp {len(self.proxy_list)} proxy từ file proxies.txt.")
            else:
                self.log("Không có proxies.txt -> chạy bằng IP máy.")
            if len(self.api_key_list) > 1:
                self.log(f"Đã nạp {len(self.api_key_list)} Smarty Key từ file smarty_keys.txt.")
            else:
                self.log("Chỉ đang dùng 1 API Key Smarty duy nhất.")
            self.log(f"Đang xoay vòng {len(USER_AGENT_POOL)} User-Agent.")

    def _session_file_path(self):
        return os.path.join(self.session_state_dir, SESSION_STATE_FILE)

    def _save_session_to_disk(self, collected_data, source_excel_path="", output_excel_path=""):
        try:
            payload = {
                "saved_at": time.strftime("%Y-%m-%d %H:%M:%S"),
                "row_count": len(collected_data),
                "source_excel_path": source_excel_path,
                "output_excel_path": output_excel_path,
                "collected_data": collected_data,
            }
            with open(self._session_file_path(), "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False)
        except Exception as e:
            self.log(f"    [!] Không lưu được phiên làm việc: {e}")

    def _load_session_from_disk(self):
        path = self._session_file_path()
        if not os.path.exists(path):
            return None
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, dict) or not isinstance(data.get("collected_data"), list):
                return None
            if not data["collected_data"]:
                return None
            return data
        except Exception:
            return None

    def _refresh_recheck_button(self):
        session = self._load_session_from_disk()
        if session:
            n = session.get("row_count", len(session.get("collected_data", [])))
            self.btn_recheck.config(text=f"Kiểm tra lại ({n} dòng)", state=tk.NORMAL)
            if not self.btn_recheck.winfo_ismapped():
                self.btn_recheck.pack(side="left", padx=(0, 10))
            info_bits = [f"Phiên gần nhất lưu lúc {session.get('saved_at', '?')}"]
            src = session.get("source_excel_path", "")
            if src:
                info_bits.append(f"nguồn: {os.path.basename(src)}")
            self.lbl_recheck_info.config(text=" - ".join(info_bits))
            if not self.lbl_recheck_info.winfo_ismapped():
                self.lbl_recheck_info.pack(side="left")
        else:
            self.btn_recheck.pack_forget()
            self.lbl_recheck_info.pack_forget()

    def start_recheck_session(self):
        session = self._load_session_from_disk()
        if not session:
            messagebox.showinfo("Không có phiên cũ", "Không tìm thấy phiên làm việc cũ nào.")
            self._refresh_recheck_button()
            return

        collected_data = session["collected_data"]
        n = len(collected_data)

        proceed = messagebox.askyesno(
            "Kiểm tra lại phiên cũ",
            f"Phiên cũ có {n} dòng. Tool sẽ không gọi lại Smarty.\nBạn có muốn tiếp tục không?"
        )
        if not proceed:
            return

        save_path = filedialog.asksaveasfilename(
            title="Lưu kết quả kiểm tra lại",
            defaultextension=".xlsx",
            initialfile="KetQua_Smarty_KiemTraLai",
        )
        if not save_path:
            return
        if not save_path.lower().endswith(".xlsx"):
            save_path += ".xlsx"

        self.btn_recheck.config(state=tk.DISABLED)
        self.btn_start.config(state=tk.DISABLED)
        self.btn_select_excel.config(state=tk.DISABLED)

        threading.Thread(
            target=self._recheck_session_worker,
            args=(collected_data, save_path, session.get("source_excel_path", "")),
            daemon=True,
        ).start()

    def _recheck_session_worker(self, collected_data, save_path, source_excel_path):
        try:
            self.log("\n===========================================")
            self.log(f"[KIỂM TRA LẠI] Đang mở lại {len(collected_data)} dòng để hỏi AI lại...")

            self.run_manual_flow(collected_data)

            self.log("\nĐang tạo lại file Excel...")
            final_path = self.export_excel(collected_data, save_path)
            self.log(f"-> Đã lưu Excel: {final_path}")

            self._save_session_to_disk(collected_data, source_excel_path=source_excel_path, output_excel_path=final_path)

            self.log("\n[THÀNH CÔNG] Đã kiểm tra lại và xuất file Excel mới!")
            messagebox.showinfo("Hoàn tất", f"Đã xuất lại thành công:\n{final_path}")
        except Exception as e:
            self.log(f"\n[LỖI] Kiểm tra lại thất bại: {e}")
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi kiểm tra lại:\n{e}")
        finally:
            self.root.after(0, self._reset_recheck_ui)

    def _reset_recheck_ui(self):
        self.btn_start.config(state=tk.NORMAL if self.excel_path else tk.DISABLED)
        self.btn_select_excel.config(state=tk.NORMAL)
        self._refresh_recheck_button()

    def stop_processing(self):
        self.stop_requested = True
        self.log("\n[HỆ THỐNG] Đang yêu cầu dừng tiến trình...")
        self.btn_stop.config(state=tk.DISABLED)

    def start_processing_thread(self):
        try:
            delay_val = float(self.delay_var.get())
            if delay_val < 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning("Cảnh báo", "Độ trễ không hợp lệ!")
            return

        if self.use_ai_var.get() and self.ai_mode_var.get() == "auto":
            keys = [self.entry_ai_key.get().strip(), self.entry_groq_key.get().strip(), self.entry_openrouter_key.get().strip()]
            if not any(keys):
                messagebox.showwarning("Cảnh báo", "Chế độ Tự động cần ít nhất 1 API Key!")
                return

        save_path = filedialog.asksaveasfilename(
            title="Lưu kết quả",
            defaultextension="",
            initialfile="KetQua_Smarty",
        )
        if not save_path:
            return

        base_path = save_path.replace('.txt', '').replace('.xlsx', '')
        self.txt_output_path = base_path + ".txt"
        self.excel_output_path = base_path + ".xlsx"

        self.stop_requested = False
        self.btn_start.config(state=tk.DISABLED, text="Đang xử lý...")
        self.btn_select_excel.config(state=tk.DISABLED)
        self.btn_stop.config(state=tk.NORMAL)
        self.entry_delay.config(state=tk.DISABLED)
        self.rb_match_strict.config(state=tk.DISABLED)
        self.rb_match_enhanced.config(state=tk.DISABLED)
        self.chk_use_ai.config(state=tk.DISABLED)
        self._set_ai_inputs_state(tk.DISABLED)

        threading.Thread(target=self.process_data, daemon=True).start()

    def process_data(self):
        collected_data = []

        self.key_status = {}
        self.proxy_status = {}
        self._key_rr_counter = 0
        self._proxy_rr_counter = 0
        self._warned_single_key_session = False

        try:
            delay_seconds = float(self.delay_var.get())
            self.log("Đang đọc dữ liệu từ file Excel...")
            df = pd.read_excel(self.excel_path)

            address_cols = ['Shipping Address1', 'Shipping Address2', 'Shipping City', 'Shipping State', 'Shipping PostalCode', 'Shipping Country']
            ref_col = 'Shipment Reference Id'
            required_cols = address_cols + [ref_col]

            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                messagebox.showerror("Lỗi dữ liệu", f"File thiếu các cột:\n{', '.join(missing_cols)}")
                self.reset_ui()
                return

            total_rows = len(df)
            match_mode = self.match_mode_var.get()
            self.log(f"Chế độ khớp Smarty: {match_mode.upper()}")

            self._warmup_proxies()

            self.log(f"Bắt đầu giai đoạn 1: Gọi API Smarty ({total_rows} dòng)...")

            f_handle = self._open_output_txt_with_retry(self.txt_output_path)
            self.txt_output_path = f_handle.name
            with f_handle as f_out:
                for index, row in df.iterrows():
                    if self.stop_requested:
                        self.log("\n[HỆ THỐNG] Tiến trình dừng bởi người dùng.")
                        break

                    ref_id = str(row[ref_col]).strip() if pd.notna(row[ref_col]) else "NO_REF"
                    parts = [str(row[col]).strip() for col in address_cols if pd.notna(row[col]) and str(row[col]).strip() != ""]
                    input_string = " ".join(parts)

                    addr1_raw = str(row['Shipping Address1']).strip() if pd.notna(row['Shipping Address1']) else ""
                    addr2_raw = str(row['Shipping Address2']).strip() if pd.notna(row['Shipping Address2']) else ""
                    city_raw = str(row['Shipping City']).strip() if pd.notna(row['Shipping City']) else ""
                    state_raw = str(row['Shipping State']).strip() if pd.notna(row['Shipping State']) else ""
                    postal_raw = str(row['Shipping PostalCode']).strip() if pd.notna(row['Shipping PostalCode']) else ""

                    if not input_string:
                        continue

                    outcome = self.call_smarty_api_with_retry(input_string, 10, index + 1)
                    result_string = outcome.get("result", "")
                    analysis = outcome.get("analysis", {}) or {}

                    dpv_match_code = analysis.get("dpv_match_code", "")
                    dpv_vacant = analysis.get("dpv_vacant", "")
                    dpv_cmra = analysis.get("dpv_cmra", "")
                    dpv_no_stat = analysis.get("dpv_no_stat", "")
                    active = analysis.get("active", "")
                    dpv_footnotes_raw = analysis.get("dpv_footnotes", "")
                    footnotes_raw = analysis.get("footnotes", "")

                    analysis_reasons = self._compute_analysis_flags(analysis)
                    input_anomaly_reasons = self._detect_input_anomalies(input_string)
                    format_reasons = self._detect_column_format_issues(addr1_raw, addr2_raw, city_raw, state_raw, postal_raw)
                    all_reasons = analysis_reasons + input_anomaly_reasons + format_reasons

                    analysis_summary = "; ".join(analysis_reasons) if analysis_reasons else (
                        "Không có cảnh báo" if analysis else "Không có dữ liệu analysis"
                    )
                    if input_anomaly_reasons:
                        analysis_summary += " | " + "; ".join(input_anomaly_reasons)
                    if format_reasons:
                        analysis_summary += " | " + "; ".join(format_reasons)

                    f_out.write(f"({input_string} : {result_string} | Analysis: {analysis_summary} [{ref_id}])\n")
                    f_out.flush()

                    item = {
                        "Reference Id": ref_id,
                        "Chuỗi đầu vào": input_string,
                        "Chuỗi đầu ra": result_string,
                        "DPV Match Code": dpv_match_code,
                        "DPV Vacant": dpv_vacant,
                        "DPV CMRA": dpv_cmra,
                        "DPV No Stat": dpv_no_stat,
                        "Active": active,
                        "DPV Footnotes": dpv_footnotes_raw,
                        "Footnotes": footnotes_raw,
                        "Trạng thái xử lý": "suspect" if all_reasons else "",
                        "Lý do đáng ngờ": ("Cảnh báo: " + "; ".join(all_reasons)) if all_reasons else "",
                    }
                    item["_analysis_reasons"] = all_reasons
                    collected_data.append(item)

                    if outcome.get("success"):
                        self.log(f"Smarty API - Dòng {index + 1}/{total_rows} -> OK")
                    else:
                        self.log(f"Smarty API - Dòng {index + 1}/{total_rows} -> LỖI: {result_string}")

                    if delay_seconds > 0:
                        time.sleep(delay_seconds)
                        if index % 30 == 0 and index > 0:
                            self.log("Đang tạm nghỉ 60 giây sau mỗi 30 dòng để tránh bị giới hạn...")
                            time.sleep(60)

                if not self.stop_requested:
                    gpt_prompt = """
=======================================================================================================================
[SYSTEM PROMPT - DÀNH CHO AI/CHATGPT/GEMINI]
Nhiệm vụ của bạn là đóng vai trò chuyên gia kiểm duyệt địa chỉ quốc tế khắt khe.
1. QUÉT TOÀN BỘ danh sách và LỌC RA những đơn hàng CÓ DẤU HIỆU ĐÁNG NGỜ hoặc LỖI.
2. Tiêu chí bắt LỖI: Không tìm thấy kết quả từ Server, Lỗi HTTP, Lỗi API, thiếu delivery_line_1 hoặc last_line.
3. Tiêu chí bắt NGHI NGỜ: Phần Analysis có cảnh báo vacant, CMRA, no_stat, active=N, dpv_match_code khác Y.
4. ĐẦU RA BẮT BUỘC TRÌNH BÀY DƯỚI DẠNG BẢNG Markdown.
=======================================================================================================================
"""
                    f_out.write(gpt_prompt)
                    f_out.flush()

            if self.use_ai_var.get() and not self.stop_requested and collected_data:
                self.log("\n===========================================")
                self.log("Bắt đầu giai đoạn 2: Kiểm tra kết quả đáng ngờ bằng AI...")

                mode = self.ai_mode_var.get()
                if mode == "manual":
                    self.run_manual_flow(collected_data)
                else:
                    self.run_auto_flow(collected_data)
            elif not self.use_ai_var.get():
                self.log("\n[INFO] AI tự động đang TẮT.")

            if collected_data:
                self.log("\nBắt đầu tạo file Excel...")
                self.excel_output_path = self.export_excel(collected_data, self.excel_output_path)
                self.log(f"-> Đã lưu Excel: {self.excel_output_path}")

                self._save_session_to_disk(
                    collected_data,
                    source_excel_path=self.excel_path,
                    output_excel_path=self.excel_output_path,
                )
                self.root.after(0, self._refresh_recheck_button)

            if not self.stop_requested:
                self.log(f"\n[THÀNH CÔNG] Toàn bộ tiến trình hoàn tất!")
                messagebox.showinfo("Hoàn tất", f"Đã xuất thành công:\n1. {self.txt_output_path}\n2. {self.excel_output_path}")

        except Exception as e:
            self.log(f"\n[LỖI] Đã xảy ra sự cố: {e}")
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra:\n{e}")
        finally:
            self.reset_ui()

    def run_auto_flow(self, collected_data):
        BATCH_SIZE = 20
        n = len(collected_data)
        self.log(f"Tổng {n} dòng -> chia thành các batch {BATCH_SIZE} dòng/lượt gọi AI.")

        idx = 0
        while idx < n:
            if self.stop_requested:
                return
            batch = collected_data[idx: idx + BATCH_SIZE]
            batch_no = idx // BATCH_SIZE + 1
            total_batches = (n + BATCH_SIZE - 1) // BATCH_SIZE
            self.log(f"-> Đang thử gọi AI tự động cho batch {batch_no}/{total_batches} ({len(batch)} dòng)...")

            results_map = self.check_with_ai_batch_auto(batch)

            if results_map is None:
                self.log("    [!] Không có nhà cung cấp AI miễn phí nào phản hồi thành công.")
                remaining_items = collected_data[idx:]
                want_manual = self.ask_switch_to_manual(
                    f"Hệ thống không gọi được AI tự động.\nCòn {len(remaining_items)} dòng chưa kiểm tra."
                )
                if want_manual:
                    self.run_manual_flow(remaining_items)
                else:
                    for item in remaining_items:
                        self._apply_ai_result_with_analysis(item, None, missing_reason="Bỏ qua kiểm tra AI")
                return

            for item in batch:
                res = results_map.get(item["Reference Id"])
                self._apply_ai_result_with_analysis(item, res)

            self.log(f"-> Batch {batch_no}/{total_batches} hoàn tất.")
            idx += BATCH_SIZE
            if idx < n:
                time.sleep(1.0)

    def check_with_ai_batch_auto(self, items):
        prompt = self._build_ai_prompt(items)
        raw = self.call_any_available_ai(prompt)
        if raw is None:
            return None
        return self._parse_ai_json_array(raw)

    def call_any_available_ai(self, prompt):
        gemini_key = self.entry_ai_key.get().strip()
        groq_key = self.entry_groq_key.get().strip()
        openrouter_key = self.entry_openrouter_key.get().strip()

        providers = []
        if gemini_key and HAS_GEMINI:
            providers.append(("Gemini", lambda: self._call_gemini_fast(prompt, gemini_key)))
        if groq_key:
            providers.append(("Groq", lambda: self._call_groq(prompt, groq_key)))
        if openrouter_key:
            providers.append(("OpenRouter", lambda: self._call_openrouter(prompt, openrouter_key)))

        if not providers:
            self.log("    [!] Chưa nhập bất kỳ API Key nào cho chế độ Tự động.")
            return None

        for name, fn in providers:
            if self.stop_requested:
                return None
            self.log(f"    -> Đang thử {name}...")
            try:
                raw = fn()
                if raw and raw.strip():
                    self.log(f"    [OK] {name} phản hồi thành công.")
                    return raw
                self.log(f"    [!] {name} trả về rỗng.")
            except Exception as e:
                self.log(f"    [!] {name} lỗi: {str(e)[:120]}")
                continue

        return None

    def _call_gemini_fast(self, prompt, api_key):
        client = genai.Client(api_key=api_key)
        models = self._quick_discover_gemini_models(client) or GEMINI_FALLBACK_MODELS
        last_err = None

        for model_name in models[:2]:
            for attempt in range(2):
                if self.stop_requested:
                    return None
                try:
                    response = client.models.generate_content(
                        model=model_name,
                        contents=prompt,
                        config=genai_types.GenerateContentConfig(
                            temperature=0.0,
                            response_mime_type="application/json"
                        )
                    )
                    return (response.text or "").strip()
                except Exception as e:
                    last_err = e
                    es = str(e).lower()
                    if "404" in es or "not found" in es:
                        break
                    if "429" in es or "503" in es or "overloaded" in es or "unavailable" in es or "quota" in es:
                        if attempt == 0:
                            time.sleep(3)
                            continue
                        break
                    break

        if last_err:
            raise last_err
        raise RuntimeError("Gemini không khả dụng")

    def _quick_discover_gemini_models(self, client):
        try:
            discovered = []
            for m in client.models.list():
                name = getattr(m, "name", "") or ""
                short = name.split("/")[-1] if "/" in name else name
                if not short:
                    continue
                actions = getattr(m, "supported_actions", None) or []
                if actions and "generateContent" not in actions:
                    continue
                low = short.lower()
                if any(x in low for x in ["embedding", "tts", "image", "video", "vision", "aqa", "gemma"]):
                    continue
                if "flash" not in low and "pro" not in low:
                    continue
                discovered.append(short)

            def priority(name):
                n = name.lower()
                if "flash-lite" in n:
                    return 1
                if n.endswith("-latest") and "flash" in n:
                    return 0
                if "flash" in n:
                    return 2
                if "pro" in n:
                    return 3
                return 4

            return sorted(set(discovered), key=priority)
        except Exception:
            return []

    def _call_groq(self, prompt, api_key):
        url = "https://api.groq.com/openai/v1/chat/completions"
        headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
        last_err = None
        for model_name in GROQ_MODELS[:2]:
            body = {"model": model_name, "messages": [{"role": "user", "content": prompt}], "temperature": 0}
            try:
                r = self.session.post(url, headers=headers, json=body, timeout=20)
                if r.status_code == 200:
                    return r.json()["choices"][0]["message"]["content"]
                elif r.status_code == 404:
                    last_err = RuntimeError(f"Model '{model_name}' không tồn tại trên Groq")
                    continue
                elif r.status_code == 429:
                    last_err = RuntimeError("Groq quá tải/hết quota (429)")
                    time.sleep(2)
                    continue
                else:
                    last_err = RuntimeError(f"Groq lỗi HTTP {r.status_code}: {r.text[:150]}")
                    continue
            except requests.exceptions.RequestException as e:
                last_err = e
                continue
        raise last_err or RuntimeError("Groq thất bại")

    def _call_openrouter(self, prompt, api_key):
        url = "https://openrouter.ai/api/v1/chat/completions"
        headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
        last_err = None
        for model_name in OPENROUTER_MODELS:
            body = {"model": model_name, "messages": [{"role": "user", "content": prompt}], "temperature": 0}
            try:
                r = self.session.post(url, headers=headers, json=body, timeout=20)
                if r.status_code == 200:
                    return r.json()["choices"][0]["message"]["content"]
                elif r.status_code == 404:
                    last_err = RuntimeError(f"Model '{model_name}' không tồn tại trên OpenRouter")
                    continue
                elif r.status_code == 429:
                    last_err = RuntimeError("OpenRouter quá tải/hết quota (429)")
                    time.sleep(2)
                    continue
                else:
                    last_err = RuntimeError(f"OpenRouter lỗi HTTP {r.status_code}: {r.text[:150]}")
                    continue
            except requests.exceptions.RequestException as e:
                last_err = e
                continue
        raise last_err or RuntimeError("OpenRouter thất bại")

    def run_manual_flow(self, items):
        if not items:
            return
        self.log(f"[THỦ CÔNG] Đang tạo Prompt cho {len(items)} dòng để bạn sao chép...")
        prompt = self._build_ai_prompt(items)

        json_text = self.get_manual_json_from_user(prompt, len(items))

        if not json_text:
            self.log("    [!] Bạn đã bỏ qua bước nhập JSON thủ công.")
            for item in items:
                self._apply_ai_result_with_analysis(item, None, missing_reason="Bỏ qua kiểm tra AI thủ công")
            return

        results_map = self._parse_ai_json_array(json_text)
        if not results_map:
            self.log("    [!] Không đọc được JSON hợp lệ từ nội dung bạn dán vào.")

        for item in items:
            res = results_map.get(item["Reference Id"])
            if res is None:
                self._apply_ai_result_with_analysis(item, None, missing_reason="Không thấy id này trong JSON đã nhập")
            else:
                self._apply_ai_result_with_analysis(item, res)

        self.log(f"    -> Đã áp dụng kết quả thủ công cho {len(items)} dòng.")

    def _apply_ai_result_with_analysis(self, item, res, missing_reason="Không nhận được kết quả từ AI cho dòng này"):
        analysis_reasons = item.get("_analysis_reasons", [])
        analysis_text = "; ".join(analysis_reasons)

        if res is None:
            if analysis_reasons:
                base_reason = item.get("Lý do đáng ngờ", "") or f"Cảnh báo Analysis: {analysis_text}"
                item["Trạng thái xử lý"] = "suspect"
                item["Lý do đáng ngờ"] = f"{base_reason} | {missing_reason}"
            else:
                item["Trạng thái xử lý"] = "unknown"
                item["Lý do đáng ngờ"] = missing_reason
            return

        ai_status = res.get("status", "unknown")
        ai_reason = (res.get("reason", "") or "").strip()

        if not analysis_reasons:
            item["Trạng thái xử lý"] = ai_status
            item["Lý do đáng ngờ"] = ai_reason
            return

        final_status = "false" if ai_status == "false" else "suspect"

        reason_parts = []
        if ai_reason:
            reason_parts.append(f"AI: {ai_reason}")
        reason_parts.append(f"Cảnh báo Analysis: {analysis_text}")
        item["Trạng thái xử lý"] = final_status
        item["Lý do đáng ngờ"] = " | ".join(reason_parts)

    def _ask_retry_cancel(self, title, message):
        result_holder = {"value": False}
        event = threading.Event()

        def _build():
            try:
                result_holder["value"] = messagebox.askretrycancel(title, message)
            finally:
                event.set()

        self.root.after(0, _build)
        event.wait()
        return result_holder["value"]

    def _alt_path_with_timestamp(self, path):
        base, ext = os.path.splitext(path)
        return f"{base}_{time.strftime('%Y%m%d_%H%M%S')}{ext}"

    def _open_output_txt_with_retry(self, path):
        current = path
        while True:
            try:
                return open(current, 'w', encoding='utf-8')
            except PermissionError:
                self.log(f"  [!] Không thể ghi file TXT '{os.path.basename(current)}'.")
                retry = self._ask_retry_cancel(
                    "Không thể ghi file TXT",
                    f"File:\n{current}\n\nBấm 'Retry' sau khi đóng file. Bấm 'Cancel' để lưu tên khác."
                )
                if retry:
                    continue
                current = self._alt_path_with_timestamp(path)
                self.log(f"  [i] Sẽ lưu file TXT sang tên khác: {current}")

    def _save_workbook_with_retry(self, wb, desired_path):
        path = desired_path
        while True:
            try:
                wb.save(path)
                return path
            except PermissionError:
                self.log(f"  [!] Không thể lưu Excel '{os.path.basename(path)}'.")
                retry = self._ask_retry_cancel(
                    "Không thể lưu file Excel",
                    f"File:\n{path}\n\nBấm 'Retry' sau khi đóng file. Bấm 'Cancel' để lưu tên khác."
                )
                if retry:
                    continue
                path = self._alt_path_with_timestamp(desired_path)
                self.log(f"  [i] Sẽ lưu file Excel sang tên khác: {path}")
            except OSError as e:
                self.log(f"  [!] Lỗi khi lưu file Excel: {e}")
                retry = self._ask_retry_cancel(
                    "Không thể lưu file Excel",
                    f"Đã xảy ra lỗi khi lưu file:\n{path}\n\nLỗi: {e}"
                )
                if retry:
                    continue
                path = self._alt_path_with_timestamp(desired_path)
                self.log(f"  [i] Sẽ lưu file Excel sang tên khác: {path}")

    def ask_switch_to_manual(self, reason_text):
        result_holder = {"value": False}
        event = threading.Event()

        def _build():
            try:
                ans = messagebox.askyesno(
                    "Không gọi được AI tự động",
                    f"{reason_text}\n\nBạn có muốn chuyển sang chế độ THỦ CÔNG không?"
                )
                result_holder["value"] = bool(ans)
            finally:
                event.set()

        self.root.after(0, _build)
        event.wait()
        return result_holder["value"]

    def get_manual_json_from_user(self, prompt_text, item_count):
        result_holder = {"value": None}
        event = threading.Event()

        def _build():
            dlg = tk.Toplevel(self.root)
            dlg.title("Chế độ thủ công - Dán Prompt vào AI ngoài")
            dlg.geometry("720x650")
            dlg.grab_set()

            try:
                webbrowser.open("https://chatgpt.com/")
            except Exception:
                pass

            tk.Label(
                dlg,
                text=(f"Có {item_count} dòng cần kiểm tra.\n"
                      "1) Dán Prompt vào AI ngoài.\n"
                      "2) Sao chép JSON AI trả về dán vào ô dưới.\n"
                      "3) Bấm 'Xác nhận'."),
                justify="left", anchor="w"
            ).pack(anchor="w", padx=10, pady=(10, 5))

            tk.Label(dlg, text="Prompt để dán vào AI ngoài:", font=("Arial", 9, "bold")).pack(anchor="w", padx=10)
            txt_prompt = scrolledtext.ScrolledText(dlg, height=14, wrap="word", font=("Consolas", 9))
            txt_prompt.insert("1.0", prompt_text)
            txt_prompt.pack(fill="both", expand=True, padx=10, pady=(0, 5))

            def copy_prompt():
                dlg.clipboard_clear()
                dlg.clipboard_append(prompt_text)
                lbl_copied.config(text="Đã sao chép vào Clipboard!", fg="#28a745")

            frame_copy = tk.Frame(dlg)
            frame_copy.pack(pady=2)
            tk.Button(frame_copy, text="Sao chép Prompt", command=copy_prompt, bg="#0078D7", fg="white", width=20).pack(side="left", padx=5)
            lbl_copied = tk.Label(frame_copy, text="", font=("Arial", 9))
            lbl_copied.pack(side="left", padx=5)

            copy_prompt()

            tk.Label(dlg, text="Dán JSON kết quả từ AI vào đây:", font=("Arial", 9, "bold")).pack(anchor="w", padx=10, pady=(10, 0))
            txt_json = scrolledtext.ScrolledText(dlg, height=10, wrap="word", font=("Consolas", 9))
            txt_json.pack(fill="both", expand=True, padx=10, pady=(0, 5))

            def on_confirm():
                content = txt_json.get("1.0", "end").strip()
                result_holder["value"] = content if content else None
                dlg.grab_release()
                dlg.destroy()
                event.set()

            def on_skip():
                result_holder["value"] = None
                dlg.grab_release()
                dlg.destroy()
                event.set()

            frame_btn = tk.Frame(dlg)
            frame_btn.pack(pady=10)
            tk.Button(frame_btn, text="Xác nhận", command=on_confirm, bg="#28a745", fg="white", width=14, font=("Arial", 10, "bold")).pack(side="left", padx=5)
            tk.Button(frame_btn, text="Bỏ qua kiểm tra AI", command=on_skip, bg="#dc3545", fg="white", width=32).pack(side="left", padx=5)

            dlg.protocol("WM_DELETE_WINDOW", on_skip)

        self.root.after(0, _build)
        event.wait()
        return result_holder["value"]

    def _build_ai_prompt(self, items):
        payload = [
            {
                "id": it["Reference Id"],
                "in": it["Chuỗi đầu vào"],
                "out": it["Chuỗi đầu ra"],
                "analysis": {
                    "dpv_match_code": it.get("DPV Match Code", ""),
                    "dpv_vacant": it.get("DPV Vacant", ""),
                    "dpv_cmra": it.get("DPV CMRA", ""),
                    "dpv_no_stat": it.get("DPV No Stat", ""),
                    "active": it.get("Active", ""),
                    "dpv_footnotes": it.get("DPV Footnotes", ""),
                    "footnotes": it.get("Footnotes", ""),
                },
            }
            for it in items
        ]

        prompt = f"""Bạn là chuyên gia kiểm duyệt địa chỉ vận chuyển quốc tế.
Dưới đây là danh sách JSON gồm nhiều đơn hàng.

Với MỖI phần tử, gán MỘT trong 3 trạng thái:
- "true": chỉ khi không có cảnh báo analysis và chuỗi khớp tốt.
- "suspect": khi có cảnh báo analysis hoặc chuỗi khớp nhưng cần xem lại.
- "false": khi lỗi hoặc sai lệch nghiêm trọng.

CHỈ trả về DUY NHẤT một mảng JSON hợp lệ, không markdown, không chữ thừa.

Danh sách đầu vào:
{json.dumps(payload, ensure_ascii=False)}
"""
        return prompt

    def _parse_ai_json_array(self, raw_text):
        text = raw_text.strip()
        text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text.strip())
        try:
            data = json.loads(text)
        except json.JSONDecodeError:
            match = re.search(r"\[.*\]", text, re.DOTALL)
            if not match:
                self.log("    [!] Không parse được JSON trả về từ AI.")
                return {}
            try:
                data = json.loads(match.group(0))
            except json.JSONDecodeError:
                self.log("    [!] JSON trả về từ AI bị sai định dạng.")
                return {}

        result_map = {}
        if isinstance(data, list):
            for entry in data:
                if not isinstance(entry, dict) or "id" not in entry:
                    continue
                status = str(entry.get("status", "")).strip().lower()
                status = status if status in ("true", "false", "suspect") else "unknown"
                result_map[str(entry["id"])] = {
                    "status": status,
                    "reason": entry.get("reason", "") or ""
                }
        return result_map

    def export_excel(self, collected_data, output_path):
        for item in collected_data:
            item.setdefault("DPV Match Code", "")
            item.setdefault("DPV Vacant", "")
            item.setdefault("DPV CMRA", "")
            item.setdefault("DPV No Stat", "")
            item.setdefault("Active", "")
            item.setdefault("DPV Footnotes", "")
            item.setdefault("Footnotes", "")
            item.setdefault("Trạng thái xử lý", "")
            item.setdefault("Lý do đáng ngờ", "")

        status_priority = {"false": 0, "suspect": 1, "unknown": 2, "": 2, "true": 3}
        sorted_data = sorted(
            collected_data,
            key=lambda x: status_priority.get(str(x.get("Trạng thái xử lý", "")).lower(), 2)
        )

        columns = [
            "Reference Id", "Chuỗi đầu vào", "Chuỗi đầu ra",
            "DPV Match Code", "DPV Vacant", "DPV CMRA", "DPV No Stat", "Active",
            "DPV Footnotes", "Footnotes",
            "Trạng thái xử lý", "Lý do đáng ngờ",
        ]
        out_df = pd.DataFrame(sorted_data, columns=columns)

        tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(tmp_fd)
        try:
            out_df.to_excel(tmp_path, index=False)

            import openpyxl
            wb = openpyxl.load_workbook(tmp_path)
            ws = wb.active

            fill_false = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            font_false = Font(color="9C0006", bold=True)
            fill_suspect = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            font_suspect = Font(color="C65911", bold=True)
            fill_unknown = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            font_unknown = Font(color="9C6500")
            fill_true = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            font_true = Font(color="006100")

            status_col_idx = columns.index("Trạng thái xử lý") + 1

            for row_idx in range(2, ws.max_row + 1):
                status_val = str(ws.cell(row=row_idx, column=status_col_idx).value or "").strip().lower()
                if status_val == "false":
                    fill, font = fill_false, font_false
                elif status_val == "suspect":
                    fill, font = fill_suspect, font_suspect
                elif status_val == "true":
                    fill, font = fill_true, font_true
                else:
                    fill, font = fill_unknown, font_unknown

                for col_idx in range(1, len(columns) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = fill
                    if col_idx == status_col_idx:
                        cell.font = font

            for col_idx, col_name in enumerate(columns, start=1):
                max_len = max([len(str(col_name))] + [len(str(row.get(col_name, ""))) for row in sorted_data])
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 60)

            ws.freeze_panes = "A2"

            final_path = self._save_workbook_with_retry(wb, output_path)
            return final_path
        finally:
            try:
                os.remove(tmp_path)
            except OSError:
                pass

    def call_smarty_api(self, street_input, identity):
        match_mode = self.match_mode_var.get() if self.match_mode_var.get() in ("strict", "enhanced") else "strict"
        params = {
            "key": identity["api_key"],
            "agent": "smarty (website:demo/single-address@latest)",
            "match": match_mode,
            "candidates": "5",
            "geocode": "true",
            "license": "us-core-cloud",
            "street": street_input,
        }

        proxy_url = identity.get("proxy_url")
        proxies = {"http": proxy_url, "https": proxy_url} if proxy_url else None

        try:
            if os.path.exists("cookie.txt"):
                try:
                    with open("cookie.txt", "r", encoding="utf-8") as cf:
                        cookie_value = cf.read().strip()
                    if cookie_value:
                        identity["headers"]["Cookie"] = cookie_value
                except Exception:
                    pass

            res = self.session.get(
                API_URL, params=params, headers=identity["headers"],
                proxies=proxies, timeout=15
            )

            rate_limited = res.status_code == 429
            if res.status_code != 200:
                try:
                    msg = res.json()['errors'][0].get('message', 'Unknown Error')
                except Exception:
                    msg = f"HTTP {res.status_code}"
                if "too many requests" in msg.lower() or "rate limit" in msg.lower():
                    rate_limited = True
                return {
                    "result": f"Lỗi API: {msg}",
                    "analysis": {},
                    "rate_limited": rate_limited,
                    "status_code": res.status_code,
                    "success": False,
                }

            data = res.json()
            if isinstance(data, list) and len(data) > 0:
                c = data[0]
                dl = c.get("delivery_line_1", "")
                ll = c.get("last_line", "")
                result_string = f"{dl} {ll}".strip() if dl or ll else "Không tìm thấy delivery_line_1 hoặc last_line"
                analysis = c.get("analysis", {}) or {}
                return {"result": result_string, "analysis": analysis, "rate_limited": False, "status_code": 200, "success": True}
            return {"result": "Không tìm thấy kết quả từ Server", "analysis": {}, "rate_limited": False, "status_code": 200, "success": False}
        except requests.exceptions.RequestException as e:
            return {
                "result": f"Lỗi mạng / Timeout: {str(e)[:120]}",
                "analysis": {}, "rate_limited": False, "network_error": True, "status_code": None, "success": False,
            }
        except json.JSONDecodeError:
            return {
                "result": "Lỗi phân tích JSON",
                "analysis": {}, "rate_limited": False, "network_error": True, "status_code": None, "success": False,
            }

    def call_smarty_api_with_retry(self, street_input, max_retries, row_num):
        base_wait = 1.5
        only_one_key = len(self.api_key_list) == 1
        single_key_fail_streak_at_row_start = self.key_status.get(self.api_key_list[0], {}).get("fail_streak", 0) if only_one_key else 0

        for attempt in range(max_retries):
            if self.stop_requested:
                return {"result": "Bị dừng", "analysis": {}}

            identity = self._next_identity()
            api_key = identity["api_key"]
            proxy_url = identity.get("proxy_url")
            outcome = self.call_smarty_api(street_input, identity)

            is_rate_limited = outcome.get("rate_limited")
            is_network_error = outcome.get("network_error")

            if not is_rate_limited and not is_network_error:
                self._mark_key_success(api_key)
                self._mark_proxy_success(proxy_url)
                return outcome

            key_label = self._mask_key(api_key)
            proxy_label = self._proxy_label(proxy_url)

            if is_rate_limited:
                cooldown, fail_streak = self._mark_key_rate_limited(api_key)
                self.log(
                    f"  [!] Dòng {row_num}: Bị giới hạn request (lần {attempt + 1}/{max_retries}, "
                    f"key {key_label}, {proxy_label}). Tạm khóa key ~{cooldown:.0f}s."
                )

                if only_one_key:
                    fail_streak_this_row = fail_streak - single_key_fail_streak_at_row_start
                    if fail_streak_this_row >= 3 and not self._warned_single_key_session:
                        self._warned_single_key_session = True
                        self.log(
                            "  [CẢNH BÁO] Chỉ có 1 API Key và key này bị 429 liên tục. "
                            "Hãy đợi lâu hơn hoặc bổ sung smarty_keys.txt."
                        )
            else:
                cooldown = self._mark_proxy_issue(proxy_url)
                self.log(
                    f"  [!] Dòng {row_num}: Lỗi mạng/Timeout (lần {attempt + 1}/{max_retries}, {proxy_label}). "
                    f"Tạm khóa proxy ~{cooldown:.0f}s."
                )

            jitter = random.uniform(0.3, 1.3)
            if self.proxy_list or len(self.api_key_list) > 1:
                time.sleep(jitter)
            else:
                sleep_for = base_wait + jitter
                self.log(f"      Không có proxy/key khác, chờ {sleep_for:.1f}s...")
                time.sleep(sleep_for)
                base_wait = min(base_wait * 1.8, 30)

        hint = (
            " (Rất có thể API Key đã hết quota. Hãy bổ sung smarty_keys.txt.)"
            if only_one_key else ""
        )
        return {
            "result": f"Lỗi: Quá nhiều request (đã xoay vòng vẫn bị giới hạn){hint}",
            "analysis": {},
            "success": False,
        }

    DPV_FOOTNOTE_INFO = {
        "AA": ("Street/city/state/ZIP hợp lệ", False),
        "A1": ("Địa chỉ KHÔNG có trong dữ liệu USPS", True),
        "BB": ("Toàn bộ địa chỉ hợp lệ", False),
        "CC": ("Thông tin phụ không được nhận diện", True),
        "C1": ("Thông tin phụ không được nhận diện và bắt buộc", True),
        "F1": ("Địa chỉ quân sự/ngoại giao", True),
        "G1": ("Địa chỉ General Delivery", True),
        "M1": ("Thiếu số nhà", True),
        "M3": ("Số nhà không hợp lệ", True),
        "N1": ("Thiếu thông tin phụ bắt buộc", True),
        "PB": ("Địa chỉ PO Box kiểu đường phố", True),
        "P1": ("Thiếu số hộp PO/RR/HC", True),
        "P3": ("Số hộp PO/RR/HC không hợp lệ", True),
        "RR": ("Có thông tin hộp thư riêng PMB", True),
        "R1": ("Không có thông tin hộp thư riêng PMB", False),
        "R7": ("Không được USPS giao tận nhà", True),
        "TA": ("Số nhà khớp sau khi bỏ chữ cái cuối", True),
        "U1": ("ZIP Code dạng unique", True),
    }

    FOOTNOTE_INFO_ONLY = {"N", "Q", "Y", "Z", "LL", "LI"}
    FOOTNOTE_INFO = {
        "A": "USPS đã sửa ZIP Code",
        "B": "USPS đã sửa chính tả thành phố/tiểu bang",
        "C": "Không xác định được ZIP",
        "D": "Địa chỉ không có trong dữ liệu USPS",
        "E": "Nhiều bản ghi cùng ZIP",
        "F": "Không tìm thấy địa chỉ trong thành phố/ZIP đã cho",
        "G": "Dùng dữ liệu Addressee để ghép",
        "H": "Thiếu số phụ",
        "I": "Dữ liệu không đủ để xác định ZIP+4",
        "J": "Địa chỉ bị trùng",
        "K": "Khớp sau khi đổi hướng",
        "L": "Suffix hoặc directional bị thêm/sửa/xóa",
        "M": "Đã sửa chính tả tên đường",
        "N": "Đã chuẩn hoá viết tắt",
        "O": "Nhiều ZIP+4 phù hợp",
        "P": "Có tên khác được ưu tiên hơn",
        "Q": "ZIP Code unique",
        "R": "EWS báo sắp có dữ liệu",
        "S": "Thông tin phụ không được nhận diện",
        "T": "Magnet street syndrome",
        "U": "Tên thành phố không chính thức",
        "V": "Không xác thực city/state khớp ZIP",
        "W": "ZIP không giao tận đường",
        "X": "ZIP unique dùng ZIP+4 mặc định",
        "Y": "Địa chỉ quân sự",
        "Z": "Khớp qua ZIPMOVE",
        "LL": "LACSLink",
        "LI": "LACSLink",
    }

    def _parse_dpv_footnotes(self, raw):
        s = str(raw or "").strip().upper()
        return [s[i:i + 2] for i in range(0, len(s) - 1, 2)] if len(s) >= 2 else ([s] if s else [])

    def _parse_footnotes(self, raw):
        s = str(raw or "").strip().upper()
        if not s:
            return []
        return [code for code in s.split("#") if code]

    SUSPICIOUS_INPUT_PATTERNS = [
        (re.compile(r"&#\d+;"), "chứa mã HTML numeric entity"),
        (re.compile(r"&#x[0-9a-fA-F]+;"), "chứa mã HTML hex entity"),
        (re.compile(r"&(amp|lt|gt|quot|apos|nbsp);", re.IGNORECASE), "chứa mã HTML named entity"),
        (re.compile(r"[\uFFFD]"), "chứa ký tự lỗi encode"),
        (re.compile(r"<[^>]+>"), "chứa thẻ HTML còn sót"),
    ]

    VALID_US_STATE_CODES = {
        "AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA", "HI", "ID", "IL", "IN", "IA",
        "KS", "KY", "LA", "ME", "MD", "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ",
        "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI", "SC", "SD", "TN", "TX", "UT", "VT",
        "VA", "WA", "WV", "WI", "WY", "DC", "AS", "GU", "MP", "PR", "VI", "AA", "AE", "AP",
    }

    def _detect_column_format_issues(self, addr1, addr2, city, state, postal):
        reasons = []

        if not city:
            reasons.append("[Lọc định dạng] Shipping City đang TRỐNG")
        else:
            if re.search(r"\d", city):
                reasons.append(f"[Lọc định dạng] Shipping City chứa chữ số bất thường ('{city}')")
            if re.search(r"[,;]", city):
                reasons.append(f"[Lọc định dạng] Shipping City chứa dấu phẩy/chấm phẩy ('{city}')")
            if len(city) > 40:
                reasons.append(f"[Lọc định dạng] Shipping City dài bất thường ({len(city)} ký tự)")

        if not state:
            reasons.append("[Lọc định dạng] Shipping State đang TRỐNG")
        elif state.upper() not in self.VALID_US_STATE_CODES:
            reasons.append(f"[Lọc định dạng] Shipping State ('{state}') không đúng chuẩn USPS")

        if not postal:
            reasons.append("[Lọc định dạng] Shipping PostalCode đang TRỐNG")

        if not addr1 and not addr2:
            reasons.append("[Lọc định dạng] Shipping Address1 và Address2 đều TRỐNG")
        else:
            addr_combo = f"{addr1} {addr2}"

            zip5_match = re.search(r"\d{5}", postal) if postal else None
            zip5 = zip5_match.group(0) if zip5_match else ""
            has_zip_leak = bool(zip5) and re.search(rf"\b{re.escape(zip5)}\b", addr_combo)

            has_state_leak = bool(state) and re.search(rf"\b{re.escape(state)}\b", addr_combo, re.IGNORECASE)
            has_city_leak = bool(city) and len(city) >= 3 and re.search(rf"\b{re.escape(city)}\b", addr_combo, re.IGNORECASE)

            if has_zip_leak:
                reasons.append(f"[Lọc định dạng] Address chứa mã ZIP ('{zip5}') trùng PostalCode")
            if has_city_leak and (has_state_leak or has_zip_leak):
                reasons.append(f"[Lọc định dạng] Address chứa tên thành phố cùng State/ZIP")

        return reasons

    def _detect_input_anomalies(self, input_string):
        reasons = []
        if not input_string:
            return reasons
        for pattern, desc in self.SUSPICIOUS_INPUT_PATTERNS:
            if pattern.search(input_string):
                reasons.append(f"Input gốc {desc}")
        return reasons

    def _compute_analysis_flags(self, analysis):
        reasons = []
        if not analysis:
            return reasons

        vacant = str(analysis.get("dpv_vacant", "")).strip().upper()
        cmra = str(analysis.get("dpv_cmra", "")).strip().upper()
        no_stat = str(analysis.get("dpv_no_stat", "")).strip().upper()
        active = str(analysis.get("active", "")).strip().upper()
        match_code = str(analysis.get("dpv_match_code", "")).strip().upper()

        if vacant == "Y":
            reasons.append("USPS đánh dấu địa chỉ BỊ BỎ TRỐNG (dpv_vacant=Y)")
        if cmra == "Y":
            reasons.append("Địa chỉ là CMRA (dpv_cmra=Y)")
        if no_stat == "Y":
            reasons.append("USPS không đủ dữ liệu xác thực (dpv_no_stat=Y)")
        if active == "N":
            reasons.append("Địa chỉ KHÔNG hoạt động (active=N)")
        if match_code and match_code != "Y":
            reasons.append(f"Mã DPV chưa xác nhận đầy đủ (dpv_match_code={match_code})")

        for code in self._parse_dpv_footnotes(analysis.get("dpv_footnotes", "")):
            info = self.DPV_FOOTNOTE_INFO.get(code)
            if info is None:
                reasons.append(f"Mã dpv_footnotes lạ ({code})")
                continue
            desc, should_block = info
            if should_block:
                reasons.append(f"Cảnh báo dpv_footnotes: {desc} ({code})")

        for code in self._parse_footnotes(analysis.get("footnotes", "")):
            if code in self.FOOTNOTE_INFO_ONLY:
                continue
            desc = self.FOOTNOTE_INFO.get(code, "Mã footnotes lạ")
            reasons.append(f"Cảnh báo footnotes: {desc} ({code}#)")

        return reasons

    def reset_ui(self):
        self.btn_start.config(state=tk.NORMAL, text="2. Bắt đầu xử lý & Xuất file (TXT + Excel)")
        self.btn_select_excel.config(state=tk.NORMAL)
        self.btn_stop.config(state=tk.DISABLED)
        self.entry_delay.config(state=tk.NORMAL)
        self.rb_match_strict.config(state=tk.NORMAL)
        self.rb_match_enhanced.config(state=tk.NORMAL)
        self.chk_use_ai.config(state=tk.NORMAL)
        if self.use_ai_var.get():
            self._set_ai_inputs_state(tk.NORMAL)

if __name__ == "__main__":
    root = tk.Tk()
    app = SmartyApp(root)
    root.mainloop()